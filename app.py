"""
博昱科技智能助手后端服务
- Flask + SQLite + OpenAI兼容API（支持DeepSeek/OpenAI/小米）
- Function Calling 实现 skill 查询逻辑
"""

import json
import os
import sqlite3
import sys
import threading
import time
from datetime import datetime, timedelta
from collections import defaultdict
from flask import Flask, request, jsonify, send_from_directory, send_file
from openai import OpenAI

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(BASE_DIR, 'data', 'supply_chain.db')

app = Flask(__name__, static_folder='static', static_url_path='')

# 禁用静态文件缓存（开发阶段，确保前端修改即时生效）
@app.after_request
def add_no_cache(response):
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    return response

# ============================================================
#  地区分类规则（与 skill 一致）
# ============================================================
REGION_US = {"US", "CA", "FA"}
REGION_UK = {"GB", "UK", "IE"}

REGION_ALIAS = {
    "美国": "美国地区", "us": "美国地区", "美国地区": "美国地区",
    "英国": "英国地区", "uk": "英国地区", "英国地区": "英国地区",
    "欧洲": "欧洲地区", "eu": "欧洲地区", "欧洲地区": "欧洲地区",
    "德国": "欧洲地区", "de": "欧洲地区",
    "法国": "欧洲地区", "fr": "欧洲地区",
    "意大利": "欧洲地区", "it": "欧洲地区",
    "西班牙": "欧洲地区", "es": "欧洲地区",
}

REGION_TO_MAP = {"美国地区": "US", "英国地区": "UK", "欧洲地区": "DE"}

# 表分类
TABLE_CATEGORIES = {
    '美国基础运费': '价卡数据',
    '美国超规费用': '价卡数据',
    '美国其他费用': '价卡数据',
    '德国基础运费': '价卡数据',
    '德国超规费用': '价卡数据',
    '德国其他费用': '价卡数据',
    '英国基础运费': '价卡数据',
    '英国超规费用': '价卡数据',
    '英国其他费用': '价卡数据',
    '产品信息': '产品数据',
    'SKU映射': '产品数据',
    '负责人': '产品数据',
    '工厂库存': '供应链数据',
    '海外仓库存': '供应链数据',
    '在途货物': '供应链数据',
    '销量': '供应链数据',
    '批次库存表': '供应链数据',
    '海运周期': '供应链数据',
}


# ============================================================
#  SQLite 查询
# ============================================================

def db_connect():
    conn = sqlite3.connect(DB_PATH, timeout=30, check_same_thread=False)
    conn.row_factory = sqlite3.Row
    # 启用 WAL 模式提升并发性能，避免读写锁冲突
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA busy_timeout=5000")
    conn.execute("PRAGMA synchronous=NORMAL")
    conn.execute("PRAGMA cache_size=-8000")  # 8MB 缓存
    return conn


def parse_sku(warehouse_sku: str) -> str:
    if not warehouse_sku:
        return ""
    if "-" in warehouse_sku:
        return warehouse_sku.split("-")[0]
    return warehouse_sku


def classify_region(country_code: str) -> str:
    if not country_code:
        return "未知"
    code = country_code.strip().upper()
    if code in REGION_US:
        return "美国地区"
    elif code in REGION_UK:
        return "英国地区"
    else:
        return "欧洲地区"


def resolve_sku(parsed_sku: str, country_code: str) -> str:
    """根据国家代码查找SKU映射"""
    try:
        conn = db_connect()
        region_name = classify_region(country_code)
        map_region = REGION_TO_MAP.get(region_name)
        if map_region:
            c = conn.cursor()
            c.execute('SELECT 匹配后SKU FROM SKU映射 WHERE 地区=? AND 原始SKU=?',
                       (map_region, parsed_sku))
            row = c.fetchone()
            if row:
                conn.close()
                return row[0]
        conn.close()
    except Exception:
        pass
    return parsed_sku


def resolve_division(mapped_sku: str, country_code: str) -> str:
    """根据SKU和国家代码查找事业部"""
    try:
        conn = db_connect()
        region_name = classify_region(country_code)
        sheet = REGION_TO_MAP.get(region_name, "DE")
        c = conn.cursor()
        c.execute('SELECT 所属事业部 FROM 产品信息 WHERE 地区=? AND SKU=?',
                   (sheet, mapped_sku))
        row = c.fetchone()
        conn.close()
        if row and row[0]:
            return row[0]
    except Exception:
        pass
    return ""


# ============================================================
#  Function Calling 工具定义
# ============================================================

TOOLS = [
    {
        "type": "function",
        "function": {
            "name": "query_sales",
            "description": "查询销量数据。支持按SKU、店铺、时间范围、地区、事业部筛选。时间范围：1)相对天数用days参数(如近7天)；2)绝对日期用start_date+end_date(如四月→start_date='2026-04-01',end_date='2026-04-30')。用户说某月时必须用start_date+end_date。",
            "parameters": {
                "type": "object",
                "properties": {
                    "sku": {"type": "string", "description": "要查询的SKU编码（如CR04800101MP）"},
                    "store": {"type": "string", "description": "按店铺账号筛选"},
                    "days": {"type": "integer", "description": "近N天（如7表示近7天），以数据最大付款时间为终点。注意：只有用户说'近N天'时才用此参数，说某月时用start_date+end_date"},
                    "start_date": {"type": "string", "description": "开始日期，格式YYYY-MM-DD（如2026-04-01）。用户说某月时用此参数"},
                    "end_date": {"type": "string", "description": "结束日期，格式YYYY-MM-DD（如2026-04-30）。用户说某月时用此参数"},
                    "region": {"type": "string", "description": "地区：美国/英国/欧洲"},
                    "division": {"type": "string", "description": "事业部：卫浴事业部/儿童事业部/客厅事业部等"},
                    "detail": {"type": "boolean", "description": "是否返回明细（默认只返回汇总）"},
                    "limit": {"type": "integer", "description": "明细行数限制，默认20"}
                },
                "required": []
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "query_inventory",
            "description": "查询海外仓库库存数据。支持按SKU、仓库名称、地区、事业部筛选，可按事业部/仓库/SKU分组汇总。地区和事业部筛选会关联产品基础信息表。数据来自易仓API，可点击更新按钮同步最新数据。",
            "parameters": {
                "type": "object",
                "properties": {
                    "sku": {"type": "string", "description": "要查询的SKU编码"},
                    "warehouse": {"type": "string", "description": "按仓库名称筛选（支持模糊匹配）"},
                    "region": {"type": "string", "description": "地区：美国/英国/欧洲。按地区筛选时关联产品基础信息表中的地区字段"},
                    "division": {"type": "string", "description": "事业部：卫浴事业部/儿童事业部/客厅事业部等。关联产品基础信息表筛选"},
                    "group_by": {"type": "string", "description": "分组方式：division(按事业部)/warehouse(按仓库)/sku(按SKU)，默认自动选择", "enum": ["division", "warehouse", "sku"]},
                    "detail": {"type": "boolean", "description": "是否返回明细（默认只返回汇总）"},
                    "limit": {"type": "integer", "description": "明细行数限制，默认30"}
                },
                "required": []
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "list_stores",
            "description": "列出所有可用的店铺账号",
            "parameters": {"type": "object", "properties": {}, "required": []}
        }
    },
    {
        "type": "function",
        "function": {
            "name": "list_skus",
            "description": "列出销量表中的SKU及总销量（可按地区/事业部筛选）",
            "parameters": {
                "type": "object",
                "properties": {
                    "region": {"type": "string", "description": "地区：美国/英国/欧洲"},
                    "division": {"type": "string", "description": "事业部"},
                    "limit": {"type": "integer", "description": "返回数量限制，默认20"}
                },
                "required": []
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "list_divisions",
            "description": "列出所有事业部",
            "parameters": {"type": "object", "properties": {}, "required": []}
        }
    },
    {
        "type": "function",
        "function": {
            "name": "query_in_transit",
            "description": "查询在途货物数据。支持按SKU、市场(美国/英国/德国)、事业部、仓库分类筛选。数据来自钉钉AI表格，需先同步。",
            "parameters": {
                "type": "object",
                "properties": {
                    "sku": {"type": "string", "description": "要查询的SKU编码（模糊匹配）"},
                    "market": {"type": "string", "description": "目的市场：US/美国、UK/英国、DE/德国"},
                    "division": {"type": "string", "description": "事业部筛选"},
                    "warehouse": {"type": "string", "description": "仓库分类筛选（如FBM美西、CG美国等）"},
                    "arrived": {"type": "boolean", "description": "是否只查已到港的货物"},
                    "detail": {"type": "boolean", "description": "是否返回明细（默认只返回汇总）"},
                    "limit": {"type": "integer", "description": "明细行数限制，默认30"}
                },
                "required": []
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "query_factory_stock",
            "description": "查询工厂在产和在库数据。支持按SKU、事业部筛选。数据来自钉钉AI表格总台账表，需先同步。",
            "parameters": {
                "type": "object",
                "properties": {
                    "sku": {"type": "string", "description": "要查询的SKU编码（精确匹配）"},
                    "division": {"type": "string", "description": "事业部/产品线筛选"},
                    "detail": {"type": "boolean", "description": "是否返回明细（默认只返回汇总）"},
                    "limit": {"type": "integer", "description": "明细行数限制，默认30"}
                },
                "required": []
            }
        }
    }
]

SYSTEM_PROMPT_TEMPLATE = """你是一个博昱科技智能助手，专注于供应链数据查询。你可以帮助用户查询销量、库存、在途和工厂库存数据。

当前日期：{current_date}
数据最新时间：{max_date}
在途数据同步时间：{transit_sync_time}
工厂库存同步时间：{factory_sync_time}
海外库存同步时间：{inventory_sync_time}
销量数据同步时间：{sales_sync_time}

核心规则：
1. 仓库SKU处理：带"-"取第一个"-"之前的部分（如CR04800101MP-FBA → CR04800101MP）；以W/w开头的SKU不计入销量
2. 地区分类：US/CA/FA→美国地区，GB/UK/IE→英国地区，其余→欧洲地区（含DE）
3. 近N天的时间范围：以数据中最大付款时间为终点，往前推N-1天
4. 销量数据来自易仓API订单，无需状态筛选
5. SKU会根据地区在SKU对应表中查找映射
6. 在途查询支持按市场(US/UK/DE)、仓库分类、事业部、SKU筛选
7. 工厂库存查询支持按SKU、产品线/事业部筛选，显示在产数量和国内在库
8. 海外库存查询支持按SKU、仓库名称、地区、事业部筛选，显示可用数量；数据来自易仓API，可点击更新按钮同步

时间处理规则（重要！）：
- 用户说"X月"、"四月"、"5月"等绝对时间 → 转换为start_date和end_date参数，如"四月"→start_date="2026-04-01", end_date="2026-04-30"
- 用户说"近N天"、"最近7天"等相对时间 → 使用days参数
- 用户没说时间范围 → 不传时间参数，返回全部数据
- start_date和end_date格式必须为YYYY-MM-DD

回答要求：
- 简洁直接，不要冗余的客套话
- 用表格展示数据汇总
- 如有地区分布，同时展示各地区的销量/占比
- 金额保留2位小数，数量用千分位

工具调用规则（重要！）：
- 当用户同时问销量和库存时，必须在同一次回复中同时调用query_sales和query_inventory两个工具
- 当用户问在途或工厂库存时，调用query_in_transit或query_factory_stock
- 不要只调一个工具就回答，必须等所有工具结果都拿到后再汇总回答
- 例：用户问"XX的销量和库存" → 同时调query_sales(sku="XX")和query_inventory(sku="XX")
"""


# ============================================================
#  工具执行函数
# ============================================================

def execute_query_sales(args: dict) -> str:
    """执行销量查询（基于易仓API订单数据，无需状态筛选）"""
    sku = args.get('sku')
    store = args.get('store')
    days = args.get('days')
    start_date = args.get('start_date')
    end_date = args.get('end_date')
    region = args.get('region')
    division = args.get('division')
    detail = args.get('detail', False)
    limit = args.get('limit', 20)

    try:
        conn = db_connect()
        c = conn.cursor()

        # 查找最大付款时间
        c.execute("SELECT MAX(付款时间) FROM 销量 WHERE 付款时间 IS NOT NULL AND 付款时间 != ''")
        max_time_row = c.fetchone()
        max_time_str = max_time_row[0] if max_time_row else None

        if not max_time_str:
            conn.close()
            return "销量表中无付款时间数据"

        max_date = max_time_str.split(' ')[0].replace('/', '-')

        # 计算时间范围
        if days:
            try:
                max_dt = datetime.strptime(max_date, '%Y-%m-%d')
            except ValueError:
                try:
                    max_dt = datetime.strptime(max_date, '%Y/%m/%d')
                except ValueError:
                    conn.close()
                    return f"无法解析最大日期: {max_date}"
            end_dt = max_dt
            start_dt = max_dt - timedelta(days=days - 1)
            start_date = start_dt.strftime('%Y-%m-%d')
            end_date = end_dt.strftime('%Y-%m-%d')

        # 构建查询 — 不再筛选状态
        conditions = ["仓库SKU IS NOT NULL", "仓库SKU != ''",
                       "(仓库SKU NOT LIKE 'W%' AND 仓库SKU NOT LIKE 'w%')"]
        params = []

        if store:
            conditions.append("账号 = ?")
            params.append(store)

        if start_date:
            sd = start_date.replace('-', '/')
            conditions.append("付款时间 >= ?")
            params.append(sd)
        if end_date:
            ed = end_date.replace('-', '/')
            conditions.append("付款时间 <= ?")
            params.append(ed + ' 23:59:59')

        if region:
            target_region = REGION_ALIAS.get(region.lower() if isinstance(region, str) else region, region)
            if target_region == "美国地区":
                conditions.append("币种 IN (?, ?)")
                params.extend(['USD', 'CAD'])
            elif target_region == "英国地区":
                conditions.append("币种 = ?")
                params.append('GBP')
            elif target_region == "欧洲地区":
                conditions.append("币种 NOT IN (?, ?, ?)")
                params.extend(['USD', 'CAD', 'GBP'])

        where_clause = ' AND '.join(conditions)

        # 查询所有符合基础条件的行
        sql = f"SELECT 账号, 国家或地区代码, 付款时间, 仓库SKU, 数量, 币种, 产品名称, 总金额 FROM 销量 WHERE {where_clause}"
        c.execute(sql, params)
        rows = c.fetchall()

        # 在Python中处理SKU映射和事业部筛选
        results = []
        for row in rows:
            warehouse_sku = row[3] if row[3] else ''
            parsed = parse_sku(warehouse_sku)
            # 用币种判定地区（USD/CAD→US, GBP→UK, EUR→DE）
            currency = row[5] if row[5] else ''
            country_code = map_currency_to_region(currency)

            # SKU映射
            mapped_sku = resolve_sku(parsed, country_code)

            # 按SKU筛选
            if sku and mapped_sku != sku and parsed != sku:
                continue

            # 按事业部筛选
            if division:
                row_div = resolve_division(mapped_sku, country_code)
                if row_div != division:
                    continue

            qty = 0
            try:
                qty = int(float(row[4])) if row[4] else 0
            except (ValueError, TypeError):
                pass

            results.append({
                '账号': row[0],
                '国家或地区代码': country_code,
                '付款时间': row[2],
                '仓库SKU': warehouse_sku,
                '映射后SKU': mapped_sku,
                '数量': qty,
                '币种': row[5] if row[5] else '',
                '产品名称': row[6] if row[6] else '',
                '总金额': row[7] if row[7] else 0,
                '地区': classify_region(country_code),
            })

        conn.close()

        if not results:
            return "未找到匹配的销量数据"

        # 汇总
        total_qty = sum(r['数量'] for r in results)

        # 地区分布
        region_summary = defaultdict(int)
        for r in results:
            region_summary[r['地区']] += r['数量']

        # SKU分布
        sku_summary = defaultdict(int)
        for r in results:
            sku_summary[r['映射后SKU']] += r['数量']

        # 构建输出
        output_parts = []

        # 查询条件
        filters = []
        if sku:
            filters.append(f"SKU={sku}")
        if store:
            filters.append(f"店铺={store}")
        if region:
            filters.append(f"地区={region}")
        if division:
            filters.append(f"事业部={division}")
        if start_date and end_date:
            filters.append(f"时间={start_date}~{end_date}")
        output_parts.append(f"查询条件: {', '.join(filters)}")
        output_parts.append(f"总销量: {total_qty:,}")

        # 地区分布表
        region_rows = []
        for r_name in sorted(region_summary.keys()):
            qty = region_summary[r_name]
            pct = f"{qty / total_qty * 100:.1f}%" if total_qty > 0 else "0%"
            region_rows.append(f"  {r_name}: {qty:,} ({pct})")
        if len(region_rows) > 1:
            output_parts.append("地区分布:\n" + "\n".join(region_rows))

        # SKU分布（多SKU时展示）
        if not sku and len(sku_summary) > 1:
            sorted_skus = sorted(sku_summary.items(), key=lambda x: x[1], reverse=True)[:20]
            sku_rows = []
            for s, q in sorted_skus:
                pct = f"{q / total_qty * 100:.1f}%" if total_qty > 0 else "0%"
                sku_rows.append(f"  {s}: {q:,} ({pct})")
            output_parts.append("SKU分布 (Top 20):\n" + "\n".join(sku_rows))

        # 明细
        if detail:
            detail_rows = results[:limit]
            lines = []
            for r in detail_rows:
                lines.append(f"  {r['付款时间'][:10] if r['付款时间'] else ''} | {r['账号']} | {r['映射后SKU']} | {r['国家或地区代码']} | {r['数量']}")
            if len(results) > limit:
                lines.append(f"  ... 还有 {len(results) - limit} 行")
            output_parts.append("明细:\n" + "\n".join(lines))

        return "\n\n".join(output_parts)

    except Exception as e:
        return f"查询出错: {str(e)}"


def execute_query_inventory(args: dict) -> str:
    """执行库存查询，支持按地区/事业部筛选和分组
    新表结构：SKU, 仓库名称, 可用数量
    """
    sku = args.get('sku')
    warehouse = args.get('warehouse')
    region = args.get('region')
    division = args.get('division')
    group_by = args.get('group_by')
    detail = args.get('detail', False)
    limit = args.get('limit', 30)

    try:
        conn = db_connect()
        c = conn.cursor()

        # 判断是否需要关联产品信息（有地区或事业部筛选时需要）
        need_join = bool(region or division)

        if need_join:
            # 关联产品信息获取事业部和地区信息
            join_sql = '''
                SELECT i.SKU, i.仓库名称, i.可用数量,
                       p.所属事业部, p.地区
                FROM 海外仓库存 i
                INNER JOIN (
                    SELECT SKU, 所属事业部, 地区
                    FROM 产品信息
                    GROUP BY SKU, 地区
                ) p ON i.SKU = p.SKU
            '''

            conditions = []
            params = []

            if sku:
                conditions.append("i.SKU = ?")
                params.append(sku)
            if warehouse:
                conditions.append("i.仓库名称 LIKE ?")
                params.append(f"%{warehouse}%")

            # 地区筛选
            if region:
                target_region = REGION_ALIAS.get(region.lower() if isinstance(region, str) else region, region)
                if target_region == "美国地区":
                    region_val = "US"
                elif target_region == "英国地区":
                    region_val = "UK"
                else:
                    region_val = "DE"
                conditions.append("p.地区 = ?")
                params.append(region_val)

            # 事业部筛选
            if division:
                conditions.append("p.所属事业部 = ?")
                params.append(division)

            where_clause = ' AND '.join(conditions) if conditions else '1=1'
            sql = f"{join_sql} WHERE {where_clause}"
            c.execute(sql, params)
            rows = c.fetchall()

            # 获取库存同步时间
            c.execute("SELECT last_sync FROM 同步日志 WHERE table_name='海外仓库存'")
            inv_sync_row = c.fetchone()
            conn.close()

            if not rows:
                return "未找到匹配的库存数据"

            # rows: (SKU, 仓库名称, 可用数量, 所属事业部, 地区)
            total_stock = sum(row[2] for row in rows)

            div_summary = defaultdict(float)
            wh_summary = defaultdict(float)
            sku_summary = defaultdict(float)

            for row in rows:
                div_name = row[3]
                wh_name = row[1]
                sku_name = row[0]
                stock = row[2]
                div_summary[div_name] += stock
                wh_summary[wh_name] += stock
                sku_summary[sku_name] += stock

            output_parts = []
            filters = []
            if sku:
                filters.append(f"SKU={sku}")
            if warehouse:
                filters.append(f"仓库={warehouse}")
            if region:
                filters.append(f"地区={region}")
            if division:
                filters.append(f"事业部={division}")
            output_parts.append(f"查询条件: {', '.join(filters)}")
            output_parts.append(f"总库存: {total_stock:,.0f}")

            # 决定分组方式
            if group_by == 'division' or (not group_by and (region or division) and not sku and not warehouse):
                sorted_divs = sorted(div_summary.items(), key=lambda x: x[1], reverse=True)
                div_rows = []
                for d, q in sorted_divs:
                    pct = f"{q / total_stock * 100:.1f}%" if total_stock > 0 else "0%"
                    div_rows.append(f"  {d}: {q:,.0f} ({pct})")
                output_parts.append("按事业部汇总:\n" + "\n".join(div_rows))
            elif group_by == 'warehouse' or (not group_by and warehouse):
                sorted_wh = sorted(wh_summary.items(), key=lambda x: x[1], reverse=True)[:20]
                wh_rows = []
                for w, q in sorted_wh:
                    pct = f"{q / total_stock * 100:.1f}%" if total_stock > 0 else "0%"
                    wh_rows.append(f"  {w}: {q:,.0f} ({pct})")
                output_parts.append("按仓库汇总:\n" + "\n".join(wh_rows))
            else:
                # 默认按SKU
                sorted_skus = sorted(sku_summary.items(), key=lambda x: x[1], reverse=True)[:20]
                sku_rows = []
                for s, q in sorted_skus:
                    sku_rows.append(f"  {s}: {q:,.0f}")
                output_parts.append("按SKU汇总 (Top 20):\n" + "\n".join(sku_rows))

                # 同时展示事业部分布
                if len(div_summary) > 1:
                    sorted_divs = sorted(div_summary.items(), key=lambda x: x[1], reverse=True)
                    div_rows = []
                    for d, q in sorted_divs:
                        pct = f"{q / total_stock * 100:.1f}%" if total_stock > 0 else "0%"
                        div_rows.append(f"  {d}: {q:,.0f} ({pct})")
                    output_parts.append("事业部分布:\n" + "\n".join(div_rows))

            # 明细
            if detail:
                detail_rows = rows[:limit]
                lines = []
                for row in detail_rows:
                    lines.append(
                        f"  {row[0]} | {row[1]} | {row[3] or ''} | 可用数量:{row[2]:,.0f}"
                    )
                if len(rows) > limit:
                    lines.append(f"  ... 还有 {len(rows) - limit} 行")
                output_parts.append("明细:\n" + "\n".join(lines))

            # 同步时间
            if inv_sync_row and inv_sync_row[0]:
                output_parts.append(f"数据同步时间: {inv_sync_row[0]}")

            return "\n\n".join(output_parts)

        else:
            # 不关联产品信息
            conditions = []
            params = []

            if sku:
                conditions.append("SKU = ?")
                params.append(sku)
            if warehouse:
                conditions.append("仓库名称 LIKE ?")
                params.append(f"%{warehouse}%")

            where_clause = ' AND '.join(conditions) if conditions else '1=1'

            c.execute(f'''
                SELECT SKU, 仓库名称, 可用数量
                FROM 海外仓库存 WHERE {where_clause}
                ORDER BY 可用数量 DESC
            ''', params)
            rows = c.fetchall()

            # 获取库存同步时间
            c.execute("SELECT last_sync FROM 同步日志 WHERE table_name='海外仓库存'")
            inv_sync_row = c.fetchone()
            conn.close()

            if not rows:
                return "未找到匹配的库存数据"

            # 汇总
            total_stock = sum(row[2] for row in rows)

            sku_summary = defaultdict(float)
            warehouse_summary = defaultdict(float)
            for row in rows:
                sku_summary[row[0]] += row[2]
                warehouse_summary[row[1]] += row[2]

            output_parts = []
            filters = []
            if sku:
                filters.append(f"SKU={sku}")
            if warehouse:
                filters.append(f"仓库={warehouse}")
            output_parts.append(f"查询条件: {', '.join(filters)}")
            output_parts.append(f"总库存: {total_stock:,.0f}")

            # 按SKU汇总
            if not sku or len(sku_summary) > 1:
                sorted_skus = sorted(sku_summary.items(), key=lambda x: x[1], reverse=True)[:20]
                sku_rows = []
                for s, q in sorted_skus:
                    sku_rows.append(f"  {s}: {q:,.0f}")
                output_parts.append("SKU库存分布:\n" + "\n".join(sku_rows))

            # 按仓库汇总
            if not warehouse or len(warehouse_summary) > 1:
                sorted_wh = sorted(warehouse_summary.items(), key=lambda x: x[1], reverse=True)[:10]
                wh_rows = []
                for w, q in sorted_wh:
                    wh_rows.append(f"  {w}: {q:,.0f}")
                output_parts.append("仓库库存分布:\n" + "\n".join(wh_rows))

            # 明细
            if detail:
                detail_rows = rows[:limit]
                lines = []
                for row in detail_rows:
                    lines.append(
                        f"  {row[0]} | {row[1]} | 可用数量:{row[2]:,.0f}"
                    )
                if len(rows) > limit:
                    lines.append(f"  ... 还有 {len(rows) - limit} 行")
                output_parts.append("明细:\n" + "\n".join(lines))

            # 同步时间
            if inv_sync_row and inv_sync_row[0]:
                output_parts.append(f"数据同步时间: {inv_sync_row[0]}")

            return "\n\n".join(output_parts)

    except Exception as e:
        return f"查询出错: {str(e)}"


def execute_list_stores(args: dict) -> str:
    """列出所有店铺"""
    try:
        conn = db_connect()
        c = conn.cursor()
        c.execute("SELECT DISTINCT 账号 FROM 销量 WHERE 账号 IS NOT NULL AND 账号 != '' ORDER BY 账号")
        rows = c.fetchall()
        conn.close()
        if not rows:
            return "暂无店铺数据"
        stores = [row[0] for row in rows]
        return "可用店铺:\n" + "\n".join(f"  {i+1}. {s}" for i, s in enumerate(stores))
    except Exception as e:
        return f"查询出错: {str(e)}"


def execute_list_skus(args: dict) -> str:
    """列出SKU"""
    region = args.get('region')
    division = args.get('division')
    limit = args.get('limit', 20)

    try:
        conn = db_connect()
        c = conn.cursor()
        c.execute("""
            SELECT 仓库SKU, SUM(CAST(数量 AS REAL)) as total
            FROM 销量
            WHERE 仓库SKU IS NOT NULL AND 仓库SKU != ''
              AND 仓库SKU NOT LIKE 'W%' AND 仓库SKU NOT LIKE 'w%'
            GROUP BY 仓库SKU
            ORDER BY total DESC
        """)
        rows = c.fetchall()

        # 处理SKU解析和映射
        sku_summary = defaultdict(float)
        for row in rows:
            parsed = parse_sku(row[0])
            # 简单汇总（不做逐行地区映射，用原始解析后的SKU）
            sku_summary[parsed] += row[1]

        sorted_skus = sorted(sku_summary.items(), key=lambda x: x[1], reverse=True)[:limit]

        output = f"SKU销量排行 (Top {limit}):\n"
        for i, (s, q) in enumerate(sorted_skus):
            output += f"  {i+1}. {s}: {q:,.0f}\n"

        conn.close()
        return output.strip()
    except Exception as e:
        return f"查询出错: {str(e)}"


def execute_list_divisions(args: dict) -> str:
    """列出所有事业部"""
    try:
        conn = db_connect()
        c = conn.cursor()
        c.execute("SELECT DISTINCT 所属事业部 FROM 产品信息 WHERE 所属事业部 IS NOT NULL AND 所属事业部 != '' ORDER BY 所属事业部")
        rows = c.fetchall()
        conn.close()
        if not rows:
            return "暂无事业部数据"
        divs = [row[0] for row in rows]
        return "可用事业部:\n" + "\n".join(f"  {i+1}. {d}" for i, d in enumerate(divs))
    except Exception as e:
        return f"查询出错: {str(e)}"


def execute_query_in_transit(args: dict) -> str:
    """执行在途货物查询"""
    sku = args.get('sku')
    market = args.get('market')
    division = args.get('division')
    warehouse = args.get('warehouse')
    arrived = args.get('arrived', False)
    detail = args.get('detail', False)
    limit = args.get('limit', 30)

    try:
        conn = db_connect()
        c = conn.cursor()

        # 检查数据是否存在
        c.execute("SELECT COUNT(*) FROM 在途货物")
        total_count = c.fetchone()[0]
        if total_count == 0:
            conn.close()
            return "在途数据为空，请先点击「刷新数据」从钉钉同步在途货物数据"

        # 构建查询条件
        conditions = []
        params = []

        if market:
            m = market.upper().replace("德国", "DE").replace("美国", "US").replace("英国", "UK")
            conditions.append("市场代码 = ?")
            params.append(m)

        if sku:
            conditions.append("SKU LIKE ?")
            params.append(f"%{sku}%")

        if division:
            conditions.append("事业部 LIKE ?")
            params.append(f"%{division}%")

        if warehouse:
            conditions.append("仓库分类 LIKE ?")
            params.append(f"%{warehouse}%")

        if arrived:
            # 到港ETA <= 今天视为已到港
            today = datetime.now().strftime('%Y-%m-%d')
            conditions.append("(到港ETA IS NOT NULL AND 到港ETA != '' AND 到港ETA <= ?)")
            params.append(today)

        where_clause = ' AND '.join(conditions) if conditions else '1=1'

        c.execute(f'''
            SELECT 市场代码, 仓库分类, SKU, SUM(数量) as 总数量, 事业部, 送仓时间
            FROM 在途货物
            WHERE {where_clause}
            GROUP BY 市场代码, 仓库分类, SKU, 送仓时间
            ORDER BY 市场代码, 仓库分类, 总数量 DESC
        ''', params)
        agg_rows = c.fetchall()

        if not agg_rows:
            conn.close()
            return "未找到匹配的在途数据"

        total_qty = sum(row[3] for row in agg_rows)

        market_names = {"US": "美国", "UK": "英国", "DE": "德国", "CA": "加拿大"}
        market_summary = defaultdict(int)
        for row in agg_rows:
            market_summary[market_names.get(row[0], row[0])] += row[3]

        output_parts = []
        filters = []
        if sku: filters.append(f"SKU={sku}")
        if market: filters.append(f"市场={market}")
        if division: filters.append(f"事业部={division}")
        if warehouse: filters.append(f"仓库={warehouse}")
        if arrived: filters.append("已到港")
        output_parts.append(f"查询条件: {', '.join(filters)}")
        output_parts.append(f"在途总量: {total_qty:,}")

        # 市场分布
        if len(market_summary) > 1:
            mkt_rows = []
            for m, q in sorted(market_summary.items(), key=lambda x: x[1], reverse=True):
                pct = f"{q / total_qty * 100:.1f}%" if total_qty > 0 else "0%"
                mkt_rows.append(f"  {m}: {q:,} ({pct})")
            output_parts.append("市场分布:\n" + "\n".join(mkt_rows))

        # 汇总表（按市场+仓库+SKU+送仓时间）
        top_rows = agg_rows[:30]
        lines = [f"  {'市场':<6} {'仓库':<10} {'SKU':<20} {'数量':>8} {'送仓时间':<14}"]
        lines.append("  " + "-" * 64)
        for row in top_rows:
            mkt_display = market_names.get(row[0], row[0])
            delivery_display = str(row[5])[:10] if row[5] else "暂无送仓时间"
            lines.append(f"  {mkt_display:<6} {row[1]:<10} {row[2]:<20} {row[3]:>8,} {delivery_display:<14}")
        if len(agg_rows) > 30:
            lines.append(f"  ... 还有 {len(agg_rows) - 30} 行")
        output_parts.append("在途汇总:\n" + "\n".join(lines))

        # 明细
        if detail:
            c.execute(f'''
                SELECT 市场代码, 仓库分类, SKU, 数量, 订单号, 事业部, 到港ETA, 送仓时间
                FROM 在途货物
                WHERE {where_clause}
                ORDER BY 市场代码, 仓库分类, SKU
                LIMIT ?
            ''', params + [limit])
            detail_rows = c.fetchall()
            dlines = [f"  {'市场':<6} {'仓库':<10} {'SKU':<18} {'数量':>6} {'订单号':<16} {'到港ETA':<14} {'送仓时间':<14} {'事业部':<10}"]
            dlines.append("  " + "-" * 104)
            for row in detail_rows:
                mkt_display = market_names.get(row[0], row[0])
                eta_display = str(row[6])[:10] if row[6] else "-"
                delivery_display = str(row[7])[:10] if row[7] else "暂无送仓时间"
                dlines.append(f"  {mkt_display:<6} {row[1]:<10} {row[2]:<18} {row[3]:>6,} {row[4] or '-':<16} {eta_display:<14} {delivery_display:<14} {row[5] or '-':<10}")
            output_parts.append("明细:\n" + "\n".join(dlines))

        # 同步时间
        c.execute("SELECT last_sync FROM 同步日志 WHERE table_name='在途货物'")
        sync_row = c.fetchone()
        conn.close()
        if sync_row:
            output_parts.append(f"数据同步时间: {sync_row[0]}")

        return "\n\n".join(output_parts)

    except Exception as e:
        return f"查询出错: {str(e)}"


def execute_query_factory_stock(args: dict) -> str:
    """执行工厂库存（在产/在库）查询"""
    sku = args.get('sku')
    division = args.get('division')
    detail = args.get('detail', False)
    limit = args.get('limit', 30)

    try:
        conn = db_connect()
        c = conn.cursor()

        # 检查数据是否存在
        c.execute("SELECT COUNT(*) FROM 工厂库存")
        total_count = c.fetchone()[0]
        if total_count == 0:
            conn.close()
            return "工厂库存数据为空，请先点击「刷新数据」从钉钉同步总台账数据"

        # 构建查询条件
        conditions = []
        params = []

        if sku:
            conditions.append("SKU = ?")
            params.append(sku)

        if division:
            conditions.append("产品线 LIKE ?")
            params.append(f"%{division}%")

        where_clause = ' AND '.join(conditions) if conditions else '1=1'

        # 汇总查询：按SKU合并（同一SKU可能有多条订单）
        c.execute(f'''
            SELECT SKU, 品名,
                   SUM(在产数量) as 总在产,
                   SUM(国内在库) as 总在库,
                   SUM(待生产数量) as 总待生产,
                   MAX(库龄天数) as 最大库龄,
                   GROUP_CONCAT(DISTINCT 工厂简写) as 工厂,
                   MAX(产品线) as 产品线
            FROM 工厂库存
            WHERE {where_clause}
            GROUP BY SKU
            ORDER BY 总在产 + 总在库 DESC
        ''', params)
        agg_rows = c.fetchall()

        if not agg_rows:
            conn.close()
            return "未找到匹配的工厂库存数据"

        total_prod = sum(row[2] for row in agg_rows)
        total_stock = sum(row[3] for row in agg_rows)
        total_pending = sum(row[4] for row in agg_rows)

        output_parts = []
        filters = []
        if sku: filters.append(f"SKU={sku}")
        if division: filters.append(f"事业部={division}")
        output_parts.append(f"查询条件: {', '.join(filters)}")
        output_parts.append(f"总在产: {total_prod:,}  |  总国内在库: {total_stock:,}  |  总待生产: {total_pending:,}")

        # 事业部分布
        div_summary = defaultdict(lambda: [0, 0, 0])
        for row in agg_rows:
            div_name = row[7] or "未分类"
            div_summary[div_name][0] += row[2]
            div_summary[div_name][1] += row[3]
            div_summary[div_name][2] += row[4]

        if len(div_summary) > 1:
            div_lines = []
            for d, (p, s, pend) in sorted(div_summary.items(), key=lambda x: x[1][0] + x[1][1], reverse=True):
                div_lines.append(f"  {d}: 在产 {p:,} | 在库 {s:,} | 待生产 {pend:,}")
            output_parts.append("事业部分布:\n" + "\n".join(div_lines))

        # SKU汇总表
        top_rows = agg_rows[:30]
        lines = [f"  {'SKU':<20} {'品名':<12} {'在产':>8} {'在库':>8} {'待生产':>8} {'库龄':>6} {'工厂':<12}"]
        lines.append("  " + "-" * 90)
        for row in top_rows:
            name_display = (row[1] or '')[:10]
            factory_display = (row[6] or '')[:10]
            lines.append(f"  {row[0]:<20} {name_display:<12} {row[2]:>8,} {row[3]:>8,} {row[4]:>8,} {row[5] or 0:>6} {factory_display:<12}")
        if len(agg_rows) > 30:
            lines.append(f"  ... 还有 {len(agg_rows) - 30} 行")
        output_parts.append("工厂库存汇总:\n" + "\n".join(lines))

        # 明细
        if detail:
            c.execute(f'''
                SELECT SKU, 品名, 在产数量, 国内在库, 待生产数量, 库龄天数,
                       工厂简写, 产品线, 订单号, 订单状态, 未交货数
                FROM 工厂库存
                WHERE {where_clause}
                ORDER BY SKU
                LIMIT ?
            ''', params + [limit])
            detail_rows = c.fetchall()
            dlines = [f"  {'SKU':<18} {'在产':>6} {'在库':>6} {'库龄':>4} {'工厂':<8} {'订单号':<14} {'状态':<8} {'未交货':>6}"]
            dlines.append("  " + "-" * 80)
            for row in detail_rows:
                dlines.append(f"  {row[0]:<18} {row[2]:>6,} {row[3]:>6,} {row[5] or 0:>4} {(row[6] or ''):<8} {(row[8] or ''):<14} {(row[9] or ''):<8} {row[10]:>6,}")
            output_parts.append("明细:\n" + "\n".join(dlines))

        # 同步时间
        c.execute("SELECT last_sync FROM 同步日志 WHERE table_name='工厂库存'")
        sync_row = c.fetchone()
        conn.close()
        if sync_row:
            output_parts.append(f"数据同步时间: {sync_row[0]}")

        return "\n\n".join(output_parts)

    except Exception as e:
        return f"查询出错: {str(e)}"


# 工具执行映射
TOOL_EXECUTORS = {
    "query_sales": execute_query_sales,
    "query_inventory": execute_query_inventory,
    "list_stores": execute_list_stores,
    "list_skus": execute_list_skus,
    "list_divisions": execute_list_divisions,
    "query_in_transit": execute_query_in_transit,
    "query_factory_stock": execute_query_factory_stock,
}


# ============================================================
#  AI 对话 API
# ============================================================

@app.route('/')
def index():
    return send_from_directory('static', 'index.html')


@app.route('/api/chat', methods=['POST'])
def chat():
    """AI对话接口，支持Function Calling"""
    try:
        return _chat_impl()
    except Exception as e:
        return jsonify({'error': f'服务器内部错误: {str(e)}'})


# ========== 同步状态管理 ==========
_sync_status = {
    'in_transit': {'status': 'idle', 'message': '', 'last_sync': None},
    'factory_stock': {'status': 'idle', 'message': '', 'last_sync': None},
    'inventory': {'status': 'idle', 'message': '', 'last_sync': None},
    'sales': {'status': 'idle', 'message': '', 'last_sync': None},
}


def _run_sync_in_background(sync_type):
    """后台线程执行同步"""
    status_key = sync_type
    try:
        _sync_status[status_key]['status'] = 'syncing'
        _sync_status[status_key]['message'] = '正在同步...'

        # 确保表存在
        conn = db_connect()
        c = conn.cursor()
        if sync_type == 'in_transit':
            c.execute('''CREATE TABLE IF NOT EXISTS 在途货物 (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                市场代码 TEXT, 市场原文 TEXT, 仓库分类 TEXT, 仓库名称 TEXT,
                SKU TEXT, 数量 INTEGER, 订单号 TEXT, 事业部 TEXT,
                到港ETA TEXT, 送仓时间 TEXT, 同步时间 TEXT
            )''')
            c.execute('''CREATE TABLE IF NOT EXISTS 同步日志 (
                table_name TEXT PRIMARY KEY, last_sync TEXT, record_count INTEGER DEFAULT 0
            )''')
        elif sync_type == 'factory_stock':
            c.execute('''CREATE TABLE IF NOT EXISTS 工厂库存 (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                SKU TEXT, 品名 TEXT, 在产数量 INTEGER DEFAULT 0,
                国内在库 INTEGER DEFAULT 0, 待生产数量 INTEGER DEFAULT 0,
                库龄天数 INTEGER DEFAULT 0, 工厂简写 TEXT, 产品线 TEXT,
                订单号 TEXT, 采购单价 REAL DEFAULT 0, 订单状态 TEXT,
                目的地 TEXT, 下单日期 TEXT, 未交货数 INTEGER DEFAULT 0, 订单数量 INTEGER DEFAULT 0,
                出货总数 REAL DEFAULT 0, 合同交期 TEXT, 实际交期 TEXT,
                位置区域 TEXT, SPU TEXT, 同步时间 TEXT
            )''')
            # 迁移：为已有表补充下单日期列
            try:
                c.execute("ALTER TABLE 工厂库存 ADD COLUMN 下单日期 TEXT")
            except sqlite3.OperationalError:
                pass  # 列已存在
            c.execute('''CREATE TABLE IF NOT EXISTS 同步日志 (
                table_name TEXT PRIMARY KEY, last_sync TEXT, record_count INTEGER DEFAULT 0
            )''')
        elif sync_type == 'inventory':
            c.execute('''CREATE TABLE IF NOT EXISTS 海外仓库存 (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                SKU TEXT, 仓库名称 TEXT, 可用数量 INTEGER DEFAULT 0,
                市场 TEXT DEFAULT ''
            )''')
            try:
                c.execute('ALTER TABLE 海外仓库存 ADD COLUMN 市场 TEXT DEFAULT \"\"')
            except:
                pass
            c.execute('''CREATE TABLE IF NOT EXISTS 同步日志 (
                table_name TEXT PRIMARY KEY, last_sync TEXT, record_count INTEGER DEFAULT 0
            )''')
        elif sync_type == 'sales':
            c.execute('''CREATE TABLE IF NOT EXISTS 销量 (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                订单ID TEXT, 平台 TEXT, 销售单号 TEXT, 仓库代码 TEXT,
                账号 TEXT, 国家或地区代码 TEXT, 币种 TEXT, 付款时间 TEXT,
                仓库SKU TEXT, 数量 INTEGER, 单价 REAL, 产品名称 TEXT,
                总金额 REAL, 销售额 REAL, 运费 REAL, 订单类型 TEXT,
                发货类型 INTEGER
            )''')
            c.execute('''CREATE TABLE IF NOT EXISTS 同步日志 (
                table_name TEXT PRIMARY KEY, last_sync TEXT, record_count INTEGER DEFAULT 0
            )''')
        conn.commit()
        conn.close()

        if sync_type == 'in_transit':
            from data import dingtalk_sync
            result = dingtalk_sync.sync_in_transit()
        elif sync_type == 'factory_stock':
            from data import dingtalk_sync
            result = dingtalk_sync.sync_factory_stock()
        elif sync_type == 'inventory':
            from data import eccang_sync
            # 获取上次同步时间作为update_time_from
            conn2 = db_connect()
            c2 = conn2.cursor()
            c2.execute("SELECT last_sync FROM 同步日志 WHERE table_name='海外仓库存'")
            row = c2.fetchone()
            update_from = None
            if row and row[0]:
                # "2026-05-14 14:30:00" → "2026-05-14"（易仓API要求横杠格式）
                date_str = str(row[0]).split(' ')[0]
                # 确保是横杠格式（兼容可能的斜杠格式存储）
                date_str = date_str.replace('/', '-')
                update_from = date_str
            conn2.close()
            result = eccang_sync.sync_inventory(update_from)
        elif sync_type == 'sales':
            from data import eccang_sales_sync
            # 增量同步：从最晚付款时间到昨天
            result = eccang_sales_sync.sync_sales()
        else:
            from data import dingtalk_sync
            result = dingtalk_sync.sync_in_transit()
            result2 = dingtalk_sync.sync_factory_stock()

        _sync_status[status_key]['status'] = 'success'
        _sync_status[status_key]['message'] = f"同步完成: {result.get('count', 0)} 条"
        _sync_status[status_key]['last_sync'] = result.get('time', '')
    except Exception as e:
        _sync_status[status_key]['status'] = 'error'
        _sync_status[status_key]['message'] = f"同步失败: {str(e)}"


@app.route('/api/sync/<sync_type>', methods=['POST'])
def sync_data(sync_type):
    """触发钉钉数据同步"""
    if sync_type not in ('in_transit', 'factory_stock', 'inventory', 'sales', 'all'):
        return jsonify({'error': f'未知同步类型: {sync_type}'}), 400

    # 检查是否正在同步
    if sync_type == 'all':
        if any(_sync_status[k]['status'] == 'syncing' for k in ('in_transit', 'factory_stock', 'inventory', 'sales')):
            return jsonify({'error': '正在同步中，请稍后再试'}), 409
    else:
        if _sync_status.get(sync_type, {}).get('status') == 'syncing':
            return jsonify({'error': '正在同步中，请稍后再试'}), 409

    # 启动后台同步
    if sync_type == 'all':
        t1 = threading.Thread(target=_run_sync_in_background, args=('in_transit',), daemon=True)
        t2 = threading.Thread(target=_run_sync_in_background, args=('factory_stock',), daemon=True)
        t3 = threading.Thread(target=_run_sync_in_background, args=('inventory',), daemon=True)
        t4 = threading.Thread(target=_run_sync_in_background, args=('sales',), daemon=True)
        t1.start()
        t2.start()
        t3.start()
        t4.start()
    else:
        t = threading.Thread(target=_run_sync_in_background, args=(sync_type,), daemon=True)
        t.start()

    return jsonify({'message': '同步已启动', 'type': sync_type})


@app.route('/api/sync/status', methods=['GET'])
def sync_status():
    """获取同步状态"""
    # 从数据库读取最新同步时间
    try:
        conn = db_connect()
        c = conn.cursor()
        c.execute("SELECT table_name, last_sync, record_count FROM 同步日志")
        db_status = {}
        for row in c.fetchall():
            db_status[row[0]] = {'last_sync': row[1], 'record_count': row[2]}

        # 表行数
        c.execute("SELECT COUNT(*) FROM 在途货物")
        transit_count = c.fetchone()[0]
        c.execute("SELECT COUNT(*) FROM 工厂库存")
        factory_count = c.fetchone()[0]
        c.execute("SELECT COUNT(*) FROM 海外仓库存")
        inventory_count = c.fetchone()[0]
        c.execute("SELECT COUNT(*) FROM 销量")
        sales_count = c.fetchone()[0]
        conn.close()
    except Exception:
        db_status = {}
        transit_count = 0
        factory_count = 0
        inventory_count = 0
        sales_count = 0

    return jsonify({
        'in_transit': {
            **_sync_status.get('in_transit', {}),
            'record_count': transit_count,
            'last_sync': db_status.get('在途货物', {}).get('last_sync') or _sync_status.get('in_transit', {}).get('last_sync'),
        },
        'factory_stock': {
            **_sync_status.get('factory_stock', {}),
            'record_count': factory_count,
            'last_sync': db_status.get('工厂库存', {}).get('last_sync') or _sync_status.get('factory_stock', {}).get('last_sync'),
        },
        'inventory': {
            **_sync_status.get('inventory', {}),
            'record_count': inventory_count,
            'last_sync': db_status.get('海外仓库存', {}).get('last_sync') or _sync_status.get('inventory', {}).get('last_sync'),
        },
        'sales': {
            **_sync_status.get('sales', {}),
            'record_count': sales_count,
            'last_sync': db_status.get('销量', {}).get('last_sync') or _sync_status.get('sales', {}).get('last_sync'),
        },
    })


def _chat_impl():
    data = request.json
    messages = data.get('messages', [])
    api_config = data.get('api_config', {})

    api_url = api_config.get('url', '').strip()
    api_key = api_config.get('key', '').strip()
    model = api_config.get('model', '').strip()

    if not api_url or not api_key:
        return jsonify({'error': '请先配置API地址和Key'}), 400

    if not model:
        model = 'deepseek-chat'

    try:
        client = OpenAI(base_url=api_url, api_key=api_key, timeout=90.0, max_retries=1)

        # 动态生成system prompt，注入当前日期和数据最新时间
        current_date = datetime.now().strftime('%Y-%m-%d')
        max_date = current_date
        try:
            conn_tmp = db_connect()
            c_tmp = conn_tmp.cursor()
            c_tmp.execute("SELECT MAX(付款时间) FROM 销量 WHERE 付款时间 IS NOT NULL AND 付款时间 != ''")
            row_tmp = c_tmp.fetchone()
            if row_tmp and row_tmp[0]:
                max_date = row_tmp[0].split(' ')[0].replace('/', '-')
            # 获取同步时间
            transit_sync_time = "未同步"
            factory_sync_time = "未同步"
            inventory_sync_time = "未同步"
            sales_sync_time = "未同步"
            c_tmp.execute("SELECT last_sync FROM 同步日志 WHERE table_name='在途货物'")
            st_row = c_tmp.fetchone()
            if st_row and st_row[0]:
                transit_sync_time = st_row[0]
            c_tmp.execute("SELECT last_sync FROM 同步日志 WHERE table_name='工厂库存'")
            st_row = c_tmp.fetchone()
            if st_row and st_row[0]:
                factory_sync_time = st_row[0]
            c_tmp.execute("SELECT last_sync FROM 同步日志 WHERE table_name='海外仓库存'")
            st_row = c_tmp.fetchone()
            if st_row and st_row[0]:
                inventory_sync_time = st_row[0]
            c_tmp.execute("SELECT last_sync FROM 同步日志 WHERE table_name='销量'")
            st_row = c_tmp.fetchone()
            if st_row and st_row[0]:
                sales_sync_time = st_row[0]
            conn_tmp.close()
        except Exception:
            transit_sync_time = "未同步"
            factory_sync_time = "未同步"
            inventory_sync_time = "未同步"
            sales_sync_time = "未同步"

        system_content = SYSTEM_PROMPT_TEMPLATE.format(
            current_date=current_date,
            max_date=max_date,
            transit_sync_time=transit_sync_time,
            factory_sync_time=factory_sync_time,
            inventory_sync_time=inventory_sync_time,
            sales_sync_time=sales_sync_time,
        )
        system_msg = {"role": "system", "content": system_content}
        all_messages = [system_msg] + messages

        # 循环式工具调用：支持AI在一次对话中调用多个工具（如同时查销量+库存）
        max_rounds = 10  # 最多10轮工具调用，防止死循环
        all_tool_calls = []  # 收集所有工具调用记录

        for round_idx in range(max_rounds):
            response = client.chat.completions.create(
                model=model,
                messages=all_messages,
                tools=TOOLS,
                tool_choice="auto",
                temperature=0.3,
            )

            choice = response.choices[0]
            assistant_msg = choice.message

            # 如果没有工具调用，返回最终回复
            if not assistant_msg.tool_calls:
                final_reply = assistant_msg.content or ''
                return jsonify({
                    'reply': final_reply,
                    'tool_calls': all_tool_calls
                })

            # 有工具调用，执行并继续
            all_messages.append(assistant_msg)

            for tool_call in assistant_msg.tool_calls:
                func_name = tool_call.function.name
                func_args = json.loads(tool_call.function.arguments)

                # 记录工具调用
                all_tool_calls.append({
                    'name': func_name,
                    'args': tool_call.function.arguments
                })

                # 执行工具
                executor = TOOL_EXECUTORS.get(func_name)
                if executor:
                    result = executor(func_args)
                else:
                    result = f"未知工具: {func_name}"

                all_messages.append({
                    "role": "tool",
                    "tool_call_id": tool_call.id,
                    "content": result
                })

        # 超过最大轮数，强制获取最终回复
        response = client.chat.completions.create(
            model=model,
            messages=all_messages,
            tools=TOOLS,
            tool_choice="none",  # 强制不调工具
            temperature=0.3,
        )
        final_reply = response.choices[0].message.content or '查询轮次过多，请缩小查询范围'

        return jsonify({
            'reply': final_reply,
            'tool_calls': all_tool_calls
        })

    except Exception as e:
        return jsonify({'error': f'AI调用失败: {str(e)}'}), 500


@app.route('/api/health', methods=['GET'])
def health():
    """轻量健康检查 - 不做重查询，避免超时误触发守护进程重启"""
    db_exists = os.path.exists(DB_PATH)
    db_size = os.path.getsize(DB_PATH) if db_exists else 0
    # 只做快速连接测试，不遍历所有表
    db_ok = False
    if db_exists:
        try:
            conn = db_connect()
            conn.execute("SELECT 1")
            db_ok = True
            conn.close()
        except Exception:
            pass

    return jsonify({
        'status': 'ok' if db_ok else 'degraded',
        'db_exists': db_exists,
        'db_ok': db_ok,
        'db_size_mb': round(db_size / 1024 / 1024, 2),
    })


@app.route('/api/tables', methods=['GET'])
def list_tables():
    """获取所有数据表及其行数和列信息"""
    try:
        conn = db_connect()
        c = conn.cursor()
        c.execute("SELECT name FROM sqlite_master WHERE type='table' AND name != 'sqlite_sequence' ORDER BY name")
        table_names = [row[0] for row in c.fetchall()]

        result = []
        for tname in table_names:
            if tname == 'sqlite_sequence':
                continue
            # 行数
            c.execute(f'SELECT COUNT(*) FROM "{tname}"')
            row_count = c.fetchone()[0]
            # 列信息
            c.execute(f'PRAGMA table_info("{tname}")')
            columns = [{'name': col[1], 'type': col[2]} for col in c.fetchall()]
            # 分类
            category = TABLE_CATEGORIES.get(tname, '其他数据')
            result.append({
                'name': tname,
                'rows': row_count,
                'columns': columns,
                'category': category,
            })
        conn.close()
        return jsonify({'tables': result, 'categories': sorted(set(TABLE_CATEGORIES.values())) + ['其他数据']})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/table/<table_name>', methods=['GET'])
def get_table_data(table_name):
    """获取指定表的数据，支持分页和搜索"""
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 100, type=int)
    search = request.args.get('search', '').strip()

    try:
        conn = db_connect()
        c = conn.cursor()

        # 安全校验表名
        c.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?", (table_name,))
        if not c.fetchone():
            conn.close()
            return jsonify({'error': f'表 {table_name} 不存在'}), 404

        # 列信息
        c.execute(f'PRAGMA table_info("{table_name}")')
        columns = [col[1] for col in c.fetchall()]

        # 总行数
        c.execute(f'SELECT COUNT(*) FROM "{table_name}"')
        total = c.fetchone()[0]

        # 搜索过滤
        where_clause = ''
        params = []
        if search:
            search_conditions = []
            for col in columns:
                search_conditions.append(f'CAST("{col}" AS TEXT) LIKE ?')
                params.append(f'%{search}%')
            where_clause = ' WHERE ' + ' OR '.join(search_conditions)

        # 过滤后总行数
        if search:
            c.execute(f'SELECT COUNT(*) FROM "{table_name}"{where_clause}', params)
            filtered_total = c.fetchone()[0]
        else:
            filtered_total = total

        # 分页查询
        offset = (page - 1) * per_page
        c.execute(f'SELECT * FROM "{table_name}"{where_clause} LIMIT ? OFFSET ?',
                  params + [per_page, offset])
        rows = c.fetchall()

        data = []
        for row in rows:
            row_dict = {}
            for i, col in enumerate(columns):
                val = row[i]
                # 处理不可序列化的类型
                if val is None:
                    row_dict[col] = None
                elif isinstance(val, bytes):
                    row_dict[col] = val.hex()
                else:
                    row_dict[col] = val
            data.append(row_dict)

        conn.close()

        return jsonify({
            'table': table_name,
            'columns': columns,
            'data': data,
            'total': total,
            'filtered_total': filtered_total,
            'page': page,
            'per_page': per_page,
            'total_pages': max(1, (filtered_total + per_page - 1) // per_page),
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/export/<table_name>', methods=['GET'])
def export_table(table_name):
    """导出指定表为 Excel 文件（支持搜索筛选）"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from flask import send_file
    import io

    # 安全检查
    allowed_tables = _get_all_table_names()
    if table_name not in allowed_tables:
        return jsonify({'error': f'表 "{table_name}" 不存在'}), 404

    search = request.args.get('search', '').strip()
    conn = db_connect()
    c = conn.cursor()

    # 获取列名
    c.execute(f'PRAGMA table_info("{table_name}")')
    columns = [col[1] for col in c.fetchall()]

    # 构建查询
    if search:
        where_parts = []
        for col in columns:
            where_parts.append(f'"{col}" LIKE ?')
        where_clause = ' WHERE ' + ' OR '.join(where_parts)
        c.execute(f'SELECT * FROM "{table_name}"{where_clause}',
                  [f'%{search}%'] * len(columns))
    else:
        c.execute(f'SELECT * FROM "{table_name}"')

    rows = c.fetchall()
    conn.close()

    # 创建 Excel
    wb = Workbook()
    ws = wb.active
    ws.title = table_name[:31]

    # 表头样式
    header_font = Font(name='微软雅黑', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='3B82F6', end_color='3B82F6', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB'),
    )
    data_alignment = Alignment(vertical='center')
    data_font = Font(name='微软雅黑', size=10)

    # 写入表头
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # 写入数据
    for row_idx, row in enumerate(rows, 2):
        for col_idx, val in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if val is None:
                cell.value = ''
            elif isinstance(val, bytes):
                cell.value = val.hex()
            else:
                cell.value = val
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = thin_border

    # 自动列宽
    for col_idx, col_name in enumerate(columns, 1):
        max_len = len(str(col_name)) * 2  # 中文字符算双倍宽
        for row_idx in range(2, min(len(rows) + 2, 101)):  # 取前100行估算
            val = ws.cell(row=row_idx, column=col_idx).value
            if val:
                cell_len = len(str(val))
                # 粗略处理中文字符
                cn_count = sum(1 for ch in str(val) if '\u4e00' <= ch <= '\u9fff')
                cell_len = cell_len + cn_count  # 中文字符翻倍
                max_len = max(max_len, cell_len)
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 4, 50)

    # 冻结首行
    ws.freeze_panes = 'A2'

    # 写入内存
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f'{table_name}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename,
    )


def _get_all_table_names():
    """获取所有表名"""
    try:
        conn = db_connect()
        c = conn.cursor()
        c.execute("SELECT name FROM sqlite_master WHERE type='table' AND name != 'sqlite_sequence' ORDER BY name")
        names = [row[0] for row in c.fetchall()]
        conn.close()
        return names
    except Exception:
        return []


# ============================================================
#  表模板 / 替换 / 追加 API
# ============================================================

def _get_table_columns(table_name):
    """获取表的列信息（含类型），跳过 id 主键列"""
    conn = db_connect()
    c = conn.cursor()
    c.execute(f'PRAGMA table_info("{table_name}")')
    cols = c.fetchall()
    conn.close()
    # columns (cid, name, type, notnull, default, pk)
    # 返回: [(name, type), ...], 跳过id列
    return [(col[1], col[2]) for col in cols if col[1].lower() != 'id']


@app.route('/api/table/<table_name>/template', methods=['GET'])
def download_table_template(table_name):
    """下载表数据模板（空Excel，含表头）"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from flask import send_file
    import io

    allowed_tables = _get_all_table_names()
    if table_name not in allowed_tables:
        return jsonify({'error': f'表 "{table_name}" 不存在'}), 404

    columns = _get_table_columns(table_name)
    if not columns:
        return jsonify({'error': f'表 "{table_name}" 无可用列'}), 400

    wb = Workbook()
    ws = wb.active
    ws.title = table_name[:31]

    header_font = Font(name='微软雅黑', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='3B82F6', end_color='3B82F6', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB'),
    )

    # 说明行
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
    note_cell = ws.cell(row=1, column=1,
        value=f'提示：请按表头填入数据，id列自动生成无需填写。替换/追加操作会按列名匹配导入。')
    note_cell.font = Font(name='微软雅黑', size=10, color='EF4444')
    note_cell.alignment = Alignment(wrap_text=True, vertical='center')

    for col_idx, (col_name, col_type) in enumerate(columns, 1):
        cell = ws.cell(row=2, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    for col_idx in range(1, len(columns) + 1):
        ws.column_dimensions[ws.cell(row=2, column=col_idx).column_letter].width = 16

    ws.row_dimensions[1].height = 28
    ws.freeze_panes = 'A3'

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'{table_name}_模板.xlsx',
    )


def _parse_uploaded_excel(file_or_bytes, columns_expected):
    """解析上传的Excel文件/bytes，返回行列表（每行为列值tuple）"""
    from openpyxl import load_workbook
    from io import BytesIO

    data = file_or_bytes if isinstance(file_or_bytes, bytes) else file_or_bytes.read()
    wb = load_workbook(BytesIO(data))
    ws = wb.active

    # 读取表头（跳过说明行）
    headers = []
    data_start_row = 1
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), 1):
        if not row or all(v is None for v in row):
            continue
        # 检查是否是表头行
        first_val = str(row[0]).strip() if row[0] else ''
        if first_val and not first_val.startswith('提示') and not first_val.startswith('说明'):
            headers = [str(v).strip() if v else '' for v in row]
            data_start_row = row_idx + 1
            break

    if not headers:
        return None, '未找到表头行'

    # 校验列名匹配
    col_names = [c[0] for c in columns_expected]
    col_map = {}  # Excel列索引 → DB列名
    for db_col, db_type in columns_expected:
        found = False
        for h_idx, h in enumerate(headers):
            if h == db_col:
                col_map[h_idx] = db_col
                found = True
                break
        if not found:
            return None, f'缺少必需列: "{db_col}"'

    # 读取数据行
    rows = []
    for row in ws.iter_rows(min_row=data_start_row, values_only=True):
        if not row or all(v is None for v in row):
            continue
        values = []
        all_empty = True
        for c_idx in range(len(headers)):
            if c_idx in col_map:
                val = row[c_idx] if c_idx < len(row) else None
                if val is not None:
                    all_empty = False
                values.append(val)
        if not all_empty:
            rows.append(values)

    return rows, None


@app.route('/api/table/<table_name>/replace', methods=['POST'])
def replace_table_data(table_name):
    """替换数据：清空表后导入Excel"""
    try:
        allowed_tables = _get_all_table_names()
        if table_name not in allowed_tables:
            return jsonify({'error': f'表 "{table_name}" 不存在'}), 404

        if 'file' not in request.files:
            return jsonify({'error': '请上传Excel文件'}), 400

        file = request.files['file']
        if not file.filename.endswith(('.xlsx', '.xls')):
            return jsonify({'error': '仅支持 .xlsx 或 .xls 格式'}), 400

        columns = _get_table_columns(table_name)
        if not columns:
            return jsonify({'error': f'表 "{table_name}" 无可用列'}), 400

        # 保存上传文件内容到内存，避免多次读取
        file_data = file.read()

        # 解析Excel（可能抛出异常，需捕获）
        try:
            rows, error = _parse_uploaded_excel(file_data, columns)
        except Exception as e:
            return jsonify({'error': f'文件解析失败: {str(e)}'}), 400
        if error:
            return jsonify({'error': error}), 400
        if not rows:
            return jsonify({'error': '文件中没有有效数据行'}), 400

        col_names = [c[0] for c in columns]
        placeholders = ', '.join(['?'] * len(col_names))
        col_list = ', '.join(f'"{c}"' for c in col_names)

        conn = db_connect()
        c = conn.cursor()
        try:
            c.execute('BEGIN TRANSACTION')
            c.execute(f'DELETE FROM "{table_name}"')
            c.executemany(f'INSERT INTO "{table_name}" ({col_list}) VALUES ({placeholders})', rows)
            conn.commit()
            count = len(rows)
        except Exception as e:
            conn.rollback()
            conn.close()
            return jsonify({'error': f'导入失败: {str(e)}'}), 500
        conn.close()

        return jsonify({'success': True, 'count': count, 'message': f'已清空并导入 {count} 条数据'})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'服务器内部错误: {str(e)}'}), 500


@app.route('/api/table/<table_name>/append', methods=['POST'])
def append_table_data(table_name):
    """追加数据：不清空，直接追加Excel数据"""
    try:
        allowed_tables = _get_all_table_names()
        if table_name not in allowed_tables:
            return jsonify({'error': f'表 "{table_name}" 不存在'}), 404

        if 'file' not in request.files:
            return jsonify({'error': '请上传Excel文件'}), 400

        file = request.files['file']
        if not file.filename.endswith(('.xlsx', '.xls')):
            return jsonify({'error': '仅支持 .xlsx 或 .xls 格式'}), 400

        columns = _get_table_columns(table_name)
        if not columns:
            return jsonify({'error': f'表 "{table_name}" 无可用列'}), 400

        file_data = file.read()
        try:
            rows, error = _parse_uploaded_excel(file_data, columns)
        except Exception as e:
            return jsonify({'error': f'文件解析失败: {str(e)}'}), 400
        if error:
            return jsonify({'error': error}), 400
        if not rows:
            return jsonify({'error': '文件中没有有效数据行'}), 400

        col_names = [c[0] for c in columns]
        placeholders = ', '.join(['?'] * len(col_names))
        col_list = ', '.join(f'"{c}"' for c in col_names)

        conn = db_connect()
        c = conn.cursor()
        try:
            c.execute('BEGIN TRANSACTION')
            c.executemany(f'INSERT INTO "{table_name}" ({col_list}) VALUES ({placeholders})', rows)
            conn.commit()
            count = len(rows)
        except Exception as e:
            conn.rollback()
            conn.close()
            return jsonify({'error': f'导入失败: {str(e)}'}), 500
        conn.close()

        return jsonify({'success': True, 'count': count, 'message': f'已追加 {count} 条数据'})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'服务器内部错误: {str(e)}'}), 500


# ============================================================
#  参数配置 API
# ============================================================

@app.route('/api/config/params', methods=['GET'])
def get_config_params():
    """获取配置参数：头程基数、资金利息"""
    try:
        conn = db_connect()
        c = conn.cursor()
        c.execute('SELECT key, value, updated_at FROM 配置参数 ORDER BY key')
        rows = c.fetchall()
        conn.close()
        result = {}
        for row in rows:
            result[row[0]] = {
                'value': row[1],
                'updated_at': row[2]
            }
        return jsonify({'params': result})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/config/params', methods=['POST'])
def update_config_params():
    """更新配置参数"""
    try:
        data = request.get_json(force=True)
    except Exception:
        return jsonify({'error': '请求体需为JSON'}), 400

    allowed_keys = {'first_leg_base', 'capital_interest'}
    updated = {}
    conn = db_connect()
    c = conn.cursor()
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    for key, val in data.items():
        if key not in allowed_keys:
            continue
        # 验证类型
        if key == 'first_leg_base':
            try:
                int_val = int(float(val))
                val = str(int_val)
            except (ValueError, TypeError):
                conn.close()
                return jsonify({'error': '头程基数必须是整数'}), 400
        elif key == 'capital_interest':
            try:
                float_val = float(val)
                if float_val < 0 or float_val > 100:
                    conn.close()
                    return jsonify({'error': '资金利息需在0~100之间'}), 400
                val = f'{float_val:.2f}'
            except (ValueError, TypeError):
                conn.close()
                return jsonify({'error': '资金利息必须是数字'}), 400

        c.execute('''
            INSERT INTO 配置参数 (key, value, updated_at) VALUES (?, ?, ?)
            ON CONFLICT(key) DO UPDATE SET value=excluded.value, updated_at=excluded.updated_at
        ''', (key, val, now))
        updated[key] = val

    conn.commit()
    conn.close()
    return jsonify({'success': True, 'updated': updated, 'time': now})


@app.route('/api/config/fedex/rates', methods=['GET'])
def get_fedex_rates():
    """获取FedEx报价（只读）：基础运费 + 超尺寸费"""
    try:
        conn = db_connect()
        c = conn.cursor()

        # 基础运费（默认查海智链）
        c.execute('SELECT weight_lb, zone2, zone3, zone4, zone5, zone6, zone7, zone8 FROM 美国基础运费 WHERE 仓库名称=? ORDER BY weight_lb', ('海智链',))
        base_rates = []
        for row in c.fetchall():
            base_rates.append({
                'weight_lb': row[0],
                'zone2': row[1],
                'zone3': row[2],
                'zone4': row[3],
                'zone5': row[4],
                'zone6': row[5],
                'zone7': row[6],
                'zone8': row[7],
            })

        # 超尺寸费（默认海智链）
        c.execute('SELECT charge_type, zone2, zone3, zone4, zone5, zone6, zone7, zone8 FROM 美国超规费用 WHERE 仓库名称=?', ('海智链',))
        oversize_rates = []
        for row in c.fetchall():
            oversize_rates.append({
                'charge_type': row[0],
                'zone2': row[1],
                'zone3': row[2],
                'zone4': row[3],
                'zone5': row[4],
                'zone6': row[5],
                'zone7': row[6],
                'zone8': row[7],
            })

        conn.close()
        return jsonify({
            'base_rates': base_rates,
            'oversize_rates': oversize_rates,
            'base_count': len(base_rates),
            'oversize_count': len(oversize_rates),
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/config/haizhi/fees', methods=['GET'])
def get_haizhi_fees():
    """获取海智链价卡五项费用数据"""
    try:
        conn = db_connect()
        c = conn.cursor()
        c.execute('''
            SELECT id, category, fee_name, condition_desc, condition_min, condition_max,
                   condition_unit, rate, unit, notes, sort_order, 仓库名称
            FROM 美国其他费用
            ORDER BY sort_order
        ''')
        rows = c.fetchall()
        conn.close()

        fees = []
        for row in rows:
            fees.append({
                'id': row[0],
                'category': row[1],
                'fee_name': row[2],
                'condition_desc': row[3],
                'condition_min': row[4],
                'condition_max': row[5],
                'condition_unit': row[6],
                'rate': row[7],
                'unit': row[8],
                'notes': row[9],
                'sort_order': row[10],
                '仓库名称': row[11],
            })

        return jsonify({'fees': fees, 'count': len(fees)})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/config/de/oversize', methods=['GET'])
def get_de_oversize():
    """获取德国超规费用（金仓/欧品居/易达云 × DPD/DHL/GLS 超规类型）"""
    try:
        conn = db_connect()
        c = conn.cursor()
        c.execute('SELECT * FROM 德国超规费用 ORDER BY id')
        rows = c.fetchall()
        conn.close()
        result = []
        for row in rows:
            result.append({
                'id': row[0], 'oversize_type': row[1], 'carrier': row[2],
                'description': row[3],
                '金仓': row[4], '欧品居': row[5], '易达云': row[6],
            })
        return jsonify({'oversize': result, 'count': len(result)})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/config/de/other-fees', methods=['GET'])
def get_de_other_fees():
    """获取德国其他费用（卸货/入库/出库/仓储 按仓库阶梯）"""
    try:
        conn = db_connect()
        c = conn.cursor()
        c.execute('SELECT * FROM 德国其他费用 ORDER BY sort_order')
        rows = c.fetchall()
        conn.close()
        fees = []
        for row in rows:
            fees.append({
                'id': row[0], 'category': row[1], 'fee_name': row[2],
                'condition_desc': row[3], 'condition_min': row[4],
                'condition_max': row[5], 'condition_unit': row[6],
                'rate': row[7], 'unit': row[8], 'notes': row[9],
                'sort_order': row[10], 'warehouse': row[11],
            })
        return jsonify({'fees': fees, 'count': len(fees)})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/config/de/rates', methods=['GET'])
def get_de_rates():
    """获取德国基础运费（金仓/欧品居/易达云 × DPD/DHL/GLS）"""
    try:
        conn = db_connect()
        c = conn.cursor()
        c.execute('SELECT * FROM 德国基础运费 ORDER BY weight_kg')
        rows = c.fetchall()
        conn.close()
        rates = []
        for row in rows:
            rates.append({
                'weight_kg': row[0],
                '金仓DPD': row[1], '金仓DHL': row[2], '金仓GLS': row[3],
                '欧品居DPD': row[4], '欧品居DHL': row[5], '欧品居GLS': row[6],
                '易达云DPD': row[7], '易达云DHL': row[8], '易达云GLS': row[9],
            })
        return jsonify({'rates': rates, 'count': len(rates)})
    except Exception as e:
        return jsonify({'error': str(e)}), 500




# ============================================================
#  尾程测算引擎
# ============================================================

import math
import re
import urllib.request
import urllib.error

CM_TO_INCH = 1 / 2.54
KG_TO_LB = 2.20462
DIM_FACTOR = 250  # 计抛系数

# reportlab 中文字体（PDF导出用，优先尝试系统字体，降级到内置CID字体）
_CN_FONT_REGISTERED = False
_CN_FONT_NAME = 'MicrosoftYaHei'

def _ensure_cn_font():
    global _CN_FONT_REGISTERED, _CN_FONT_NAME
    if _CN_FONT_REGISTERED:
        return
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    import os
    # 尝试注册系统微软雅黑
    font_path = r'C:\Windows\Fonts\msyh.ttc'
    if os.path.exists(font_path):
        try:
            pdfmetrics.registerFont(TTFont('MicrosoftYaHei', font_path, subfontIndex=0))
            _CN_FONT_NAME = 'MicrosoftYaHei'
            _CN_FONT_REGISTERED = True
            return
        except Exception:
            pass
    # 降级：尝试 simhei.ttf
    font_path2 = r'C:\Windows\Fonts\simhei.ttf'
    if os.path.exists(font_path2):
        try:
            pdfmetrics.registerFont(TTFont('SimHei', font_path2))
            _CN_FONT_NAME = 'SimHei'
            _CN_FONT_REGISTERED = True
            return
        except Exception:
            pass
    # 最终降级：使用 reportlab 内置 CID 字体
    from reportlab.pdfbase.cidfonts import UnicodeCIDFont
    pdfmetrics.registerFont(UnicodeCIDFont('STSong-Light'))
    _CN_FONT_NAME = 'STSong-Light'
    _CN_FONT_REGISTERED = True


def cm_to_inch(cm_val):
    """厘米转英寸"""
    return cm_val * CM_TO_INCH


def kg_to_lb(kg_val):
    """千克转磅"""
    return kg_val * KG_TO_LB


def cubic_cm_to_cbm(cm3_val):
    """立方厘米转立方米"""
    return cm3_val / 1000000.0


def round_up(val):
    """向上取整"""
    return math.ceil(val)


def calc_dimensional_weight(l_in, w_in, h_in):
    """计算材积重（磅）
    l_in, w_in, h_in: 英寸值（原始非取整）
    规则：各自向上取整后相乘，结果除以计抛系数，再向上取整
    """
    l = round_up(l_in)
    w = round_up(w_in)
    h = round_up(h_in)
    return round_up((l * w * h) / DIM_FACTOR)


def check_ahs_dimensions(l_in, w_in, h_in):
    """检查 AHS - Dimensions 条件
    返回: (是否触发, 详情dict)
    """
    dims = sorted([round_up(l_in), round_up(w_in), round_up(h_in)], reverse=True)
    longest = dims[0]
    second = dims[1]
    third = dims[2]
    side_perimeter = 2 * (second + third)
    volume = longest * second * third

    triggers = []
    if longest > 48:
        triggers.append(f'最长边 {longest}" > 48"')
    if second > 30:
        triggers.append(f'第二长边 {second}" > 30"')
    if longest + side_perimeter > 105:
        triggers.append(f'最长边+侧面周长 {longest + side_perimeter}" > 105"')
    if volume > 10368:
        triggers.append(f'体积 {volume} cu in > 10368')

    return len(triggers) > 0, {
        'longest': longest,
        'second': second,
        'third': third,
        'side_perimeter': side_perimeter,
        'longest_plus_perimeter': longest + side_perimeter,
        'volume_cu_in': volume,
        'triggers': triggers
    }


def check_oversize(l_in, w_in, h_in, actual_weight_lb):
    """检查 Oversize Charge 条件
    返回: (是否触发, 详情dict)
    """
    dims = sorted([round_up(l_in), round_up(w_in), round_up(h_in)], reverse=True)
    longest = dims[0]
    second = dims[1]
    third = dims[2]
    side_perimeter = 2 * (second + third)
    volume = longest * second * third

    triggers = []
    if longest > 96:
        triggers.append(f'最长边 {longest}" > 96"')
    if longest + side_perimeter > 130:
        triggers.append(f'最长边+侧面周长 {longest + side_perimeter}" > 130"')
    if volume > 17280:
        triggers.append(f'体积 {volume} cu in > 17280')
    if actual_weight_lb > 110:
        triggers.append(f'实重 {actual_weight_lb:.1f} lb > 110 lb')

    return len(triggers) > 0, {
        'longest': longest,
        'side_perimeter': side_perimeter,
        'longest_plus_perimeter': longest + side_perimeter,
        'volume_cu_in': volume,
        'actual_weight_lb': round(actual_weight_lb, 1),
        'triggers': triggers
    }


def lookup_base_rate(billing_weight_lb, zone, warehouse='海智链'):
    """查FedEx基础运费表（默认海智链，保持向后兼容）"""
    conn = db_connect()
    c = conn.cursor()
    lookup_w = round_up(billing_weight_lb)
    c.execute(
        'SELECT weight_lb, zone2, zone3, zone4, zone5, zone6, zone7, zone8 '
        'FROM 美国基础运费 WHERE 仓库名称=? AND weight_lb >= ? ORDER BY weight_lb LIMIT 1',
        (warehouse, lookup_w)
    )
    row = c.fetchone()
    conn.close()

    if not row:
        c2 = db_connect().cursor()
        c2.execute('SELECT MAX(weight_lb) FROM 美国基础运费 WHERE 仓库名称=?', (warehouse,))
        max_w = c2.fetchone()[0]
        c2.connection.close()
        c3 = db_connect().cursor()
        c3.execute(
            'SELECT weight_lb, zone2, zone3, zone4, zone5, zone6, zone7, zone8 '
            'FROM 美国基础运费 WHERE 仓库名称=? AND weight_lb = ?', (warehouse, max_w)
        )
        row = c3.fetchone()
        c3.connection.close()

    zone_col = f'zone{zone}'
    idx = ['weight_lb', 'zone2', 'zone3', 'zone4', 'zone5', 'zone6', 'zone7', 'zone8'].index(zone_col)
    return row[0], row[idx]


def lookup_oversize_rate(zone, fee_type, warehouse='海智链'):
    """查超尺寸费率（默认海智链，保持向后兼容）"""
    conn = db_connect()
    c = conn.cursor()
    zone_col = f'zone{zone}'
    c.execute(f'SELECT charge_type, {zone_col} FROM 美国超规费用 WHERE charge_type = ? AND 仓库名称=?', (fee_type, warehouse))
    row = c.fetchone()
    conn.close()
    return row[1] if row else 0.0


def lookup_haizhi_fee(category, condition_value=None, warehouse='海智链'):
    """查其他费用（按仓库过滤，默认海智链保持向后兼容）"""
    conn = db_connect()
    c = conn.cursor()
    c.execute(
        'SELECT rate, unit, condition_min, condition_max, condition_desc '
        'FROM 美国其他费用 WHERE category = ? AND 仓库名称 = ? ORDER BY sort_order',
        (category, warehouse)
    )
    rows = c.fetchall()
    conn.close()

    if not rows:
        return None

    # 固定费率（无阶梯）
    if category in ('residential_surcharge',):
        return {'rate': rows[0][0], 'unit': rows[0][1]}

    # 阶梯费率
    if condition_value is not None:
        for r in rows:
            rate, unit, cmin, cmax, cdesc = r
            if cmin is not None and condition_value >= cmin:
                if cmax is None or condition_value <= cmax:
                    return {'rate': rate, 'unit': unit, 'condition_desc': cdesc,
                            'condition_min': cmin, 'condition_max': cmax}
        # 未精确命中任何阶梯，找下一个更高的阶梯（处理阶梯间小数间隙）
        for r in rows:
            rate, unit, cmin, cmax, cdesc = r
            if cmin is not None and condition_value < cmin:
                return {'rate': rate, 'unit': unit, 'condition_desc': cdesc,
                        'condition_min': cmin, 'condition_max': cmax}
        # 超出所有范围，用最后一个
        return {'rate': rows[-1][0], 'unit': rows[-1][1],
                'condition_desc': rows[-1][4]}

    return {'rate': rows[0][0], 'unit': rows[0][1]}


def lookup_haizhi_fee_all(category, warehouse='海智链'):
    """查其他费用全部行（用于卸货费等多行场景）"""
    conn = db_connect()
    c = conn.cursor()
    c.execute(
        'SELECT rate, unit, condition_min, condition_max, condition_desc, fee_name '
        'FROM 美国其他费用 WHERE category = ? AND 仓库名称 = ? ORDER BY sort_order',
        (category, warehouse)
    )
    rows = c.fetchall()
    conn.close()
    return [{'rate': r[0], 'unit': r[1], 'condition_min': r[2],
             'condition_max': r[3], 'condition_desc': r[4], 'fee_name': r[5]} for r in rows]


def calc_cumulative_storage_fee(days, cbm, warehouse='海智链'):
    """计算累计仓储费
    按库龄天数，分段累进计算
    从数据库读取对应仓库的阶梯费率
    """
    # 从数据库读取该仓库的仓储费阶梯
    conn = db_connect()
    c = conn.cursor()
    c.execute(
        'SELECT condition_min, condition_max, rate FROM 美国其他费用 '
        'WHERE category = ? AND 仓库名称 = ? ORDER BY sort_order',
        ('storage_fee', warehouse)
    )
    db_rows = c.fetchall()
    conn.close()

    if not db_rows:
        # 降级：使用海智链默认阶梯
        tiers = [
            (0, 60, 0.0),
            (60, 90, 0.50),
            (90, 180, 0.80),
            (180, 360, 1.30),
            (360, float('inf'), 3.00),
        ]
    else:
        tiers = []
        for r in db_rows:
            tier_min = r[0] if r[0] is not None else 0
            tier_max = float('inf') if r[1] is None else r[1]
            rate_val = r[2] if r[2] is not None else 0.0
            tiers.append((tier_min, tier_max, rate_val))
    total_fee = 0.0
    breakdown = []

    remaining = days
    for tier_min, tier_max, rate in tiers:
        if remaining <= 0:
            break
        if rate == 0.0:
            days_in_tier = min(remaining, tier_max - tier_min)
            breakdown.append({
                'tier': f'{tier_min} < days ≤ {tier_max}天',
                'days': days_in_tier,
                'rate': 0.0,
                'fee': 0.0,
                'note': '免费'
            })
            remaining -= days_in_tier
            continue

        tier_range = tier_max - tier_min
        effective_tier_max = tier_max
        if tier_max == float('inf'):
            effective_tier_max = tier_min + remaining

        days_in_tier = min(remaining, tier_range) if tier_max != float('inf') else remaining
        fee = days_in_tier * rate * cbm
        total_fee += fee
        breakdown.append({
            'tier': f'{tier_min} < days ≤ {tier_max}天' if tier_max != float('inf') else f'days > {tier_min}天',
            'days': round(days_in_tier, 1),
            'rate': rate,
            'fee': round(fee, 2)
        })
        remaining -= days_in_tier

    return round(total_fee, 2), breakdown


def fetch_fuel_surcharge():
    """获取FedEx Ground燃油附加费率
    三级策略：
    1. 优先从数据库缓存读取（有效期7天），由定时任务通过WebFetch更新
    2. 尝试实时从FedEx官网抓取（需浏览器级User-Agent + SSL）
    3. 使用数据库缓存（即使过期）或默认值 27.25%
    """
    from datetime import datetime, timedelta
    import ssl

    # === 策略1：数据库缓存 ===
    try:
        conn = db_connect()
        c = conn.cursor()
        c.execute("SELECT value FROM 配置参数 WHERE key = 'fedex_fuel_rate'")
        row_rate = c.fetchone()
        c.execute("SELECT value FROM 配置参数 WHERE key = 'fedex_fuel_rate_updated'")
        row_date = c.fetchone()
        conn.close()

        if row_rate and row_rate[0]:
            cached_rate = float(row_rate[0])
            if row_date and row_date[0]:
                try:
                    updated = datetime.strptime(row_date[0].strip(), '%Y-%m-%d')
                    if datetime.now() - updated < timedelta(days=7):
                        print(f'[燃油费率] 使用数据库缓存 {cached_rate}%（{row_date[0]}更新）')
                        return cached_rate
                except ValueError:
                    pass
            # 缓存过期但仍有值，标记为降级使用
            print(f'[燃油费率] 数据库缓存已过期，尝试实时抓取（当前缓存 {cached_rate}%）')
            db_fallback = cached_rate
        else:
            db_fallback = None
    except Exception as e:
        print(f'[燃油费率] 读取数据库缓存失败: {e}')
        db_fallback = None

    # === 策略2：实时抓取 ===
    url = 'https://www.fedex.com/en-us/shipping/fuel-surcharge.html'
    try:
        # 使用完整浏览器请求头 + 忽略SSL验证，绕过FedEx WAF
        req = urllib.request.Request(url, headers={
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36',
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8',
            'Accept-Language': 'en-US,en;q=0.9',
            'Cache-Control': 'no-cache',
        })
        ctx = ssl.create_default_context()
        ctx.check_hostname = False
        ctx.verify_mode = ssl.CERT_NONE
        with urllib.request.urlopen(req, timeout=15, context=ctx) as resp:
            html = resp.read().decode('utf-8', errors='ignore')

        # 检查是否被WAF拦截
        if len(html) < 3000 and 'System Down' in html:
            raise Exception('FedEx WAF拦截')

        # 匹配 <td> 中的日期区间
        date_td_pattern = re.compile(
            r'<td[^>]*>\s*'
            r'((?:January|February|March|April|May|June|July|August|September|October|November|December|'
            r'Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\w*\s+\d{1,2},\s*\d{4})'
            r'\s*[–\-—\u2013\u2014]\s*'
            r'((?:January|February|March|April|May|June|July|August|September|October|November|December|'
            r'Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\w*\s+\d{1,2},\s*\d{4})'
            r'\s*</td>',
            re.IGNORECASE | re.DOTALL
        )

        today = datetime.now().date()
        date_formats = ['%B %d, %Y', '%b %d, %Y', '%B %d,%Y', '%b %d,%Y']
        matches = list(date_td_pattern.finditer(html))

        for match in matches:
            start_str = match.group(1).strip()
            end_str = match.group(2).strip()

            start_date = end_date = None
            for fmt in date_formats:
                try:
                    start_date = datetime.strptime(start_str, fmt).date()
                    end_date = datetime.strptime(end_str, fmt).date()
                    break
                except ValueError:
                    continue

            if start_date and end_date and start_date <= today <= end_date:
                after_html = html[match.end():]
                pct_match = re.search(
                    r'<td[^>]*>\s*(\d+\.?\d*)\s*%\s*</td>',
                    after_html, re.IGNORECASE
                )
                if pct_match:
                    rate = float(pct_match.group(1))
                    print(f'[燃油费率] 实时抓取成功 {start_str}-{end_str}，费率 {rate}%')
                    _cache_fuel_rate(rate)
                    return rate
                break

        # 降级：使用第一个日期区间的费率
        if matches:
            after_html = html[matches[0].end():]
            pct_match = re.search(
                r'<td[^>]*>\s*(\d+\.?\d*)\s*%\s*</td>',
                after_html, re.IGNORECASE
            )
            if pct_match:
                rate = float(pct_match.group(1))
                print(f'[燃油费率] 实时抓取（未匹配今日，用最新）费率 {rate}%')
                _cache_fuel_rate(rate)
                return rate

        # 备用正则
        alt_match = re.search(
            r'FedEx\s*Ground.*?(\d+\.?\d*)\s*%',
            html, re.DOTALL | re.IGNORECASE
        )
        if alt_match:
            rate = float(alt_match.group(1))
            print(f'[燃油费率] 实时抓取（备用正则）费率 {rate}%')
            _cache_fuel_rate(rate)
            return rate

        raise Exception('HTML解析失败，所有正则未匹配')
    except Exception as e:
        print(f'[燃油费率] 实时抓取失败: {e}')

    # === 策略3：数据库缓存降级 / 默认值 ===
    if db_fallback is not None:
        print(f'[燃油费率] 降级使用过期数据库缓存 {db_fallback}%')
        return db_fallback
    print('[燃油费率] 使用默认值 27.25%')
    return 27.25


def _cache_fuel_rate(rate):
    """将燃油费率写入数据库缓存"""
    try:
        from datetime import date
        conn = db_connect()
        c = conn.cursor()
        today_str = date.today().isoformat()
        c.execute("INSERT OR REPLACE INTO 配置参数 (key, value, updated_at) VALUES ('fedex_fuel_rate', ?, ?)",
                  (str(rate), today_str))
        c.execute("INSERT OR REPLACE INTO 配置参数 (key, value, updated_at) VALUES ('fedex_fuel_rate_updated', ?, ?)",
                  (today_str, today_str))
        conn.commit()
        conn.close()
        print(f'[燃油费率] 已缓存到数据库: {rate}%（{today_str}）')
    except Exception as e:
        print(f'[燃油费率] 缓存写入失败: {e}')


def calculate_single_item(item, zone, warehouse='海智链'):
    """计算单个SKU的尾程费用
    item: {sku, qty, length_cm, width_cm, height_cm, weight_kg}
    warehouse: 仓库名称（海智链/大健云仓/发现）
    """
    result = {
        'sku': item.get('sku', '-'),
        'qty': item.get('qty', 1),
    }

    # 单位换算
    l_cm = item['length_cm']
    w_cm = item['width_cm']
    h_cm = item['height_cm']
    wt_kg = item['weight_kg']

    l_in = cm_to_inch(l_cm)
    w_in = cm_to_inch(w_cm)
    h_in = cm_to_inch(h_cm)
    wt_lb = kg_to_lb(wt_kg)

    result['dims_cm'] = f'{l_cm:.1f}×{w_cm:.1f}×{h_cm:.1f}'
    result['dims_in'] = f'{l_in:.2f}×{w_in:.2f}×{h_in:.2f}'
    result['length_cm'] = round(l_cm, 1)
    result['width_cm'] = round(w_cm, 1)
    result['height_cm'] = round(h_cm, 1)
    result['weight_kg'] = round(wt_kg, 2)
    result['weight_lb'] = round(wt_lb, 2)

    # 向上取整后的尺寸
    l_ceil = round_up(l_in)
    w_ceil = round_up(w_in)
    h_ceil = round_up(h_in)
    result['dims_ceil_in'] = f'{l_ceil}×{w_ceil}×{h_ceil}'

    # 材积重
    dim_wt = calc_dimensional_weight(l_in, w_in, h_in)
    result['dim_weight_lb'] = dim_wt
    result['volume_cu_in'] = l_ceil * w_ceil * h_ceil
    result['volume_cbm'] = round(cubic_cm_to_cbm(l_cm * w_cm * h_cm), 6)

    # AHS - Dimensions 检查
    ahs_dim_triggered, ahs_dim_detail = check_ahs_dimensions(l_in, w_in, h_in)
    result['ahs_dimensions'] = {
        'triggered': ahs_dim_triggered,
        'detail': ahs_dim_detail
    }

    # AHS - Weight 检查
    ahs_weight_triggered = wt_lb > 50
    result['ahs_weight_triggered'] = ahs_weight_triggered

    # Oversize 检查
    os_triggered, os_detail = check_oversize(l_in, w_in, h_in, wt_lb)
    result['oversize'] = {
        'triggered': os_triggered,
        'detail': os_detail
    }
    # 超规信息（cm/kg 单位，方便前端展示）
    oversize_cm = []
    dims_cm_sorted = sorted([math.ceil(l_cm), math.ceil(w_cm), math.ceil(h_cm)], reverse=True)
    longest_cm = dims_cm_sorted[0]
    second_cm = dims_cm_sorted[1]
    third_cm = dims_cm_sorted[2]
    side_perimeter_cm = 2 * (second_cm + third_cm)
    # AHS-Dim 阈值: 最长边122cm(48"), 第二长边76cm(30"), 最长+侧面周长267cm(105"), 体积170L(10368cu in)
    if longest_cm > 122:
        oversize_cm.append(f'最长边{int(longest_cm)}cm > 122cm(超{int(longest_cm-122)}cm)')
    if second_cm > 76:
        oversize_cm.append(f'第二长边{int(second_cm)}cm > 76cm(超{int(second_cm-76)}cm)')
    if longest_cm + side_perimeter_cm > 267:
        oversize_cm.append(f'最长+周长{int(longest_cm+side_perimeter_cm)}cm > 267cm(超{int(longest_cm+side_perimeter_cm-267)}cm)')
    # Oversize 阈值: 最长边244cm(96"), 最长+周长330cm(130"), 实重50kg(110lb)
    if longest_cm > 244:
        oversize_cm.append(f'最长边{int(longest_cm)}cm > 244cm(超{int(longest_cm-244)}cm)')
    if longest_cm + side_perimeter_cm > 330:
        oversize_cm.append(f'最长+周长{int(longest_cm+side_perimeter_cm)}cm > 330cm(超{int(longest_cm+side_perimeter_cm-330)}cm)')
    if wt_kg > 50:
        oversize_cm.append(f'实重{wt_kg:.1f}kg > 50kg(超{wt_kg-50:.1f}kg)')
    # AHS-Weight 检查（cm/kg）：50lb ≈ 22.68kg
    if wt_kg > 22.68:
        oversize_cm.append(f'实重{wt_kg:.1f}kg > 22.68kg(AHS-Wt)(超{wt_kg-22.68:.1f}kg)')
    result['oversize_cm'] = oversize_cm

    # 计费重
    billing_wt = max(dim_wt, wt_lb)
    result['billing_weight_initial'] = round(billing_wt, 1)

    if ahs_dim_triggered:
        billing_wt = max(billing_wt, 40)
        result['billing_weight_after_ahs_dim'] = round(billing_wt, 1)

    if os_triggered:
        billing_wt = max(billing_wt, 90)
        result['billing_weight_after_oversize'] = round(billing_wt, 1)

    result['billing_weight_final'] = round(billing_wt, 1)

    # 计费重上限检查（150lb），超标标记为无法运送
    WEIGHT_LIMIT_US_LB = 150
    if billing_wt > WEIGHT_LIMIT_US_LB:
        result['exceeds_limit'] = True
        result['limit_reason'] = f'计费重 {billing_wt:.1f} lb 超过上限 {WEIGHT_LIMIT_US_LB} lb，无法运送'
        result['base_rate'] = 0
        result['base_freight'] = 0
        result['oversize_fee_type'] = None
        result['oversize_fee_rate'] = 0
        result['oversize_fee_amount'] = 0
        return result

    # 基础运费
    lookup_w, base_rate = lookup_base_rate(billing_wt, zone, warehouse)
    result['base_rate_lookup_weight'] = lookup_w
    result['base_rate'] = base_rate
    result['base_freight'] = round(base_rate * item['qty'], 2)

    # 超尺寸费
    applicable_oversize_fees = []
    if os_triggered:
        rate = lookup_oversize_rate(zone, 'Oversize', warehouse)
        applicable_oversize_fees.append(('Oversize', rate))
    elif ahs_dim_triggered:
        rate = lookup_oversize_rate(zone, 'AHS_Dimensions', warehouse)
        applicable_oversize_fees.append(('AHS_Dimensions', rate))
    elif ahs_weight_triggered:
        rate = lookup_oversize_rate(zone, 'AHS_Weight', warehouse)
        applicable_oversize_fees.append(('AHS_Weight', rate))

    # 如果多个触发，择大收取
    if applicable_oversize_fees:
        best = max(applicable_oversize_fees, key=lambda x: x[1])
        result['oversize_fee_type'] = best[0]
        result['oversize_fee_rate'] = best[1]
        result['oversize_fee_amount'] = round(best[1] * item['qty'], 2)
    else:
        result['oversize_fee_type'] = None
        result['oversize_fee_rate'] = 0
        result['oversize_fee_amount'] = 0

    return result


def calculate_last_mile(data):
    """完整的尾程测算"""
    items = data.get('items', [])
    zone = int(data.get('zone', 8))
    warehouse = data.get('warehouse', '海智链')
    unloading_type = data.get('unloading_type', 'none')  # 'none', 'fcl', 'lcl'
    storage_days = float(data.get('storage_days', 0))
    remote_mode = data.get('remote_surcharge_mode', 'none')  # 'none', 'estimate'
    peak_mode = data.get('peak_surcharge_mode', 'none')  # 'none', 'active'
    fuel_rate_pct_input = data.get('fuel_rate_pct')  # 用户手动输入或从接口获取的值

    if zone < 2 or zone > 8:
        return {'error': 'Zone必须在2-8之间'}

    # 处理每个item：查数据库或直接用输入值
    item_results = []
    total_base_freight = 0
    total_oversize_fee = 0
    total_residential = 0
    total_unloading = 0
    total_receiving = 0
    total_outbound = 0
    total_storage = 0
    total_cbm = 0
    valid_cbm = 0       # 合规SKU总体积（用于分摊）
    valid_qty = 0       # 合规SKU总件数（用于分摊）
    total_qty = 0
    max_actual_weight_lb = 0

    conn = db_connect()
    c = conn.cursor()

    for item in items:
        length_cm = item.get('length_cm')
        width_cm = item.get('width_cm')
        height_cm = item.get('height_cm')
        weight_kg = item.get('weight_kg')
        qty = int(item.get('qty', 1))
        sku = item.get('sku', '').strip()

        # 如果提供了SKU，从数据库读取尺寸
        if sku and (length_cm is None or weight_kg is None):
            c.execute(
                'SELECT 长度, 宽度, 高度, 重量 FROM 产品信息 WHERE SKU = ? LIMIT 1',
                (sku,)
            )
            row = c.fetchone()
            if row:
                if length_cm is None:
                    length_cm = row[0]
                if width_cm is None:
                    width_cm = row[1]
                if height_cm is None:
                    height_cm = row[2]
                if weight_kg is None:
                    weight_kg = row[3]

        if length_cm is None or width_cm is None or height_cm is None or weight_kg is None:
            conn.close()
            return {'error': f'SKU "{sku or "手动输入"}" 缺少长宽高或重量数据'}

        calc_item = {
            'sku': sku or '-',
            'qty': qty,
            'length_cm': float(length_cm),
            'width_cm': float(width_cm),
            'height_cm': float(height_cm),
            'weight_kg': float(weight_kg),
        }

        r = calculate_single_item(calc_item, zone, warehouse)
        if r.get('exceeds_limit'):
            # 超重无法运送，计入exceeded列表，不计入费用合计
            item_results.append(r)
            total_cbm += r['volume_cbm'] * qty
            total_qty += qty
            if r['weight_lb'] > max_actual_weight_lb:
                max_actual_weight_lb = r['weight_lb']
        else:
            item_results.append(r)
            total_base_freight += r['base_freight']
            total_oversize_fee += r['oversize_fee_amount']
            total_cbm += r['volume_cbm'] * qty
            valid_cbm += r['volume_cbm'] * qty
            total_qty += qty
            valid_qty += qty
            if r['weight_lb'] > max_actual_weight_lb:
                max_actual_weight_lb = r['weight_lb']

    conn.close()

    # 住宅附加费（每个包裹，仅合规SKU）
    residential_fee = lookup_haizhi_fee('residential_surcharge', warehouse=warehouse)
    total_residential = round(residential_fee['rate'] * valid_qty, 2) if residential_fee else 0

    # 卸货费（从数据库按仓库读取）
    unloading_detail = {}
    total_unloading = 0
    if unloading_type in ('fcl', 'lcl'):
        unloading_rows = lookup_haizhi_fee_all('unloading_fee', warehouse)
        fcl_rate = None
        lcl_rate = None
        for row in unloading_rows:
            name = row.get('fee_name', '')
            if '整柜' in name or 'FCL' in name.upper():
                fcl_rate = row['rate']
            elif '散货' in name or '散' in name or 'LCL' in name.upper():
                lcl_rate = row['rate']
        if unloading_type == 'fcl' and fcl_rate:
            unloading_fee_rate = fcl_rate / 65.0
            total_unloading = round(valid_cbm * unloading_fee_rate, 2)
            unloading_detail = {'type': '整柜/FCL', 'formula': f'{valid_cbm:.4f} m³ × {fcl_rate} / 65', 'fee': total_unloading}
        elif unloading_type == 'lcl' and lcl_rate:
            unloading_fee_rate = lcl_rate / 1.8
            total_unloading = round(valid_cbm * unloading_fee_rate, 2)
            unloading_detail = {'type': '散货/LCL', 'formula': f'{valid_cbm:.4f} m³ × {lcl_rate} / 1.8', 'fee': total_unloading}
    if not unloading_detail:
        unloading_detail = {'type': '无', 'fee': 0}

    # 入库上架费（每个产品按自身实重分别查阶梯，跳过超限项）
    total_receiving = 0
    receiving_details = []
    for r in item_results:
        if r.get('exceeds_limit'):
            r['receiving_rate'] = 0
            r['receiving_fee'] = 0
            continue
        item_wt_lb = r['weight_lb']
        fee_data = lookup_haizhi_fee('receiving_fee', item_wt_lb, warehouse)
        rate = fee_data['rate'] if fee_data else 0
        item_fee = round(rate * r['qty'], 2)
        total_receiving += item_fee
        r['receiving_rate'] = rate
        r['receiving_fee'] = item_fee
        receiving_details.append({
            'sku': r['sku'],
            'weight_lb': round(item_wt_lb, 1),
            'rate': rate,
            'fee': item_fee
        })
    total_receiving = round(total_receiving, 2)

    # 出库费（按仓库不同计费逻辑，仅合规SKU）
    total_outbound = 0
    if warehouse == '海智链':
        outbound_fee_data = lookup_haizhi_fee('outbound_fee', warehouse=warehouse)
        outbound_rate = outbound_fee_data['rate'] if outbound_fee_data else 1.0
        total_outbound = round(outbound_rate * valid_qty, 2)
    elif warehouse == '大健云仓':
        for r in item_results:
            if r.get('exceeds_limit'): continue
            wt_lb = math.ceil(r['weight_lb'])
            if wt_lb <= 20:
                fee = 1.2
            else:
                fee = 1.2 + (wt_lb - 20) * 0.05
            total_outbound += round(fee * r['qty'], 2)
    else:
        for r in item_results:
            if r.get('exceeds_limit'): continue
            item_wt_lb = r['weight_lb']
            fee_data = lookup_haizhi_fee('outbound_fee', item_wt_lb, warehouse)
            rate = fee_data['rate'] if fee_data else 0
            total_outbound += round(rate * r['qty'], 2)
    total_outbound = round(total_outbound, 2)

    # 仓储费（累计计算，仅合规SKU体积）
    valid_cbm_rounded = round(valid_cbm, 4)
    if storage_days > 0 and valid_cbm_rounded > 0:
        total_storage, storage_breakdown = calc_cumulative_storage_fee(storage_days, valid_cbm_rounded, warehouse)
    else:
        total_storage = 0
        storage_breakdown = []

    # 偏远附加费
    remote_fee = 0
    remote_detail = ''
    if remote_mode == 'estimate':
        remote_fee = round(total_base_freight * 0.09, 2)
        remote_detail = f'基础运费 × 9% = {total_base_freight} × 0.09'

    # 燃油附加费：优先使用用户传入的值，否则实时抓取
    if fuel_rate_pct_input is not None:
        try:
            fuel_rate_pct = float(fuel_rate_pct_input)
        except (ValueError, TypeError):
            return {'error': '燃油费率格式无效，请输入数字百分比'}
    else:
        fuel_rate_pct = fetch_fuel_surcharge()
    fuel_base = total_base_freight + total_oversize_fee + remote_fee + total_residential
    fuel_fee = round(fuel_base * fuel_rate_pct / 100.0, 2)

    # 旺季附加费
    peak_fee = 0
    peak_detail = ''
    if peak_mode == 'active':
        subtotal_before_peak = (
            total_base_freight + total_oversize_fee + total_residential +
            total_unloading + total_receiving + total_outbound +
            total_storage + remote_fee + fuel_fee
        )
        peak_fee = round(subtotal_before_peak * 0.011, 2)
        peak_detail = f'所有费用 × 1.1% = {subtotal_before_peak} × 0.011'

    # 总费用
    grand_total = round(
        total_base_freight + total_oversize_fee + total_residential +
        total_unloading + total_receiving + total_outbound +
        total_storage + remote_fee + fuel_fee + peak_fee, 2
    )

    # 统计超重SKU
    exceeded_count = sum(1 for r in item_results if r.get('exceeds_limit'))

    return {
        'zone': zone,
        'item_count': len(item_results),
        'exceeded_count': exceeded_count,
        'total_qty': total_qty,
        'total_cbm': round(total_cbm, 4),

        'items': item_results,

        'base_freight': total_base_freight,
        'oversize_fee': total_oversize_fee,
        'residential_surcharge': total_residential,
        'unloading_fee': total_unloading,
        'unloading_detail': unloading_detail,
        'receiving_fee': total_receiving,
        'receiving_details': receiving_details,
        'outbound_fee': total_outbound,
        'storage_fee': total_storage,
        'storage_days': storage_days,
        'storage_breakdown': storage_breakdown,
        'remote_surcharge': remote_fee,
        'remote_mode': remote_mode,
        'remote_detail': remote_detail,
        'fuel_surcharge': fuel_fee,
        'fuel_rate_pct': fuel_rate_pct,
        'fuel_formula': f'({total_base_freight} + {total_oversize_fee} + {remote_fee} + {total_residential}) × {fuel_rate_pct}%',
        'peak_surcharge': peak_fee,
        'peak_mode': peak_mode,
        'peak_detail': peak_detail,
        'grand_total': grand_total,
    }


# ============================================================
#  德国尾程测算引擎（金仓 × DPD/DHL/GLS）
# ============================================================

def check_de_oversize(l_cm, w_cm, h_cm, wt_kg, carrier):
    """检查德国超规条件，返回 (超规类型, [触发描述列表]) 或 (None, [])
    费率由调用方按仓库从数据库查询"""
    dims = sorted([l_cm, w_cm, h_cm], reverse=True)
    longest = dims[0]
    mid = dims[1]
    short = dims[2]
    perimeter = 2 * (mid + short)
    girth_length = longest + perimeter
    volume_liters = (l_cm * w_cm * h_cm) / 1000.0

    if carrier == 'DPD':
        triggers = []
        if wt_kg >= 40 or longest >= 250 or girth_length >= 330:
            if wt_kg >= 40: triggers.append(f'实重{wt_kg:.1f}kg≥40kg')
            if longest >= 250: triggers.append(f'最长边{longest:.0f}cm≥250cm')
            if girth_length >= 330: triggers.append(f'围长{girth_length:.0f}cm≥330cm')
            return '超规3', triggers
        triggers = []
        if 175 <= longest < 250 and 300 <= girth_length < 330:
            triggers.append(f'最长边{longest:.0f}cm≥175cm')
            triggers.append(f'围长{girth_length:.0f}cm≥300cm')
            return '超规2', triggers
        triggers = []
        rst = False
        if (120 <= longest < 175 or mid >= 60 or short >= 60) and girth_length < 300:
            if 120 <= longest < 175: triggers.append(f'最长边{longest:.0f}cm≥120cm')
            if mid >= 60: triggers.append(f'次长边{mid:.0f}cm≥60cm')
            if short >= 60: triggers.append(f'最短边{short:.0f}cm≥60cm')
            rst = True
        if volume_liters >= 150:
            triggers.append(f'体积{volume_liters:.0f}L≥150L')
            rst = True
        if rst:
            return '超规1', triggers

    elif carrier == 'DHL':
        triggers = []
        rst = False
        if longest >= 120:
            triggers.append(f'最长边{longest:.0f}cm≥120cm')
            rst = True
        if mid >= 60:
            triggers.append(f'次长边{mid:.0f}cm≥60cm')
            rst = True
        if short >= 60:
            triggers.append(f'最短边{short:.0f}cm≥60cm')
            rst = True
        if girth_length >= 300:
            triggers.append(f'围长{girth_length:.0f}cm≥300cm')
            rst = True
        if rst:
            return '超规1', triggers

    elif carrier == 'GLS':
        triggers = []
        if wt_kg >= 40 or girth_length >= 300 or longest >= 200 or mid >= 80 or short >= 60:
            if wt_kg >= 40: triggers.append(f'实重{wt_kg:.1f}kg≥40kg')
            if longest >= 200: triggers.append(f'最长边{longest:.0f}cm≥200cm')
            if mid >= 80: triggers.append(f'次长边{mid:.0f}cm≥80cm')
            if short >= 60: triggers.append(f'最短边{short:.0f}cm≥60cm')
            if girth_length >= 300: triggers.append(f'围长{girth_length:.0f}cm≥300cm')
            return '超规2', triggers
        if 120 < longest <= 150 and mid <= 80 and short <= 60:
            triggers.append(f'最长边{longest:.0f}cm>120cm')
            return '超规1', triggers

    return None, []


def get_de_oversize_fee(oversize_type, carrier, warehouse='金仓'):
    """按仓库+渠道查询超规费率"""
    if not oversize_type:
        return 0
    conn = db_connect()
    c = conn.cursor()
    c.execute('''SELECT [{wh}] FROM 德国超规费用 
                 WHERE oversize_type=? AND carrier=?'''.format(wh=warehouse),
              (oversize_type, carrier))
    row = c.fetchone()
    conn.close()
    if row and row[0] is not None:
        return float(row[0])
    return 0


def lookup_de_base_rate(weight_kg, carrier, warehouse='金仓'):
    """查德国基础运费：按仓库+渠道查表，取>=weight_kg的最小行"""
    conn = db_connect()
    c = conn.cursor()
    col = f'{warehouse}{carrier}'
    c.execute(f'SELECT weight_kg, [{col}] FROM 德国基础运费 WHERE [{col}] IS NOT NULL ORDER BY weight_kg')
    rows = c.fetchall()
    conn.close()
    if not rows:
        return 0, 0
    for r in rows:
        if r[0] >= weight_kg:
            return r[0], r[1]
    return rows[-1][0], rows[-1][1]


def calc_single_item_de(item, carrier, fuel_rate_pct=0, warehouse='金仓', dim_factor=None):
    """计算单个SKU在指定渠道下的费用
    dim_factor: 体积重除数，None=不计抛直接用实重，5000/6000=除以该值
    """
    l_cm = float(item['length_cm'])
    w_cm = float(item['width_cm'])
    h_cm = float(item['height_cm'])
    wt_kg = float(item['weight_kg'])
    qty = int(item.get('qty', 1))

    volume_cbm = round(l_cm * w_cm * h_cm / 1_000_000, 6)

    # 计费重：有dim_factor时取max(体积重/dim_factor, 实重)，否则用实重
    if dim_factor and dim_factor > 0:
        vol_wt_kg = round(l_cm * w_cm * h_cm / dim_factor, 2)
        billing_wt_kg = max(vol_wt_kg, wt_kg)
    else:
        vol_wt_kg = None
        billing_wt_kg = wt_kg

    # 计费重上限检查（40kg），超标标记为无法运送
    WEIGHT_LIMIT_DE_KG = 40
    if billing_wt_kg > WEIGHT_LIMIT_DE_KG:
        return {
            'carrier': carrier,
            'length_cm': round(l_cm, 1), 'width_cm': round(w_cm, 1), 'height_cm': round(h_cm, 1),
            'weight_kg': round(wt_kg, 2), 'volume_cbm': volume_cbm,
            'exceeds_limit': True,
            'limit_reason': f'计费重 {billing_wt_kg:.1f} kg 超过上限 {WEIGHT_LIMIT_DE_KG} kg，无法运送',
            'base_rate': 0, 'base_freight': 0,
            'oversize_type': None, 'oversize_fee_rate': 0, 'oversize_fee_amount': 0,
            'fuel_rate_pct': fuel_rate_pct, 'fuel_fee': 0,
        }

    # 基础运费
    lookup_w, base_rate = lookup_de_base_rate(billing_wt_kg, carrier, warehouse)
    base_freight = round(base_rate * qty, 2)

    # 超规检查（判定类型 → 按仓库查费率）
    os_type, os_triggers = check_de_oversize(l_cm, w_cm, h_cm, wt_kg, carrier)
    os_fee = get_de_oversize_fee(os_type, carrier, warehouse)
    oversize_amount = round(os_fee * qty, 2) if os_type else 0

    # 燃油附加费 = (基础运费 + 超规费) × 燃油费率%
    fuel_base = base_freight + oversize_amount
    fuel_fee = round(fuel_base * fuel_rate_pct / 100.0, 2)

    result = {
        'carrier': carrier,
        'length_cm': round(l_cm, 1), 'width_cm': round(w_cm, 1), 'height_cm': round(h_cm, 1),
        'weight_kg': round(wt_kg, 2),
        'volume_cbm': volume_cbm,
        'base_rate_lookup_kg': lookup_w,
        'base_rate': base_rate,
        'base_freight': base_freight,
        'oversize_type': os_type,
        'oversize_triggers': os_triggers,
        'oversize_fee_rate': os_fee,
        'oversize_fee_amount': oversize_amount,
        'fuel_rate_pct': fuel_rate_pct,
        'fuel_fee': fuel_fee,
    }
    if vol_wt_kg is not None:
        result['vol_weight_kg'] = vol_wt_kg
        result['billing_weight_kg'] = billing_wt_kg
    return result


def calc_last_mile_de(data):
    """德国尾程测算 — 金仓/欧品居/易达云 DPD/DHL/GLS 三渠道取最低
    DIM除数：金仓仅GLS=6000，欧品居全渠道=5000，易达云全渠道=6000
    """
    items = data.get('items', [])
    warehouse = data.get('warehouse', '金仓')
    unloading_type = data.get('unloading_type', 'none')
    storage_days = float(data.get('storage_days', 0))
    peak_mode = data.get('peak_mode', 'none')
    # 燃油附加费率：每个仓库×每个渠道独立输入（共9个参数）
    # 参数名格式: fuel_rate_{仓库key}_{渠道小写}
    WH_KEY_MAP = {'金仓': 'jincang', '欧品居': 'oupinju', '易达云': 'yidayun'}
    wh_key_map_reverse = {v: k for k, v in WH_KEY_MAP.items()}
    wh_key = WH_KEY_MAP.get(warehouse, 'jincang')

    # 读取所有9个燃油附加费率（新格式），同时兼容旧格式（3个全局参数）
    fuel_rates_all = {}
    for wk in ['jincang', 'oupinju', 'yidayun']:
        fuel_rates_all[wk] = {}
        for ck in ['dpd', 'dhl', 'gls']:
            param_name = f'fuel_rate_{wk}_{ck}'
            fuel_rates_all[wk][ck] = float(data.get(param_name, 0))

    # 检查是否使用了新格式（至少有1个非零值）
    has_new_format = any(
        fuel_rates_all[wk][ck] != 0
        for wk in ['jincang', 'oupinju', 'yidayun']
        for ck in ['dpd', 'dhl', 'gls']
    )

    if has_new_format:
        # 使用新格式：按仓库取对应费率
        fuel_rate_dpd = fuel_rates_all[wh_key]['dpd']
        fuel_rate_dhl = fuel_rates_all[wh_key]['dhl']
        fuel_rate_gls = fuel_rates_all[wh_key]['gls']
    else:
        # 兼容旧格式：3个全局参数
        fuel_rate_dpd = float(data.get('fuel_rate_dpd', 0))
        fuel_rate_dhl = float(data.get('fuel_rate_dhl', 0))
        fuel_rate_gls = float(data.get('fuel_rate_gls', 0))

    # DIM除数映射（体积重 = 长×宽×高 / dim_factor）
    DIM_FACTOR_MAP = {
        '金仓': {'DPD': None, 'DHL': None, 'GLS': 6000},
        '欧品居': {'DPD': 5000, 'DHL': 5000, 'GLS': 5000},
        '易达云': {'DPD': 6000, 'DHL': 6000, 'GLS': 6000},
    }
    dim_map = DIM_FACTOR_MAP.get(warehouse, DIM_FACTOR_MAP['金仓'])

    if not items:
        return {'error': '请至少添加一个产品'}

    carriers = ['DPD', 'DHL', 'GLS']
    fuel_rates = {'DPD': fuel_rate_dpd, 'DHL': fuel_rate_dhl, 'GLS': fuel_rate_gls}
    item_results = []
    total_base_freight = 0
    total_oversize_fee = 0
    total_fuel_fee = 0
    total_qty = 0
    total_cbm = 0
    valid_qty = 0       # 合规SKU件数
    valid_cbm = 0       # 合规SKU体积

    # 按渠道汇总（所有SKU各渠道的费用合计）
    de_carrier_totals = {c: {'base': 0, 'oversize': 0, 'fuel': 0} for c in carriers}

    for item in items:
        sku = item.get('sku', '-')
        qty = int(item.get('qty', 1))
        length_cm = item.get('length_cm')
        width_cm = item.get('width_cm')
        height_cm = item.get('height_cm')
        weight_kg = item.get('weight_kg')

        if length_cm is None or weight_kg is None:
            conn2 = db_connect()
            c2 = conn2.cursor()
            c2.execute('SELECT 长度, 宽度, 高度, 重量 FROM 产品信息 WHERE SKU = ? LIMIT 1', (sku,))
            row = c2.fetchone()
            conn2.close()
            if row:
                if length_cm is None: length_cm = row[0]
                if width_cm is None: width_cm = row[1]
                if height_cm is None: height_cm = row[2]
                if weight_kg is None: weight_kg = row[3]

        if length_cm is None or width_cm is None or height_cm is None or weight_kg is None:
            return {'error': f'SKU "{sku or "手动输入"}" 缺少长宽高或重量数据'}

        base_item = {'sku': sku, 'qty': qty, 'length_cm': float(length_cm),
                     'width_cm': float(width_cm), 'height_cm': float(height_cm),
                     'weight_kg': float(weight_kg)}

        # 计算每个渠道的费用（含燃油附加费，按仓库DIM规则）
        carrier_costs = {}
        for carrier in carriers:
            cr = calc_single_item_de(base_item, carrier, fuel_rates[carrier], warehouse, dim_map[carrier])
            cr_total = cr['base_freight'] + cr['oversize_fee_amount'] + cr['fuel_fee']
            carrier_costs[carrier] = {'result': cr, 'total': cr_total}

        # 检查第一个渠道是否超重（所有渠道共用同一计费重）
        first_cr = carrier_costs[carriers[0]]['result']
        if first_cr.get('exceeds_limit'):
            # 超重无法运送，仅记录警告信息
            first_cr['sku'] = sku
            first_cr['qty'] = qty
            item_results.append(first_cr)
            total_qty += qty
            total_cbm += first_cr['volume_cbm'] * qty
            continue

        # 选最优
        best = min(carrier_costs.items(), key=lambda x: x[1]['total'])
        best_carrier = best[0]
        best_result = best[1]['result']
        best_total = best[1]['total']

        best_result['sku'] = sku
        best_result['qty'] = qty
        best_result['all_carriers'] = {c: {'base': v['result']['base_freight'],
                                            'oversize': v['result']['oversize_fee_amount'],
                                            'fuel': v['result']['fuel_fee'],
                                            'fuel_rate': v['result']['fuel_rate_pct'],
                                            'total': v['total']} for c, v in carrier_costs.items()}
        item_results.append(best_result)
        total_base_freight += best_result['base_freight']
        total_oversize_fee += best_result['oversize_fee_amount']
        total_fuel_fee += best_result['fuel_fee']
        total_qty += qty
        total_cbm += best_result['volume_cbm'] * qty
        valid_qty += qty
        valid_cbm += best_result['volume_cbm'] * qty

        # 各渠道独立汇总（DE）
        for c in carriers:
            cc = carrier_costs.get(c)
            if cc and not cc['result'].get('exceeds_limit'):
                de_carrier_totals[c]['base'] += round(cc['result']['base_freight'] * qty, 2)
                de_carrier_totals[c]['oversize'] += round(cc['result']['oversize_fee_amount'] * qty, 2)
                de_carrier_totals[c]['fuel'] += round(cc['result']['fuel_fee'] * qty, 2)

    total_cbm = round(total_cbm, 4)
    valid_cbm = round(valid_cbm, 4)

    # -- 仓库费用（金仓，从德国其他费用表查）--
    conn = db_connect()
    c = conn.cursor()

    # 卸货费
    unloading_fee = 0
    unloading_detail = {'type': '无', 'fee': 0}
    if unloading_type in ('fcl', 'lcl'):
        c.execute("SELECT fee_name, rate FROM 德国其他费用 WHERE category='unloading_fee' AND 仓库名称=? ORDER BY sort_order", (warehouse,))
        urows = c.fetchall()
        fcl_rate = next((r[1] for r in urows if '整柜' in str(r[0]) or 'FCL' in str(r[0]).upper()), None)
        lcl_rate = next((r[1] for r in urows if '散货' in str(r[0]) or 'LCL' in str(r[0]).upper()), None)
        if unloading_type == 'fcl' and fcl_rate:
            unloading_fee = round(valid_cbm * fcl_rate / 65.0, 2)
            unloading_detail = {'type': '整柜/FCL', 'fee': unloading_fee}
        elif unloading_type == 'lcl' and lcl_rate:
            unloading_fee = round(valid_cbm * lcl_rate / 1.8, 2)
            unloading_detail = {'type': '散货/LCL', 'fee': unloading_fee}

    # 入库费（按实重阶梯）
    total_receiving = 0
    c.execute("SELECT condition_min, condition_max, rate FROM 德国其他费用 WHERE category='receiving_fee' AND 仓库名称=? ORDER BY sort_order", (warehouse,))
    recv_tiers = [(r[0] or 0, float('inf') if r[1] is None else r[1], r[2]) for r in c.fetchall()]
    for r in item_results:
        if r.get('exceeds_limit'): continue
        item_wt = r['weight_kg']
        rate = 0
        for tmin, tmax, trate in recv_tiers:
            if item_wt >= tmin and item_wt <= tmax:
                rate = trate
                break
        total_receiving += round(rate * r['qty'], 2)
    total_receiving = round(total_receiving, 2)

    # 出库费（按实重阶梯）
    total_outbound = 0
    c.execute("SELECT condition_min, condition_max, rate, notes FROM 德国其他费用 WHERE category='outbound_fee' AND 仓库名称=? ORDER BY sort_order", (warehouse,))
    ob_rows = c.fetchall()
    for r in item_results:
        if r.get('exceeds_limit'): continue
        item_wt = math.ceil(r['weight_kg'])
        rate = 0
        for omin, omax, orate, onotes in ob_rows:
            tmin = omin or 0
            tmax = float('inf') if omax is None else omax
            if item_wt >= tmin and item_wt <= tmax:
                rate = orate
                # 金仓特殊公式：≥16kg时基础1.4，每超15后每kg+0.1
                if onotes and '每超' in str(onotes) and item_wt > 15:
                    rate = 1.4 + (item_wt - 15) * 0.1
                break
        total_outbound += round(rate * r['qty'], 2)
    total_outbound = round(total_outbound, 2)

    conn.close()

    # 仓储费（累进）
    total_storage = 0
    storage_breakdown = []
    if storage_days > 0 and valid_cbm > 0:
        conn2 = db_connect()
        c2 = conn2.cursor()
        c2.execute("SELECT condition_min, condition_max, rate FROM 德国其他费用 WHERE category='storage_fee' AND 仓库名称=? ORDER BY sort_order", (warehouse,))
        st_rows = c2.fetchall()
        conn2.close()
        if st_rows:
            tiers = []
            for sr in st_rows:
                tmin_raw = sr[0] or 0
                tmin = (tmin_raw - 1) if tmin_raw > 0 else 0  # DB存1索引，转0索引
                tmax = float('inf') if sr[1] is None else sr[1]
                trate = sr[2] or 0
                tiers.append((tmin, tmax, trate, tmin_raw))
            remaining = storage_days
            for tmin, tmax, trate, tmin_raw in tiers:
                if remaining <= 0:
                    break
                days_in_tier = min(remaining, tmax - tmin)
                fee_in_tier = round(days_in_tier * valid_cbm * trate, 2)
                if trate == 0:
                    storage_breakdown.append({'tier': f'{tmin_raw}-{tmax if tmax != float("inf") else "+"}天', 'days': days_in_tier, 'rate': 0, 'fee': 0, 'note': '免费'})
                else:
                    storage_breakdown.append({'tier': f'{tmin_raw}-{tmax if tmax != float("inf") else "+"}天', 'days': days_in_tier, 'rate': trate, 'fee': fee_in_tier})
                total_storage += fee_in_tier
                remaining -= days_in_tier
    total_storage = round(total_storage, 2)

    # 旺季附加费
    peak_fee = 0
    if peak_mode == 'active':
        subtotal = total_base_freight + total_oversize_fee + total_fuel_fee + unloading_fee + total_receiving + total_outbound + total_storage
        peak_fee = round(subtotal * 0.005, 2)

    # 总费用
    grand_total = round(total_base_freight + total_oversize_fee + total_fuel_fee + unloading_fee +
                        total_receiving + total_outbound + total_storage + peak_fee, 2)

    # 统计超重SKU
    exceeded_count = sum(1 for r in item_results if r.get('exceeds_limit'))

    return {
        'items': item_results,
        'item_count': len(item_results),
        'exceeded_count': exceeded_count,
        'total_qty': total_qty,
        'total_cbm': total_cbm,
        'base_freight': total_base_freight,
        'oversize_fee': total_oversize_fee,
        'unloading_fee': unloading_fee,
        'unloading_detail': unloading_detail,
        'receiving_fee': total_receiving,
        'outbound_fee': total_outbound,
        'storage_fee': total_storage,
        'storage_days': storage_days,
        'storage_breakdown': storage_breakdown,
        'fuel_fee': total_fuel_fee,
        'fuel_rate_dpd': fuel_rate_dpd,
        'fuel_rate_dhl': fuel_rate_dhl,
        'fuel_rate_gls': fuel_rate_gls,
        # 新增：返回所有9个仓库×渠道的燃油附加费率
        'fuel_rates_all': {
            'jincang': {'dpd': fuel_rates_all.get('jincang', {}).get('dpd', 0),
                        'dhl': fuel_rates_all.get('jincang', {}).get('dhl', 0),
                        'gls': fuel_rates_all.get('jincang', {}).get('gls', 0)},
            'oupinju': {'dpd': fuel_rates_all.get('oupinju', {}).get('dpd', 0),
                        'dhl': fuel_rates_all.get('oupinju', {}).get('dhl', 0),
                        'gls': fuel_rates_all.get('oupinju', {}).get('gls', 0)},
            'yidayun': {'dpd': fuel_rates_all.get('yidayun', {}).get('dpd', 0),
                        'dhl': fuel_rates_all.get('yidayun', {}).get('dhl', 0),
                        'gls': fuel_rates_all.get('yidayun', {}).get('gls', 0)},
        },
        'peak_surcharge': peak_fee,
        'peak_mode': peak_mode,
        'grand_total': grand_total,
        'carrier_totals': {c: dict(t) for c, t in de_carrier_totals.items()},
        'best_carrier': min(de_carrier_totals, key=lambda c: de_carrier_totals[c]['base'] + de_carrier_totals[c]['oversize'] + de_carrier_totals[c]['fuel']),
    }


# ============================================================
#  德国利润测算引擎
# ============================================================

def fetch_eur_cny_rate():
    """获取欧元兑人民币汇率（从shishihuilv.com/EUR-CNY-1.html爬取）
    HTML结构：<p class="to">7.8199216 人民币（CNY）</p>
    """
    import urllib.request
    import urllib.error
    import re
    try:
        url = 'https://www.shishihuilv.com/EUR-CNY-1.html'
        req = urllib.request.Request(url, headers={
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
            'Accept-Language': 'zh-CN,zh;q=0.9,en;q=0.8',
        })
        with urllib.request.urlopen(req, timeout=15) as resp:
            html = resp.read().decode('utf-8', errors='ignore')
            # 精确匹配 <p class="to">数字 人民币（CNY）</p>
            m = re.search(r'<p\s+class="to">\s*([\d.]+)\s*人民币', html)
            if not m:
                # 备用：匹配所有 class="to" 的 <p> 标签
                m = re.search(r'class="to"[^>]*>\s*([\d.]+)', html)
            if m:
                rate = float(m.group(1))
                if 5 < rate < 12:
                    return round(rate, 4)
    except Exception:
        pass
    try:
        url = 'https://open.er-api.com/v6/latest/EUR'
        req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
        with urllib.request.urlopen(req, timeout=10) as resp:
            data = json.loads(resp.read().decode('utf-8'))
            rate = data.get('rates', {}).get('CNY')
            if rate and rate > 0:
                return round(rate, 4)
    except Exception:
        pass
    return 7.80


def calculate_profit_de(data):
    """德国利润测算：与US一致框架 + 16%VAT"""
    items = data.get('items', [])
    exchange_rate = float(data.get('exchange_rate', 7.80))
    tax_per_cbm = float(data.get('tax_per_cbm', data.get('tax_total', 0)))
    management_rate = float(data.get('management_rate', 0.04))
    complaint_rate = float(data.get('complaint_rate', 0.05))
    platform_rate = float(data.get('platform_rate', 0.15))
    ad_rate = float(data.get('ad_rate', 0.15))
    profit_rate = float(data.get('profit_rate', 0.30))
    vat_rate = 0.16  # 德国VAT固定16%

    if not items:
        return {'error': '请至少添加一个产品'}
    if exchange_rate <= 0:
        return {'error': '汇率必须大于0'}

    total_rates = platform_rate + ad_rate + profit_rate + vat_rate
    if total_rates >= 1:
        return {'error': f'平台({platform_rate*100}%)+广告({ad_rate*100}%)+利润({profit_rate*100}%)+VAT(16%) = {total_rates*100}% 超过100%'}

    conn = db_connect()
    c = conn.cursor()
    c.execute("SELECT value FROM 配置参数 WHERE key = 'first_leg_base'")
    row = c.fetchone()
    first_leg_base = int(float(row[0])) if row else 600
    conn.close()

    missing_cost_skus = []
    for item in items:
        sku = item.get('sku', '-')
        purchase_price = item.get('purchase_price_manual')
        if purchase_price is None and sku and sku != '-':
            try:
                conn2 = db_connect()
                c2 = conn2.cursor()
                c2.execute('SELECT 采购价 FROM 产品信息 WHERE SKU = ? LIMIT 1', (sku,))
                row2 = c2.fetchone()
                conn2.close()
                if row2 and row2[0]:
                    purchase_price = float(row2[0])
            except Exception:
                pass
        if not purchase_price or purchase_price <= 0:
            missing_cost_skus.append(sku)
        item['_purchase_price'] = float(purchase_price) if purchase_price else 0.0

    if missing_cost_skus:
        return {'error': f'以下产品缺少采购成本：{", ".join(missing_cost_skus)}'}

    total_cbm = sum(item.get('volume_cbm', 0) * item.get('qty', 1) for item in items)
    if total_cbm <= 0:
        total_cbm = 1

    item_results = []
    for item in items:
        sku = item.get('sku', '-')
        qty = int(item.get('qty', 1))
        volume_cbm = float(item.get('volume_cbm', 0))
        purchase_price = item.get('_purchase_price', 0)
        purchase_cost = round(float(purchase_price) * qty / exchange_rate, 2)
        head_leg_cost = round(first_leg_base * volume_cbm * qty / exchange_rate, 2)
        tax_cost = round(tax_per_cbm * volume_cbm * qty, 2) if tax_per_cbm > 0 else 0
        last_mile_cost = round(float(item.get('last_mile_cost_excl_storage', 0)) / qty, 2)
        storage_cost = round(float(item.get('storage_cost', 0)) / qty, 2)
        subtotal_5 = purchase_cost + head_leg_cost + tax_cost + last_mile_cost + storage_cost
        management_cost = round(management_rate * subtotal_5, 2)
        complaint_cost = round(complaint_rate * subtotal_5, 2)
        fixed_cost = purchase_cost + head_leg_cost + tax_cost + last_mile_cost + storage_cost + management_cost + complaint_cost

        denominator = 1 - platform_rate - ad_rate - profit_rate - vat_rate
        price = round(fixed_cost / denominator, 2) if denominator > 0 else 0
        profit = round(price * profit_rate, 2)
        vat_amount = round(price * vat_rate, 2)
        calc_profit_rate = round(profit / price, 4) if price > 0 else 0

        item_results.append({
            'sku': sku, 'qty': qty, 'volume_cbm': round(volume_cbm, 4),
            'purchase_cost': purchase_cost, 'purchase_price': purchase_price,
            'head_leg_cost': head_leg_cost, 'tax_cost': tax_cost,
            'last_mile_cost': last_mile_cost, 'storage_cost': storage_cost,
            'management_cost': management_cost, 'complaint_cost': complaint_cost,
            'vat_amount': vat_amount,
            'subtotal_5': subtotal_5, 'fixed_cost': fixed_cost,
            'price': price, 'profit': profit, 'profit_rate': calc_profit_rate,
        })

    summary = {
        'purchase_cost': sum(r['purchase_cost'] for r in item_results),
        'head_leg_cost': sum(r['head_leg_cost'] for r in item_results),
        'tax_cost': sum(r['tax_cost'] for r in item_results),
        'last_mile_cost': sum(r['last_mile_cost'] for r in item_results),
        'storage_cost': sum(r['storage_cost'] for r in item_results),
        'management_cost': sum(r['management_cost'] for r in item_results),
        'complaint_cost': sum(r['complaint_cost'] for r in item_results),
        'vat_amount': sum(r['vat_amount'] for r in item_results),
        'fixed_cost': sum(r['fixed_cost'] for r in item_results),
        'price': sum(r['price'] for r in item_results),
        'profit': sum(r['profit'] for r in item_results),
    }
    summary['profit_rate'] = round(summary['profit'] / summary['price'], 4) if summary['price'] > 0 else 0

    return {
        'items': item_results,
        'total_qty': sum(r['qty'] for r in item_results),
        'total_cbm': round(total_cbm, 4),
        'exchange_rate': exchange_rate,
        'first_leg_base': first_leg_base,
        'tax_per_cbm': tax_per_cbm,
        'management_rate': round(management_rate * 100, 1),
        'complaint_rate': round(complaint_rate * 100, 1),
        'platform_rate': round(platform_rate * 100, 1),
        'ad_rate': round(ad_rate * 100, 1),
        'target_profit_rate': round(profit_rate * 100, 1),
        'vat_rate': 16,
        'summary': summary,
    }


# ============================================================
#  德国 API 端点
# ============================================================

@app.route('/api/calc/de/last-mile', methods=['POST'])
def api_calc_last_mile_de():
    """德国尾程测算"""
    try:
        data = request.get_json(force=True)
        if not data:
            return jsonify({'error': '请提供测算数据'}), 400
        result = calc_last_mile_de(data)
        if 'error' in result:
            return jsonify(result), 400
        return jsonify(result)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/calc/de/exchange-rate', methods=['GET'])
def api_get_eur_rate():
    """获取欧元兑人民币汇率"""
    try:
        rate = fetch_eur_cny_rate()
        return jsonify({'rate': rate, 'currency_pair': 'EUR/CNY', 'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/calc/de/profit', methods=['POST'])
def api_calc_profit_de():
    """德国利润测算"""
    try:
        data = request.get_json(force=True)
        if not data:
            return jsonify({'error': '请提供测算数据'}), 400
        result = calculate_profit_de(data)
        if 'error' in result:
            return jsonify(result), 400
        return jsonify(result)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ============================================================
#  英国尾程测算引擎（易达云 × 5渠道 / 大健云仓 × 3渠道）
# ============================================================

UK_WAREHOUSES = ['易达云', '大健云仓']

# 易达云渠道列表（5个）
UK_YIDAYUN_CARRIERS = ['Royalmail', 'Yodel', 'Hermes', 'Parcelforce', 'DPD']
# 大健云仓渠道列表（3个）
UK_GJG_CARRIERS = ['Hermes', 'Whistl', 'TNT']


def check_uk_oversize(l_cm, w_cm, h_cm, wt_kg, carrier, warehouse):
    """检查英国超规条件，返回 (超规类型, [触发描述]) 或 (None, [])"""
    dims = sorted([l_cm, w_cm, h_cm], reverse=True)
    longest = dims[0]
    mid = dims[1]
    short = dims[2]
    perimeter = 2 * (mid + short)
    girth_length = longest + perimeter
    volume_cbm = round(l_cm * w_cm * h_cm / 1_000_000, 6)

    if warehouse == '易达云':
        if carrier == 'Royalmail':
            triggers = []
            if wt_kg > 20: triggers.append(f'实重{wt_kg:.1f}kg>20kg')
            if l_cm > 61: triggers.append(f'长{l_cm:.0f}cm>61cm')
            if w_cm > 46: triggers.append(f'宽{w_cm:.0f}cm>46cm')
            if h_cm > 46: triggers.append(f'高{h_cm:.0f}cm>46cm')
            if volume_cbm > 0.129: triggers.append(f'体积{volume_cbm:.3f}m³>0.129m³')
            if triggers: return '超规1', triggers

        elif carrier == 'Yodel':
            if longest > 170 or volume_cbm > 0.28:
                triggers = []
                if longest > 170: triggers.append(f'最长边{longest:.0f}cm>170cm')
                if volume_cbm > 0.28: triggers.append(f'体积{volume_cbm:.3f}m³>0.28m³')
                return '超规3', triggers
            if (120 < longest <= 170) or (0.23 < volume_cbm <= 0.28):
                triggers = []
                if 120 < longest <= 170: triggers.append(f'最长边{longest:.0f}cm>120cm')
                if 0.23 < volume_cbm <= 0.28: triggers.append(f'体积{volume_cbm:.3f}m³>0.23m³')
                return '超规2', triggers
            if (90 < longest <= 120) or (0.113 < volume_cbm <= 0.23):
                triggers = []
                if 90 < longest <= 120: triggers.append(f'最长边{longest:.0f}cm>90cm')
                if 0.113 < volume_cbm <= 0.23: triggers.append(f'体积{volume_cbm:.3f}m³>0.113m³')
                return '超规1', triggers

        elif carrier == 'Hermes':
            if volume_cbm >= 0.04:
                return '超规1', [f'体积{volume_cbm:.3f}m³≥0.04m³']

        elif carrier == 'Parcelforce':
            triggers = []
            if wt_kg > 30: triggers.append(f'实重{wt_kg:.1f}kg>30kg')
            if longest > 150: triggers.append(f'最长边{longest:.0f}cm>150cm')
            if girth_length > 300: triggers.append(f'围长{girth_length:.0f}cm>300cm')
            if volume_cbm > 0.22: triggers.append(f'体积{volume_cbm:.3f}m³>0.22m³')
            if triggers: return '超规1', triggers

        elif carrier == 'DPD':
            triggers = []
            if longest > 100: triggers.append(f'最长边{longest:.0f}cm>100cm')
            if mid > 70: triggers.append(f'次长边{mid:.0f}cm>70cm')
            if short > 60: triggers.append(f'最短边{short:.0f}cm>60cm')
            if (l_cm + w_cm + h_cm) > 230: triggers.append(f'长宽高和{l_cm+w_cm+h_cm:.0f}cm>230cm')
            if wt_kg > 30: triggers.append(f'实重{wt_kg:.1f}kg>30kg')
            if triggers: return '超规1', triggers

    elif warehouse == '大健云仓':
        if carrier == 'Whistl':
            if wt_kg > 30 or longest > 170 or volume_cbm > 0.28:
                triggers = []
                if wt_kg > 30: triggers.append(f'实重{wt_kg:.1f}kg>30kg')
                if longest > 170: triggers.append(f'最长边{longest:.0f}cm>170cm')
                if volume_cbm > 0.28: triggers.append(f'体积{volume_cbm:.3f}m³>0.28m³')
                return '超规3', triggers
            if (120 < longest < 170) or (0.23 < volume_cbm < 0.28):
                triggers = []
                if 120 < longest < 170: triggers.append(f'最长边{longest:.0f}cm>120cm')
                if 0.23 < volume_cbm < 0.28: triggers.append(f'体积{volume_cbm:.3f}m³>0.23m³')
                return '超规2', triggers
            triggers = []
            if wt_kg > 17: triggers.append(f'实重{wt_kg:.1f}kg>17kg')
            if longest > 90: triggers.append(f'最长边{longest:.0f}cm>90cm')
            if volume_cbm > 0.113: triggers.append(f'体积{volume_cbm:.3f}m³>0.113m³')
            if triggers: return '超规1', triggers

        elif carrier == 'TNT':
            if longest >= 243:
                return '超规2', [f'最长边{longest:.0f}cm≥243cm']
            triggers = []
            if l_cm > 121: triggers.append(f'长{l_cm:.0f}cm>121cm')
            if w_cm > 76: triggers.append(f'宽{w_cm:.0f}cm>76cm')
            if h_cm > 76: triggers.append(f'高{h_cm:.0f}cm>76cm')
            if wt_kg > 31: triggers.append(f'实重{wt_kg:.1f}kg>31kg')
            if triggers: return '超规1', triggers

        elif carrier == 'Hermes':
            return None, []

    return None, []


def get_uk_oversize_fee(oversize_type, warehouse, carrier):
    """按仓库+渠道查询英国超规费率"""
    if not oversize_type:
        return 0
    conn = db_connect()
    c = conn.cursor()
    c.execute('''SELECT 费用 FROM 英国超规费用 
                 WHERE 超规类型=? AND 仓库=? AND 渠道=?''',
              (oversize_type, warehouse, carrier))
    row = c.fetchone()
    conn.close()
    if row and row[0] is not None:
        return float(row[0])
    return 0


def lookup_uk_base_rate(weight_kg, carrier, warehouse='易达云'):
    """查英国基础运费：按仓库+渠道查表，取>=weight_kg的最小行"""
    conn = db_connect()
    c = conn.cursor()
    col = f'{warehouse}{carrier}'
    c.execute(f'SELECT weight_kg, [{col}] FROM 英国基础运费 WHERE [{col}] IS NOT NULL ORDER BY weight_kg')
    rows = c.fetchall()
    conn.close()
    if not rows:
        return 0, 0
    for r in rows:
        if r[0] >= weight_kg:
            return r[0], r[1]
    return rows[-1][0], rows[-1][1]


def calc_single_item_uk(item, carrier, fuel_rate_pct=0, warehouse='易达云', dim_factor=None):
    """计算单个SKU在英国指定渠道下的费用
    dim_factor: 体积重除数，None=不计抛直接用实重，5000/6000=除以该值
    """
    l_cm = float(item['length_cm'])
    w_cm = float(item['width_cm'])
    h_cm = float(item['height_cm'])
    wt_kg = float(item['weight_kg'])
    qty = int(item.get('qty', 1))

    volume_cbm = round(l_cm * w_cm * h_cm / 1_000_000, 6)

    # 计费重：有dim_factor时取max(体积重/dim_factor, 实重)，否则用实重
    if dim_factor and dim_factor > 0:
        vol_wt_kg = round(l_cm * w_cm * h_cm / dim_factor, 2)
        billing_wt_kg = max(vol_wt_kg, wt_kg)
    else:
        vol_wt_kg = None
        billing_wt_kg = wt_kg

    # 基础运费（先查表，若费率≥900表示该渠道不支持此重量段）
    lookup_w, base_rate = lookup_uk_base_rate(billing_wt_kg, carrier, warehouse)

    # 费率≥900表示不支持（如大健云仓Hermes 16kg起=999）
    if base_rate >= 900:
        return {
            'carrier': carrier,
            'length_cm': round(l_cm, 1), 'width_cm': round(w_cm, 1), 'height_cm': round(h_cm, 1),
            'weight_kg': round(wt_kg, 2), 'volume_cbm': volume_cbm,
            'exceeds_limit': True,
            'limit_reason': f'计费重 {billing_wt_kg:.1f} kg 超出该渠道支持范围，无法运送',
            'base_rate': 0, 'base_freight': 0,
            'oversize_type': None, 'oversize_fee_rate': 0, 'oversize_fee_amount': 0,
            'fuel_rate_pct': fuel_rate_pct, 'fuel_fee': 0,
        }

    base_freight = round(base_rate * qty, 2)

    # 超规检查（判定类型 → 按仓库+渠道查费率）
    os_type, os_triggers = check_uk_oversize(l_cm, w_cm, h_cm, wt_kg, carrier, warehouse)
    os_fee = get_uk_oversize_fee(os_type, warehouse, carrier)
    oversize_amount = round(os_fee * qty, 2) if os_type else 0

    # 燃油附加费 = (基础运费 + 超规费) × 燃油费率%
    fuel_base = base_freight + oversize_amount
    fuel_fee = round(fuel_base * fuel_rate_pct / 100.0, 2)

    result = {
        'carrier': carrier,
        'length_cm': round(l_cm, 1), 'width_cm': round(w_cm, 1), 'height_cm': round(h_cm, 1),
        'weight_kg': round(wt_kg, 2),
        'volume_cbm': volume_cbm,
        'base_rate_lookup_kg': lookup_w,
        'base_rate': base_rate,
        'base_freight': base_freight,
        'oversize_type': os_type,
        'oversize_triggers': os_triggers,
        'oversize_fee_rate': os_fee,
        'oversize_fee_amount': oversize_amount,
        'fuel_rate_pct': fuel_rate_pct,
        'fuel_fee': fuel_fee,
    }
    if vol_wt_kg is not None:
        result['vol_weight_kg'] = vol_wt_kg
        result['billing_weight_kg'] = billing_wt_kg
    return result


def calc_last_mile_uk(data):
    """英国尾程测算 — 易达云(5渠道) / 大健云仓(3渠道) 取最低
    DIM除数：易达云全渠道=5000，大健云仓全渠道=5000
    """
    items = data.get('items', [])
    warehouse = data.get('warehouse', '易达云')
    unloading_type = data.get('unloading_type', 'none')
    storage_days = float(data.get('storage_days', 0))
    peak_mode = data.get('peak_mode', 'none')

    # 根据仓库确定渠道列表
    # 燃油附加费率：每个仓库×每个渠道独立输入（共8个参数）
    # 参数名格式: fuel_rate_{仓库key}_{渠道小写}
    if warehouse == '易达云':
        carriers = UK_YIDAYUN_CARRIERS
    elif warehouse == '大健云仓':
        carriers = UK_GJG_CARRIERS
    else:
        return {'error': f'不支持的仓库: {warehouse}'}

    # 读取所有8个燃油附加费率（新格式），同时兼容旧格式（1个全局参数）
    fuel_rates_all = {}
    YIDAYUN_CARRIER_KEYS = ['royalmail', 'yodel', 'hermes', 'parcelforce', 'dpd']
    DAJIAN_CARRIER_MAP = {'hermes': 'Hermes', 'whistl': 'Whistl', 'tnt': 'TNT'}
    YIDAYUN_CARRIER_MAP = {k: c for k, c in zip(YIDAYUN_CARRIER_KEYS, UK_YIDAYUN_CARRIERS)}

    # 读取易达云5个渠道
    fuel_rates_all['yidayun'] = {}
    for ck, cn in YIDAYUN_CARRIER_MAP.items():
        fuel_rates_all['yidayun'][cn] = float(data.get(f'fuel_rate_yidayun_{ck}', 0))

    # 读取大健云仓3个渠道
    fuel_rates_all['dajian'] = {}
    for ck, cn in DAJIAN_CARRIER_MAP.items():
        fuel_rates_all['dajian'][cn] = float(data.get(f'fuel_rate_dajian_{ck}', 0))

    # 检查是否使用了新格式（至少有1个非零值）
    has_new_format = any(
        fuel_rates_all[wk][ck] != 0
        for wk in ['yidayun', 'dajian']
        for ck in (YIDAYUN_CARRIER_MAP.values() if wk == 'yidayun' else DAJIAN_CARRIER_MAP.values())
    )

    if has_new_format:
        # 使用新格式：按仓库取对应费率
        wk_key = 'yidayun' if warehouse == '易达云' else 'dajian'
        fuel_rates = fuel_rates_all[wk_key]
    else:
        # 兼容旧格式：1个全局参数 — 易达云→Yodel/Parcelforce/DPD，大健云仓→TNT
        fuel_rate = float(data.get('fuel_rate', data.get('fuel_rate_pct', 0)))
        if warehouse == '易达云':
            fuel_rates = {
                'Royalmail': 0, 'Yodel': fuel_rate, 'Hermes': 0,
                'Parcelforce': fuel_rate, 'DPD': fuel_rate,
            }
        else:
            fuel_rates = {
                'Hermes': 0, 'Whistl': 0, 'TNT': fuel_rate,
            }

    # DIM除数映射（体积重 = 长×宽×高 / dim_factor）
    # 英国统一使用5000
    DIM_FACTOR_MAP = {
        '易达云': {c: 5000 for c in UK_YIDAYUN_CARRIERS},
        '大健云仓': {c: 5000 for c in UK_GJG_CARRIERS},
    }
    dim_map = DIM_FACTOR_MAP.get(warehouse, DIM_FACTOR_MAP['易达云'])

    if not items:
        return {'error': '请至少添加一个产品'}

    item_results = []
    total_base_freight = 0
    total_oversize_fee = 0
    total_fuel_fee = 0
    total_qty = 0
    total_cbm = 0
    valid_qty = 0
    valid_cbm = 0

    # 按渠道汇总（所有SKU各渠道的费用合计）
    carrier_totals = {c: {'base': 0, 'oversize': 0, 'fuel': 0} for c in carriers}

    for item in items:
        sku = item.get('sku', '-')
        qty = int(item.get('qty', 1))
        length_cm = item.get('length_cm')
        width_cm = item.get('width_cm')
        height_cm = item.get('height_cm')
        weight_kg = item.get('weight_kg')

        if length_cm is None or weight_kg is None:
            conn2 = db_connect()
            c2 = conn2.cursor()
            c2.execute('SELECT 长度, 宽度, 高度, 重量 FROM 产品信息 WHERE SKU = ? LIMIT 1', (sku,))
            row = c2.fetchone()
            conn2.close()
            if row:
                if length_cm is None: length_cm = row[0]
                if width_cm is None: width_cm = row[1]
                if height_cm is None: height_cm = row[2]
                if weight_kg is None: weight_kg = row[3]

        if length_cm is None or width_cm is None or height_cm is None or weight_kg is None:
            return {'error': f'SKU "{sku or "手动输入"}" 缺少长宽高或重量数据'}

        base_item = {'sku': sku, 'qty': qty, 'length_cm': float(length_cm),
                     'width_cm': float(width_cm), 'height_cm': float(height_cm),
                     'weight_kg': float(weight_kg)}

        # 计算每个渠道的费用
        carrier_costs = {}
        for carrier in carriers:
            cr = calc_single_item_uk(base_item, carrier, fuel_rates[carrier], warehouse, dim_map[carrier])
            cr_total = cr['base_freight'] + cr['oversize_fee_amount'] + cr['fuel_fee']
            carrier_costs[carrier] = {'result': cr, 'total': cr_total}

        # 检查是否所有渠道都超出限制
        valid_carriers = {c: v for c, v in carrier_costs.items() if not v['result'].get('exceeds_limit')}
        exceeded_carriers = {c: v for c, v in carrier_costs.items() if v['result'].get('exceeds_limit')}

        if not valid_carriers:
            # 所有渠道都不支持，记录警告
            first_cr = carrier_costs[carriers[0]]['result']
            first_cr['sku'] = sku
            first_cr['qty'] = qty
            first_cr['all_carriers'] = {c: {'base': 0, 'oversize': 0, 'fuel': 0, 'fuel_rate': v['result']['fuel_rate_pct'], 'total': 0, 'exceeded': True} for c, v in carrier_costs.items()}
            item_results.append(first_cr)
            total_qty += qty
            total_cbm += first_cr['volume_cbm'] * qty
            continue

        # 在有效渠道中选最优（最低总费用）
        best = min(valid_carriers.items(), key=lambda x: x[1]['total'])
        best_carrier = best[0]
        best_result = best[1]['result']
        best_total = best[1]['total']

        best_result['sku'] = sku
        best_result['qty'] = qty
        # 所有渠道对比信息（包含超出限制的标记）
        best_result['all_carriers'] = {}
        for c, v in carrier_costs.items():
            if c in exceeded_carriers:
                best_result['all_carriers'][c] = {'base': 0, 'oversize': 0, 'fuel': 0,
                                                   'fuel_rate': v['result']['fuel_rate_pct'],
                                                   'total': 0, 'exceeded': True,
                                                   'reason': v['result'].get('limit_reason', '')}
            else:
                best_result['all_carriers'][c] = {'base': v['result']['base_freight'],
                                                   'oversize': v['result']['oversize_fee_amount'],
                                                   'fuel': v['result']['fuel_fee'],
                                                   'fuel_rate': v['result']['fuel_rate_pct'],
                                                   'total': v['total']}
        item_results.append(best_result)
        total_base_freight += best_result['base_freight']
        total_oversize_fee += best_result['oversize_fee_amount']
        total_fuel_fee += best_result['fuel_fee']
        total_qty += qty
        total_cbm += best_result['volume_cbm'] * qty
        valid_qty += qty
        valid_cbm += best_result['volume_cbm'] * qty

        # 各渠道独立汇总（UK）
        for c in carriers:
            cc = carrier_costs.get(c)
            if cc and not cc['result'].get('exceeds_limit'):
                carrier_totals[c]['base'] += round(cc['result']['base_freight'] * qty, 2)
                carrier_totals[c]['oversize'] += round(cc['result']['oversize_fee_amount'] * qty, 2)
                carrier_totals[c]['fuel'] += round(cc['result']['fuel_fee'] * qty, 2)

    total_cbm = round(total_cbm, 4)
    valid_cbm = round(valid_cbm, 4)

    # -- 仓库费用（从英国其他费用表查）--
    conn = db_connect()
    c = conn.cursor()

    # 卸货费
    unloading_fee = 0
    unloading_detail = {'type': '无', 'fee': 0}
    if unloading_type in ('fcl', 'lcl'):
        c.execute("SELECT fee_name, rate FROM 英国其他费用 WHERE category='unloading_fee' AND 仓库名称=? ORDER BY sort_order", (warehouse,))
        urows = c.fetchall()
        fcl_rate = next((r[1] for r in urows if '整柜' in str(r[0]) or 'FCL' in str(r[0]).upper()), None)
        lcl_rate = next((r[1] for r in urows if '散货' in str(r[0]) or 'LCL' in str(r[0]).upper()), None)
        if unloading_type == 'fcl' and fcl_rate:
            # FCL按柜分摊：假设65方/柜
            unloading_fee = round(valid_cbm * fcl_rate / 65.0, 2)
            unloading_detail = {'type': '整柜/FCL', 'fee': unloading_fee}
        elif unloading_type == 'lcl' and lcl_rate:
            # LCL按散货分摊：假设1.8方/散货单位
            unloading_fee = round(valid_cbm * lcl_rate / 1.8, 2)
            unloading_detail = {'type': '散货/LCL', 'fee': unloading_fee}

    # 查验费（仅易达云有此费用）
    inspection_fee = 0
    if warehouse == '易达云':
        c.execute("SELECT rate FROM 英国其他费用 WHERE category='inspection_fee' AND 仓库名称='易达云'")
        insp_row = c.fetchone()
        if insp_row and insp_row[0]:
            # 每个SKU收一次，不论数量；手动输入无SKU的每条单独计算
            unique_keys = []
            for idx, r in enumerate(item_results):
                if r.get('exceeds_limit'):
                    continue
                sku = r.get('sku', '-')
                if sku and sku != '-':
                    unique_keys.append(sku)
                else:
                    unique_keys.append(f'_manual_{idx}')
            inspection_fee = round(float(insp_row[0]) * len(set(unique_keys)), 2)

    # 入库费（仅易达云有此费用，按实重阶梯）
    total_receiving = 0
    if warehouse == '易达云':
        c.execute("SELECT condition_min, condition_max, rate FROM 英国其他费用 WHERE category='inbound_fee' AND 仓库名称=? ORDER BY sort_order", (warehouse,))
        recv_tiers = [(r[0] or 0, float('inf') if r[1] is None else r[1], r[2]) for r in c.fetchall()]
        for r in item_results:
            r['receiving_rate'] = 0
            r['receiving_item_fee'] = 0
            if r.get('exceeds_limit'): continue
            item_wt = math.ceil(r['weight_kg'])
            rate = 0
            for tmin, tmax, trate in recv_tiers:
                if item_wt >= tmin and item_wt <= tmax:
                    rate = trate
                    break
            item_fee = round(rate * r['qty'], 2)
            r['receiving_rate'] = rate
            r['receiving_item_fee'] = item_fee
            total_receiving += item_fee
        total_receiving = round(total_receiving, 2)

    # 出库费（按实重阶梯，两仓不同）
    total_outbound = 0
    c.execute("SELECT condition_min, condition_max, rate FROM 英国其他费用 WHERE category='outbound_fee' AND 仓库名称=? ORDER BY sort_order", (warehouse,))
    ob_rows = c.fetchall()
    for r in item_results:
        r['outbound_rate'] = 0
        r['outbound_item_fee'] = 0
        if r.get('exceeds_limit'): continue
        item_wt = math.ceil(r['weight_kg'])
        rate = 0
        for omin, omax, orate in ob_rows:
            tmin = omin or 0
            tmax = float('inf') if omax is None else omax
            if item_wt >= tmin and item_wt <= tmax:
                rate = orate
                break
        item_fee = round(rate * r['qty'], 2)
        r['outbound_rate'] = rate
        r['outbound_item_fee'] = item_fee
        total_outbound += item_fee
    total_outbound = round(total_outbound, 2)

    conn.close()

    # 仓储费（按库龄累进阶梯）
    total_storage = 0
    storage_breakdown = []
    if storage_days > 0 and valid_cbm > 0:
        conn2 = db_connect()
        c2 = conn2.cursor()
        c2.execute("SELECT condition_min, condition_max, rate FROM 英国其他费用 WHERE category='storage_fee' AND 仓库名称=? ORDER BY sort_order", (warehouse,))
        st_rows = c2.fetchall()
        conn2.close()
        if st_rows:
            tiers = []
            for sr in st_rows:
                tmin_raw = sr[0] or 0
                tmin = (tmin_raw - 1) if tmin_raw > 0 else 0
                tmax = float('inf') if sr[1] is None else sr[1]
                trate = sr[2] or 0
                tiers.append((tmin, tmax, trate, tmin_raw))
            remaining = storage_days
            for tmin, tmax, trate, tmin_raw in tiers:
                if remaining <= 0:
                    break
                days_in_tier = min(remaining, tmax - tmin)
                fee_in_tier = round(days_in_tier * valid_cbm * trate, 2)
                if trate == 0:
                    storage_breakdown.append({'tier': f'{tmin_raw}-{tmax if tmax != float("inf") else "+"}天', 'days': days_in_tier, 'rate': 0, 'fee': 0, 'note': '免费'})
                else:
                    storage_breakdown.append({'tier': f'{tmin_raw}-{tmax if tmax != float("inf") else "+"}天', 'days': days_in_tier, 'rate': trate, 'fee': fee_in_tier})
                total_storage += fee_in_tier
                remaining -= days_in_tier
    total_storage = round(total_storage, 2)

    # 旺季附加费
    peak_fee = 0
    if peak_mode == 'active':
        subtotal = total_base_freight + total_oversize_fee + total_fuel_fee + unloading_fee + inspection_fee + total_receiving + total_outbound + total_storage
        peak_fee = round(subtotal * 0.005, 2)

    # 总费用
    subtotal_no_storage = total_base_freight + total_oversize_fee + total_fuel_fee + unloading_fee + inspection_fee + total_receiving + total_outbound
    grand_total = round(subtotal_no_storage + total_storage + peak_fee, 2)

    # 统计超重SKU
    exceeded_count = sum(1 for r in item_results if r.get('exceeds_limit'))

    # 确定整体最优渠道（各渠道运输费用最低）
    best_carrier = min(carrier_totals, key=lambda c: carrier_totals[c]['base'] + carrier_totals[c]['oversize'] + carrier_totals[c]['fuel'])

    return {
        'items': item_results,
        'item_count': len(item_results),
        'exceeded_count': exceeded_count,
        'total_qty': total_qty,
        'total_cbm': total_cbm,
        'base_freight': total_base_freight,
        'oversize_fee': total_oversize_fee,
        'unloading_fee': unloading_fee,
        'unloading_detail': unloading_detail,
        'inspection_fee': inspection_fee,
        'receiving_fee': total_receiving,
        'outbound_fee': total_outbound,
        'storage_fee': total_storage,
        'storage_days': storage_days,
        'storage_breakdown': storage_breakdown,
        'fuel_fee': total_fuel_fee,
        'fuel_rates': fuel_rates,
        'carrier_totals': {c: dict(t) for c, t in carrier_totals.items()},
        'peak_surcharge': peak_fee,
        'peak_mode': peak_mode,
        'grand_total': grand_total,
        'subtotal_no_storage': subtotal_no_storage,
        'best_carrier': best_carrier,
    }


# ============================================================
#  英国利润测算引擎
# ============================================================

def fetch_gbp_cny_rate():
    """获取英镑兑人民币汇率（从shishihuilv.com/GBP-CNY-1.html爬取）"""
    import urllib.request
    import urllib.error
    import re
    try:
        url = 'https://www.shishihuilv.com/GBP-CNY-1.html'
        req = urllib.request.Request(url, headers={
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
            'Accept-Language': 'zh-CN,zh;q=0.9,en;q=0.8',
        })
        with urllib.request.urlopen(req, timeout=15) as resp:
            html = resp.read().decode('utf-8', errors='ignore')
            m = re.search(r'<p\s+class="to">\s*([\d.]+)\s*人民币', html)
            if not m:
                m = re.search(r'class="to"[^>]*>\s*([\d.]+)', html)
            if m:
                rate = float(m.group(1))
                if 5 < rate < 15:
                    return round(rate, 4)
    except Exception:
        pass
    try:
        url = 'https://open.er-api.com/v6/latest/GBP'
        req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
        with urllib.request.urlopen(req, timeout=10) as resp:
            data = json.loads(resp.read().decode('utf-8'))
            rate = data.get('rates', {}).get('CNY')
            if rate and rate > 0:
                return round(rate, 4)
    except Exception:
        pass
    return 9.40


def calculate_profit_uk(data):
    """英国利润测算：与US/DE一致框架 + 16.67%VAT"""
    items = data.get('items', [])
    exchange_rate = float(data.get('exchange_rate', 9.40))
    tax_per_cbm = float(data.get('tax_per_cbm', data.get('tax_total', 0)))
    management_rate = float(data.get('management_rate', 0.04))
    complaint_rate = float(data.get('complaint_rate', 0.05))
    platform_rate = float(data.get('platform_rate', 0.15))
    ad_rate = float(data.get('ad_rate', 0.15))
    profit_rate = float(data.get('profit_rate', 0.30))
    vat_rate = 0.1667  # 英国VAT 16.67%

    if not items:
        return {'error': '请至少添加一个产品'}
    if exchange_rate <= 0:
        return {'error': '汇率必须大于0'}

    total_rates = platform_rate + ad_rate + profit_rate + vat_rate
    if total_rates >= 1:
        return {'error': f'平台({platform_rate*100}%)+广告({ad_rate*100}%)+利润({profit_rate*100}%)+VAT(16.67%) = {total_rates*100}% 超过100%'}

    conn = db_connect()
    c = conn.cursor()
    c.execute("SELECT value FROM 配置参数 WHERE key = 'first_leg_base'")
    row = c.fetchone()
    first_leg_base = int(float(row[0])) if row else 600
    conn.close()

    missing_cost_skus = []
    for item in items:
        sku = item.get('sku', '-')
        purchase_price = item.get('purchase_price_manual')
        if purchase_price is None and sku and sku != '-':
            try:
                conn2 = db_connect()
                c2 = conn2.cursor()
                c2.execute('SELECT 采购价 FROM 产品信息 WHERE SKU = ? LIMIT 1', (sku,))
                row2 = c2.fetchone()
                conn2.close()
                if row2 and row2[0]:
                    purchase_price = float(row2[0])
            except Exception:
                pass
        if not purchase_price or purchase_price <= 0:
            missing_cost_skus.append(sku)
        item['_purchase_price'] = float(purchase_price) if purchase_price else 0.0

    if missing_cost_skus:
        return {'error': f'以下产品缺少采购成本：{", ".join(missing_cost_skus)}'}

    total_cbm = sum(item.get('volume_cbm', 0) * item.get('qty', 1) for item in items)
    if total_cbm <= 0:
        total_cbm = 1

    item_results = []
    for item in items:
        sku = item.get('sku', '-')
        qty = int(item.get('qty', 1))
        volume_cbm = float(item.get('volume_cbm', 0))
        purchase_price = item.get('_purchase_price', 0)
        purchase_cost = round(float(purchase_price) * qty / exchange_rate, 2)
        head_leg_cost = round(first_leg_base * volume_cbm * qty / exchange_rate, 2)
        tax_cost = round(tax_per_cbm * volume_cbm * qty, 2) if tax_per_cbm > 0 else 0
        last_mile_cost = round(float(item.get('last_mile_cost_excl_storage', 0)) / qty, 2)
        storage_cost = round(float(item.get('storage_cost', 0)) / qty, 2)
        subtotal_5 = purchase_cost + head_leg_cost + tax_cost + last_mile_cost + storage_cost
        management_cost = round(management_rate * subtotal_5, 2)
        complaint_cost = round(complaint_rate * subtotal_5, 2)
        fixed_cost = purchase_cost + head_leg_cost + tax_cost + last_mile_cost + storage_cost + management_cost + complaint_cost

        denominator = 1 - platform_rate - ad_rate - profit_rate - vat_rate
        price = round(fixed_cost / denominator, 2) if denominator > 0 else 0
        profit = round(price * profit_rate, 2)
        vat_amount = round(price * vat_rate, 2)
        calc_profit_rate = round(profit / price, 4) if price > 0 else 0

        item_results.append({
            'sku': sku, 'qty': qty, 'volume_cbm': round(volume_cbm, 4),
            'purchase_cost': purchase_cost, 'purchase_price': purchase_price,
            'head_leg_cost': head_leg_cost, 'tax_cost': tax_cost,
            'last_mile_cost': last_mile_cost, 'storage_cost': storage_cost,
            'management_cost': management_cost, 'complaint_cost': complaint_cost,
            'vat_amount': vat_amount,
            'subtotal_5': subtotal_5, 'fixed_cost': fixed_cost,
            'price': price, 'profit': profit, 'profit_rate': calc_profit_rate,
        })

    summary = {
        'purchase_cost': sum(r['purchase_cost'] for r in item_results),
        'head_leg_cost': sum(r['head_leg_cost'] for r in item_results),
        'tax_cost': sum(r['tax_cost'] for r in item_results),
        'last_mile_cost': sum(r['last_mile_cost'] for r in item_results),
        'storage_cost': sum(r['storage_cost'] for r in item_results),
        'management_cost': sum(r['management_cost'] for r in item_results),
        'complaint_cost': sum(r['complaint_cost'] for r in item_results),
        'vat_amount': sum(r['vat_amount'] for r in item_results),
        'fixed_cost': sum(r['fixed_cost'] for r in item_results),
        'price': sum(r['price'] for r in item_results),
        'profit': sum(r['profit'] for r in item_results),
    }
    summary['profit_rate'] = round(summary['profit'] / summary['price'], 4) if summary['price'] > 0 else 0

    return {
        'items': item_results,
        'total_qty': sum(r['qty'] for r in item_results),
        'total_cbm': round(total_cbm, 4),
        'exchange_rate': exchange_rate,
        'first_leg_base': first_leg_base,
        'tax_per_cbm': tax_per_cbm,
        'management_rate': round(management_rate * 100, 1),
        'complaint_rate': round(complaint_rate * 100, 1),
        'platform_rate': round(platform_rate * 100, 1),
        'ad_rate': round(ad_rate * 100, 1),
        'target_profit_rate': round(profit_rate * 100, 1),
        'vat_rate': 16.67,
        'summary': summary,
    }


# ============================================================
#  英国 API 端点
# ============================================================

@app.route('/api/calc/uk/last-mile', methods=['POST'])
def api_calc_last_mile_uk():
    """英国尾程测算"""
    try:
        data = request.get_json(force=True)
        if not data:
            return jsonify({'error': '请提供测算数据'}), 400
        result = calc_last_mile_uk(data)
        if 'error' in result:
            return jsonify(result), 400
        return jsonify(result)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/calc/uk/exchange-rate', methods=['GET'])
def api_get_gbp_rate():
    """获取英镑兑人民币汇率"""
    try:
        rate = fetch_gbp_cny_rate()
        return jsonify({'rate': rate, 'currency_pair': 'GBP/CNY', 'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/calc/uk/profit', methods=['POST'])
def api_calc_profit_uk():
    """英国利润测算"""
    try:
        data = request.get_json(force=True)
        if not data:
            return jsonify({'error': '请提供测算数据'}), 400
        result = calculate_profit_uk(data)
        if 'error' in result:
            return jsonify(result), 400
        return jsonify(result)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/calc/uk/last-mile-all', methods=['POST'])
def api_calc_uk_last_mile_all():
    """一键测算英国2个仓库的尾程费用，返回费用对比"""
    try:
        data = request.get_json(force=True)
        if not data:
            return jsonify({'error': '请提供测算参数'}), 400

        items = data.get('items', [])
        if not items:
            return jsonify({'error': '请至少添加一个产品'}), 400

        unloading_type = data.get('unloading_type', 'none')
        storage_days = float(data.get('storage_days', 0))
        peak_mode = data.get('peak_mode', 'none')

        # 读取所有8个仓库×渠道的燃油附加费率
        fuel_rate_yidayun_royalmail = float(data.get('fuel_rate_yidayun_royalmail', 0))
        fuel_rate_yidayun_yodel = float(data.get('fuel_rate_yidayun_yodel', 0))
        fuel_rate_yidayun_hermes = float(data.get('fuel_rate_yidayun_hermes', 0))
        fuel_rate_yidayun_parcelforce = float(data.get('fuel_rate_yidayun_parcelforce', 0))
        fuel_rate_yidayun_dpd = float(data.get('fuel_rate_yidayun_dpd', 0))
        fuel_rate_dajian_hermes = float(data.get('fuel_rate_dajian_hermes', 0))
        fuel_rate_dajian_whistl = float(data.get('fuel_rate_dajian_whistl', 0))
        fuel_rate_dajian_tnt = float(data.get('fuel_rate_dajian_tnt', 0))

        # 仓库→燃油费率映射
        UK_FUEL_PARAMS = {
            '易达云': {
                'fuel_rate_yidayun_royalmail': fuel_rate_yidayun_royalmail,
                'fuel_rate_yidayun_yodel': fuel_rate_yidayun_yodel,
                'fuel_rate_yidayun_hermes': fuel_rate_yidayun_hermes,
                'fuel_rate_yidayun_parcelforce': fuel_rate_yidayun_parcelforce,
                'fuel_rate_yidayun_dpd': fuel_rate_yidayun_dpd,
            },
            '大健云仓': {
                'fuel_rate_dajian_hermes': fuel_rate_dajian_hermes,
                'fuel_rate_dajian_whistl': fuel_rate_dajian_whistl,
                'fuel_rate_dajian_tnt': fuel_rate_dajian_tnt,
            },
        }

        warehouse_results = {}
        optimal_wh = None
        optimal_total = float('inf')

        for wh_name in UK_WAREHOUSES:
            wh_fuels = UK_FUEL_PARAMS[wh_name]
            lm_input = {
                'items': items,
                'warehouse': wh_name,
                'unloading_type': unloading_type,
                'storage_days': storage_days,
                'peak_mode': peak_mode,
                **wh_fuels,
            }

            lm_result = calc_last_mile_uk(lm_input)
            if 'error' in lm_result:
                return jsonify({'error': f'{wh_name} 尾程测算失败: {lm_result["error"]}'}), 400

            warehouse_results[wh_name] = lm_result

            grand_total = lm_result.get('grand_total', float('inf'))
            if grand_total < optimal_total:
                optimal_total = grand_total
                optimal_wh = wh_name

        return jsonify({
            'warehouses': warehouse_results,
            'optimal_warehouse': optimal_wh,
            'optimal_total': optimal_total,
            'item_count': len(items),
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ============================================================
#  英国配置查询 API
# ============================================================

@app.route('/api/config/uk/rates', methods=['GET'])
def get_uk_rates():
    """获取英国基础运费（易达云×5渠道 + 大健云仓×3渠道）"""
    conn = db_connect()
    c = conn.cursor()
    c.execute('SELECT * FROM 英国基础运费 ORDER BY weight_kg')
    rows = c.fetchall()
    conn.close()
    rates = []
    for row in rows:
        rates.append({
            'weight_kg': row[0],
            '易达云Royalmail': row[1], '易达云Yodel': row[2], '易达云Hermes': row[3],
            '易达云Parcelforce': row[4], '易达云DPD': row[5],
            '大健云仓Hermes': row[6], '大健云仓Whistl': row[7], '大健云仓TNT': row[8],
        })
    return jsonify({'rates': rates, 'count': len(rates)})


@app.route('/api/config/uk/oversize', methods=['GET'])
def get_uk_oversize():
    """获取英国超规费用"""
    conn = db_connect()
    c = conn.cursor()
    c.execute('SELECT * FROM 英国超规费用 ORDER BY id')
    rows = c.fetchall()
    conn.close()
    result = []
    for row in rows:
        result.append({
            'id': row[0], 'oversize_type': row[1], 'warehouse': row[2],
            'carrier': row[3], 'description': row[4], 'fee': row[5],
        })
    return jsonify({'oversize': result, 'count': len(result)})


@app.route('/api/config/uk/other-fees', methods=['GET'])
def get_uk_other_fees():
    """获取英国其他费用（卸货/查验/入库/出库/仓储 按仓库阶梯）"""
    conn = db_connect()
    c = conn.cursor()
    c.execute('SELECT * FROM 英国其他费用 ORDER BY sort_order')
    rows = c.fetchall()
    conn.close()
    fees = []
    for row in rows:
        fees.append({
            'id': row[0], 'category': row[1], 'fee_name': row[2],
            'condition_desc': row[3], 'condition_min': row[4],
            'condition_max': row[5], 'condition_unit': row[6],
            'rate': row[7], 'unit': row[8], 'notes': row[9],
            'sort_order': row[10], 'warehouse': row[11],
        })
    return jsonify({'fees': fees, 'count': len(fees)})


@app.route('/api/calc/last-mile', methods=['POST'])
def api_calc_last_mile():
    """尾程测算API"""
    try:
        data = request.get_json(force=True)
        if not data:
            return jsonify({'error': '请提供测算参数'}), 400

        result = calculate_last_mile(data)
        if 'error' in result:
            return jsonify(result), 400

        return jsonify(result)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/calc/fuel-rate', methods=['GET'])
def api_fetch_fuel_rate():
    """获取FedEx最新燃油附加费率"""
    try:
        rate = fetch_fuel_surcharge()
        return jsonify({'fuel_rate_pct': rate, 'source': 'fedex.com'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ============================================================
#  多仓库尾程测算 API（最优模式使用）
# ============================================================

@app.route('/api/calc/last-mile-all', methods=['POST'])
def api_calc_last_mile_all():
    """一键测算美国3个仓库的尾程费用，返回最优仓库"""
    try:
        data = request.get_json(force=True)
        if not data:
            return jsonify({'error': '请提供测算参数'}), 400

        items = data.get('items', [])
        if not items:
            return jsonify({'error': '请至少添加一个产品'}), 400

        zone = int(data.get('zone', 8))
        unloading_type = data.get('unloading_type', 'none')
        storage_days = float(data.get('storage_days', 0))
        remote_mode = data.get('remote_surcharge_mode', 'none')
        peak_mode = data.get('peak_surcharge_mode', 'none')
        fuel_rate_pct = data.get('fuel_rate_pct')

        warehouse_results = {}
        optimal_wh = None
        optimal_total = float('inf')

        for wh_name in US_WAREHOUSES:
            lm_input = {
                'items': items,
                'zone': zone,
                'warehouse': wh_name,
                'unloading_type': unloading_type,
                'storage_days': storage_days,
                'remote_surcharge_mode': remote_mode,
                'peak_surcharge_mode': peak_mode,
                'fuel_rate_pct': fuel_rate_pct
            }
            lm_result = calculate_last_mile(lm_input)
            if 'error' in lm_result:
                return jsonify({'error': f'{wh_name} 尾程测算失败: {lm_result["error"]}'}), 400

            warehouse_results[wh_name] = lm_result

            grand_total = lm_result.get('grand_total', float('inf'))
            if grand_total < optimal_total:
                optimal_total = grand_total
                optimal_wh = wh_name

        return jsonify({
            'warehouses': warehouse_results,
            'optimal': optimal_wh,
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/template/last-mile', methods=['GET'])
def download_last_mile_template():
    """下载尾程测算Excel模板"""
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = '尾程测算输入'

    header_font = Font(name='微软雅黑', bold=True, size=11)
    header_fill = PatternFill('solid', fgColor='4472C4')
    header_font_white = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # 说明行
    ws.merge_cells('A1:F1')
    note_cell = ws.cell(row=1, column=1,
        value='说明：① SKU 与 长宽高重量 二选一填写（填了SKU可留空尺寸重量，系统自动从数据库查询）；② 数量 必填。')
    note_cell.font = Font(name='微软雅黑', size=10, color='FF0000')
    note_cell.alignment = Alignment(wrap_text=True, vertical='center')

    headers = ['SKU', '数量 *必填', '长(cm)', '宽(cm)', '高(cm)', '重(kg)']
    header_row = 2
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    data_row = 3
    for col, v in enumerate(['GD04102141BK', 1, 132, 71, 14.5, 18.95], 1):
        cell = ws.cell(row=data_row, column=col, value=v)
        cell.border = thin_border
        cell.font = Font(name='微软雅黑', size=10, color='808080')

    ws.row_dimensions[1].height = 28

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 10
    for c in ['C', 'D', 'E', 'F']:
        ws.column_dimensions[c].width = 14

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='尾程测算模板.xlsx'
    )


@app.route('/api/template/de-last-mile', methods=['GET'])
def download_de_last_mile_template():
    """下载德国尾程测算Excel模板"""
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = '德国尾程测算输入'

    header_font_white = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='4472C4')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    ws.merge_cells('A1:F1')
    note_cell = ws.cell(row=1, column=1,
        value='说明：① SKU 与 长宽高重量 二选一填写（填了SKU可留空尺寸，系统自动查数据库）；② 数量 必填。')
    note_cell.font = Font(name='微软雅黑', size=10, color='FF0000')
    note_cell.alignment = Alignment(wrap_text=True, vertical='center')

    headers = ['SKU', '数量 *必填', '长(cm)', '宽(cm)', '高(cm)', '重(kg)']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 12
    for c in ['C', 'D', 'E', 'F']:
        ws.column_dimensions[c].width = 14

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='德国尾程测算模板.xlsx'
    )


@app.route('/api/template/uk-last-mile', methods=['GET'])
def download_uk_last_mile_template():
    """下载英国尾程测算Excel模板"""
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = '英国尾程测算输入'

    header_font_white = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='4472C4')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    ws.merge_cells('A1:F1')
    note_cell = ws.cell(row=1, column=1,
        value='说明：① SKU 与 长宽高重量 二选一填写（填了SKU可留空尺寸，系统自动查数据库）；② 数量 必填。')
    note_cell.font = Font(name='微软雅黑', size=10, color='FF0000')
    note_cell.alignment = Alignment(wrap_text=True, vertical='center')

    headers = ['SKU', '数量 *必填', '长(cm)', '宽(cm)', '高(cm)', '重(kg)']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 12
    for c in ['C', 'D', 'E', 'F']:
        ws.column_dimensions[c].width = 14

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='英国尾程测算模板.xlsx'
    )


@app.route('/api/import/last-mile', methods=['POST'])
def import_last_mile_excel():
    """导入Excel文件，解析SKU列表"""
    from io import BytesIO
    from openpyxl import load_workbook

    if 'file' not in request.files:
        return jsonify({'error': '请上传Excel文件'}), 400

    file = request.files['file']
    if not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({'error': '仅支持.xlsx或.xls格式'}), 400

    try:
        wb = load_workbook(BytesIO(file.read()))
        ws = wb.active
        items = []

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or all(v is None for v in row):
                continue
            # 跳过说明行和表头行
            first_val = str(row[0]).strip() if row[0] else ''
            if first_val in ('SKU', 'SKU *必填') or first_val.startswith('说明'):
                continue
            sku = str(row[0]).strip() if row[0] else ''
            qty = int(row[1]) if row[1] and str(row[1]).strip() else 1
            length = float(row[2]) if row[2] and str(row[2]).strip() else None
            width = float(row[3]) if row[3] and str(row[3]).strip() else None
            height = float(row[4]) if row[4] and str(row[4]).strip() else None
            weight = float(row[5]) if row[5] and str(row[5]).strip() else None

            if not sku and any(v is None for v in [length, width, height, weight]):
                continue

            item = {'qty': max(1, qty)}
            if sku: item['sku'] = sku
            if length is not None: item['length_cm'] = length
            if width is not None: item['width_cm'] = width
            if height is not None: item['height_cm'] = height
            if weight is not None: item['weight_kg'] = weight
            items.append(item)

        return jsonify({'items': items, 'count': len(items)})
    except Exception as e:
        return jsonify({'error': f'文件解析失败: {str(e)}'}), 400


@app.route('/api/export/last-mile', methods=['POST'])
def export_last_mile_result():
    """导出尾程测算结果为Excel"""
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    data = request.get_json(force=True)
    if not data:
        return jsonify({'error': '无导出数据'}), 400

    wb = Workbook()
    ws = wb.active
    ws.title = '尾程测算结果'

    header_font = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='4472C4')
    total_fill = PatternFill('solid', fgColor='C6EFCE')
    total_font = Font(name='微软雅黑', bold=True, size=12, color='006100')
    normal_font = Font(name='微软雅黑', size=10)
    sub_font = Font(name='微软雅黑', size=9, color='808080')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Sheet 1: 费用汇总
    ws.merge_cells('A1:C1')
    ws.cell(row=1, column=1, value=f'尾程测算结果 - Zone {data.get("zone","")}').font = Font(name='微软雅黑', bold=True, size=14)

    fee_headers = ['费用项目', '金额(USD)', '说明']
    for col, h in enumerate(fee_headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

    # 入库上架费说明
    rd = data.get('receiving_details', [])
    recv_note = '、'.join([f"{d['sku']}({d['weight_lb']}lb): {d['rate']}" for d in rd]) if rd else '按各产品实重分别计费'

    fees = [
        ('基础运费', data.get('base_freight', 0), f'Zone {data.get("zone","")}，共 {data.get("total_qty",0)} 件'),
        ('超尺寸费', data.get('oversize_fee', 0), ''),
        ('住宅附加费', data.get('residential_surcharge', 0), f'{data.get("total_qty",0)} 件 × 2.50'),
        ('卸货费', data.get('unloading_fee', 0), data.get('unloading_detail', {}).get('type', '')),
        ('入库上架费', data.get('receiving_fee', 0), recv_note),
        ('出库费', data.get('outbound_fee', 0), f'{data.get("total_qty",0)} 件 × 1.00'),
        ('仓储费', data.get('storage_fee', 0), f'{data.get("storage_days",0)} 天'),
        ('偏远附加费', data.get('remote_surcharge', 0), data.get('remote_mode', '')),
        ('燃油附加费', data.get('fuel_surcharge', 0), f'费率 {data.get("fuel_rate_pct",0)}%'),
        ('旺季附加费', data.get('peak_surcharge', 0), data.get('peak_mode', '')),
    ]

    row = 4
    for name, amount, note in fees:
        ws.cell(row=row, column=1, value=name).font = normal_font
        ws.cell(row=row, column=2, value=amount).font = normal_font
        ws.cell(row=row, column=2).number_format = '#,##0.00'
        ws.cell(row=row, column=3, value=note).font = sub_font
        for c in range(1, 4):
            ws.cell(row=row, column=c).border = thin_border
        row += 1

    ws.cell(row=row, column=1, value='总费用').font = total_font
    ws.cell(row=row, column=1).fill = total_fill
    ws.cell(row=row, column=2, value=data.get('grand_total', 0)).font = total_font
    ws.cell(row=row, column=2).fill = total_fill
    ws.cell(row=row, column=2).number_format = '#,##0.00'
    ws.cell(row=row, column=3, value='USD').font = total_font
    ws.cell(row=row, column=3).fill = total_fill
    for c in range(1, 4):
        ws.cell(row=row, column=c).border = thin_border

    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 40

    # Sheet 2: 产品明细
    ws2 = wb.create_sheet('产品明细')
    detail_headers = ['SKU', '数量', '尺寸(cm)', '尺寸(in)', '取整(in)', '实重(kg)', '实重(lb)', '材积重(lb)', '体积(cu in)', '计费重(lb)', '查表重(lb)', '基础费率', '基础运费', '超尺寸类型', '超尺寸详情', '超尺寸费']
    for col, h in enumerate(detail_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    items = data.get('items', [])
    for i, item in enumerate(items):
        r = i + 2
        vals = [
            item.get('sku', '-'), item.get('qty', 1),
            item.get('dims_cm', ''), item.get('dims_in', ''), item.get('dims_ceil_in', ''),
            item.get('weight_kg', ''), item.get('weight_lb', ''), item.get('dim_weight_lb', ''),
            item.get('volume_cu_in', ''), item.get('billing_weight_final', ''),
            item.get('base_rate_lookup_weight', ''), item.get('base_rate', ''),
            item.get('base_freight', ''), item.get('oversize_fee_type', '') or '-',
            '; '.join(item.get('oversize_cm', []) or []) or '-',
            item.get('oversize_fee_amount', '')
        ]
        for c, v in enumerate(vals, 1):
            cell = ws2.cell(row=r, column=c, value=v)
            cell.font = normal_font
            cell.border = thin_border

    ws2.column_dimensions['A'].width = 18
    for c in ['B', 'C', 'D', 'E', 'F']:
        ws2.column_dimensions[c].width = 14
    for c in ['G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P']:
        ws2.column_dimensions[c].width = 13

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='尾程测算结果.xlsx'
    )


@app.route('/api/export/uk-last-mile', methods=['POST'])
def export_uk_last_mile_result():
    """导出英国尾程测算结果为Excel（多渠道显示）"""
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    data = request.get_json(force=True)
    if not data:
        return jsonify({'error': '无导出数据'}), 400

    wb = Workbook()
    ws = wb.active
    ws.title = '费用汇总'

    hf = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
    hfill = PatternFill('solid', fgColor='4472C4')
    nf = Font(name='微软雅黑', size=10)
    tf = Font(name='微软雅黑', bold=True, size=12, color='006100')
    tfill = PatternFill('solid', fgColor='D9F2D9')
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))

    # 确定渠道列表（从fuel_rates或carrier_totals提取）
    fuel_rates = data.get('fuel_rates', {})
    carriers = list(fuel_rates.keys()) if fuel_rates else list((data.get('carrier_totals') or {}).keys())
    if not carriers:
        # fallback: 从items中提取
        items = data.get('items', [])
        for it in items:
            ac = it.get('all_carriers', {})
            if ac and not it.get('exceeds_limit'):
                carriers = list(ac.keys())
                break
    best_carrier = data.get('best_carrier', '')
    ct = data.get('carrier_totals', {})
    total_qty = data.get('total_qty', 0)

    # Sheet 1: 费用汇总（多渠道列）
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=1 + len(carriers))
    ws.cell(row=1, column=1, value='英国尾程测算 - 费用汇总').font = Font(name='微软雅黑', bold=True, size=14)

    # 表头
    ws.cell(row=3, column=1, value='费用项目').font = hf
    ws.cell(row=3, column=1).fill = hfill; ws.cell(row=3, column=1).border = border
    for ci, c in enumerate(carriers):
        col = ci + 2
        label = c + (' ★' if c == best_carrier else '')
        cell = ws.cell(row=3, column=col, value=label)
        cell.font = hf; cell.fill = hfill; cell.border = border
        cell.alignment = Alignment(horizontal='center')

    # 费用行
    def _ctval(c, k):
        d = ct.get(c) or {}
        return d.get(k, 0)

    fee_rows = [
        ('基础运费 (£)', lambda c: _ctval(c, 'base')),
        ('超规附加费 (£)', lambda c: _ctval(c, 'oversize')),
        ('燃油附加费 (£)', lambda c: _ctval(c, 'fuel')),
    ]
    r = 4
    for name, fn in fee_rows:
        ws.cell(row=r, column=1, value=name).font = nf
        ws.cell(row=r, column=1).border = border
        for ci, c in enumerate(carriers):
            cell = ws.cell(row=r, column=ci + 2, value=round(fn(c), 2))
            cell.font = nf; cell.border = border; cell.number_format = '#,##0.00'
        r += 1

    # 运输小计
    ws.cell(row=r, column=1, value='运输小计 (£)').font = Font(name='微软雅黑', bold=True, size=10)
    ws.cell(row=r, column=1).border = border
    for ci, c in enumerate(carriers):
        v = round(_ctval(c, 'base') + _ctval(c, 'oversize') + _ctval(c, 'fuel'), 2)
        cell = ws.cell(row=r, column=ci + 2, value=v)
        cell.font = Font(name='微软雅黑', bold=True, size=10); cell.border = border; cell.number_format = '#,##0.00'
    r += 1

    # 仓库费用（不分渠道）
    wh_fees = [
        ('卸货费 (£)', data.get('unloading_fee', 0)),
        ('查验费 (£)', data.get('inspection_fee', 0)),
        ('入库费 (£)', data.get('receiving_fee', 0)),
        ('出库费 (£)', data.get('outbound_fee', 0)),
        ('仓储费 (£) ' + str(data.get('storage_days', 0)) + '天', data.get('storage_fee', 0)),
        ('旺季附加费 (£)', data.get('peak_surcharge', 0)),
    ]
    for name, val in wh_fees:
        ws.cell(row=r, column=1, value=name).font = nf; ws.cell(row=r, column=1).border = border
        for ci in range(len(carriers)):
            cell = ws.cell(row=r, column=ci + 2, value=round(val, 2))
            cell.font = nf; cell.border = border; cell.number_format = '#,##0.00'
        r += 1

    # 总费用
    ws.cell(row=r, column=1, value='总费用 (£)').font = tf; ws.cell(row=r, column=1).fill = tfill
    ws.cell(row=r, column=1).border = border
    for ci, c in enumerate(carriers):
        t = round(_ctval(c, 'base') + _ctval(c, 'oversize') + _ctval(c, 'fuel') +
                  data.get('unloading_fee', 0) + data.get('inspection_fee', 0) +
                  data.get('receiving_fee', 0) + data.get('outbound_fee', 0) +
                  data.get('storage_fee', 0) + data.get('peak_surcharge', 0), 2)
        cell = ws.cell(row=r, column=ci + 2, value=t)
        cell.font = tf; cell.fill = tfill; cell.border = border; cell.number_format = '#,##0.00'
    r += 1

    ws.column_dimensions['A'].width = 22
    for ci in range(len(carriers)):
        ws.column_dimensions[chr(66 + ci)].width = 15

    # Sheet 2: SKU明细
    ws2 = wb.create_sheet('产品明细')
    detail_headers = ['SKU', '数量', '体积(m³)', '实重(kg)', '最优渠道',
                      '基础运费', '超规类型', '超规详情', '超规费', '燃油费', '小计']
    for c in carriers:
        detail_headers.append(c + '基础')
        detail_headers.append(c + '超规')
        detail_headers.append(c + '燃油')
    for c in carriers:
        detail_headers.append(c + '最终费用')
    detail_headers += ['卸货分摊', '查验分摊', '入库费', '出库费', '仓储分摊', '旺季分摊']

    for col, h in enumerate(detail_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = hf; cell.fill = hfill; cell.alignment = Alignment(horizontal='center'); cell.border = border

    items = data.get('items', [])
    valid_count = sum(1 for it in items if not it.get('exceeds_limit')) or 1
    for i, it in enumerate(items):
        row_idx = i + 2
        qty = it.get('qty', 1)
        vals = [it.get('sku', '-'), qty, round(it.get('volume_cbm', 0), 4),
                it.get('weight_kg', 0),
                it.get('carrier') if not it.get('exceeds_limit') else '超限']
        # 最优渠道费用
        if it.get('exceeds_limit'):
            vals += ['', '', '', '', '', '']
        else:
            vals += [round(it.get('base_freight', 0), 2),
                     it.get('oversize_type') or '-',
                     '; '.join(it.get('oversize_triggers') or []) if it.get('oversize_triggers') else '-',
                     round(it.get('oversize_fee_amount', 0), 2),
                     round(it.get('fuel_fee', 0), 2),
                     round(it.get('base_freight', 0) + it.get('oversize_fee_amount', 0) + it.get('fuel_fee', 0), 2)]
        # 各渠道费用
        ac = it.get('all_carriers', {})
        for c in carriers:
            d = ac.get(c) or {}
            vals.append(round(d.get('base', 0), 2) if not it.get('exceeds_limit') else '')
            vals.append(round(d.get('oversize', 0), 2) if not it.get('exceeds_limit') else '')
            vals.append(round(d.get('fuel', 0), 2) if not it.get('exceeds_limit') else '')
        # 分摊
        u_sh = round(data.get('unloading_fee', 0) / valid_count, 2)
        i_sh = round(data.get('inspection_fee', 0) / valid_count, 2)
        r_sh = round(data.get('receiving_fee', 0) / valid_count, 2)
        o_sh = round(data.get('outbound_fee', 0) / valid_count, 2)
        s_sh = round(data.get('storage_fee', 0) / valid_count, 2)
        p_sh = round(data.get('peak_surcharge', 0) / valid_count, 2)
        # 各渠道最终费用 = 运输费 + 所有仓库分摊费
        for c in carriers:
            d = ac.get(c) or {}
            transport = round((d.get('base', 0) if not it.get('exceeds_limit') else 0) +
                              (d.get('oversize', 0) if not it.get('exceeds_limit') else 0) +
                              (d.get('fuel', 0) if not it.get('exceeds_limit') else 0), 2)
            vals.append(round(transport + u_sh + i_sh + r_sh + o_sh + s_sh + p_sh, 2) if not it.get('exceeds_limit') else '')
        # 仓库分摊费用
        if it.get('exceeds_limit'):
            vals += ['', '', '', '', '', '', '']
        else:
            vals += [u_sh, i_sh, r_sh, o_sh, s_sh, p_sh]
        for ci, v in enumerate(vals, 1):
            cell = ws2.cell(row=row_idx, column=ci, value=v)
            cell.font = nf; cell.border = border

    ws2.column_dimensions['A'].width = 18
    for c in ['B', 'C', 'D', 'E']:
        ws2.column_dimensions[c].width = 10
    for ci in range(len(carriers) * 3 + 6):
        ws2.column_dimensions[chr(70 + ci) if ci < 21 else 'AA'].width = 12

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name='英国尾程测算.xlsx')


@app.route('/api/export/de-last-mile', methods=['POST'])
def export_de_last_mile_result():
    """导出德国尾程测算结果为Excel"""
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    data = request.get_json(force=True)
    if not data:
        return jsonify({'error': '无导出数据'}), 400

    wb = Workbook()
    ws = wb.active
    ws.title = '德国尾程测算'

    hf = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
    hfill = PatternFill('solid', fgColor='4472C4')
    nf = Font(name='微软雅黑', size=10)
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))

    # 明细表头
    detail_headers = ['SKU', '数量', '尺寸(cm)', '重量(kg)', '体积(m³)',
                      '最优渠道', '基础运费', '超规类型', '超规详情', '超规费', '燃油附加费', '渠道小计',
                      'DPD基础', 'DPD超规', 'DPD燃油',
                      'DHL基础', 'DHL超规', 'DHL燃油',
                      'GLS基础', 'GLS超规', 'GLS燃油',
                      '卸货分摊', '入库分摊', '出库分摊', '仓储分摊', '旺季分摊',
                      'SKU小计']
    for col, h in enumerate(detail_headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = hf; c.fill = hfill; c.alignment = Alignment(horizontal='center'); c.border = border

    items = data.get('items', [])
    total_cbm = data.get('total_cbm', 1) or 1
    total_qty = data.get('total_qty', 1) or 1
    unloading = data.get('unloading_fee', 0)
    receiving = data.get('receiving_fee', 0)
    outbound = data.get('outbound_fee', 0)
    storage = data.get('storage_fee', 0)
    peak = data.get('peak_surcharge', 0)

    for i, r in enumerate(items, 2):
        ac = r.get('all_carriers', {})
        dpd = ac.get('DPD', {}) or {}
        dhl = ac.get('DHL', {}) or {}
        gls = ac.get('GLS', {}) or {}
        fuel_fee = r.get('fuel_fee', 0)
        qty = r.get('qty', 1)
        vol_cbm = r.get('volume_cbm', 0)
        # 仓库费用按体积占比分摊
        vol_ratio = (vol_cbm * qty / total_cbm) if total_cbm > 0 else 0
        unloading_share = round(unloading * vol_ratio, 2)
        receiving_share = round(receiving * vol_ratio, 2)
        storage_share = round(storage * vol_ratio, 2)
        peak_share = round(peak * vol_ratio, 2)
        # 出库费按件数占比分摊
        qty_ratio = qty / total_qty if total_qty > 0 else 0
        outbound_share = round(outbound * qty_ratio, 2)

        carrier_subtotal = round(r.get('base_freight', 0) + r.get('oversize_fee_amount', 0) + fuel_fee, 2)
        sku_total = round(carrier_subtotal + unloading_share + receiving_share + outbound_share + storage_share + peak_share, 2)

        vals = [r.get('sku', ''), qty,
                f'{r.get("length_cm", 0)}×{r.get("width_cm", 0)}×{r.get("height_cm", 0)}',
                r.get('weight_kg', 0), round(vol_cbm, 6),
                r.get('carrier', ''), r.get('base_freight', 0), r.get('oversize_type') or '-', '; '.join(r.get('oversize_triggers') or []) or '-', r.get('oversize_fee_amount', 0),
                fuel_fee, carrier_subtotal,
                dpd.get('base', 0), dpd.get('oversize', 0), dpd.get('fuel', 0),
                dhl.get('base', 0), dhl.get('oversize', 0), dhl.get('fuel', 0),
                gls.get('base', 0), gls.get('oversize', 0), gls.get('fuel', 0),
                unloading_share, receiving_share, outbound_share, storage_share, peak_share,
                sku_total]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=i, column=col, value=v)
            c.font = nf; c.border = border

    # 汇总区
    sr = len(items) + 3
    summary = [
        ('费用汇总', '金额(€)', ''),
        ('基础运费', data.get('base_freight', 0), ''),
        ('超尺寸费', data.get('oversize_fee', 0), ''),
        ('燃油附加费', data.get('fuel_fee', 0), f'DPD:{data.get("fuel_rate_dpd",0)}% DHL:{data.get("fuel_rate_dhl",0)}% GLS:{data.get("fuel_rate_gls",0)}%'),
        ('卸货费', data.get('unloading_fee', 0), ''),
        ('入库费', data.get('receiving_fee', 0), ''),
        ('出库费', data.get('outbound_fee', 0), ''),
        ('仓储费', data.get('storage_fee', 0), f'{data.get("storage_days",0)}天'),
        ('旺季附加费', data.get('peak_surcharge', 0), ''),
        ('总费用', data.get('grand_total', 0), ''),
    ]
    for i, (label, val, note) in enumerate(summary):
        c1 = ws.cell(row=sr + i, column=1, value=label)
        c2 = ws.cell(row=sr + i, column=2, value=val)
        c3 = ws.cell(row=sr + i, column=3, value=note)
        ft = Font(name='微软雅黑', bold=True, size=10) if i == 0 else nf
        c1.font = ft; c2.font = ft; c3.font = nf
        c1.border = border; c2.border = border; c3.border = border

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['C'].width = 20
    for c in ['B', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y']:
        ws.column_dimensions[c].width = 12

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'德国尾程测算_{datetime.now().strftime("%Y%m%d")}.xlsx')


# ============================================================
#  英国+德国 最优模式多仓库 Excel 导出（每仓库一个Sheet）
# ============================================================

@app.route('/api/export/de-last-mile-all', methods=['POST'])
def export_de_last_mile_all():
    """导出德国最优模式（3个仓库各自一个Sheet）"""
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    data = request.get_json(force=True)
    warehouses = data.get('warehouses', {})
    if not warehouses:
        return jsonify({'error': '无仓库数据'}), 400

    hf = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
    hfill = PatternFill('solid', fgColor='4472C4')
    nf = Font(name='微软雅黑', size=10)
    tf = Font(name='微软雅黑', bold=True, size=11, color='006100')
    tfill = PatternFill('solid', fgColor='D9F2D9')
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))
    carriers_list = ['DPD', 'DHL', 'GLS']

    wb = Workbook()
    wb.remove(wb.active)

    for wh_name in DE_WAREHOUSES:
        wh_data = warehouses.get(wh_name)
        if not wh_data:
            continue
        ct = wh_data.get('carrier_totals', {})
        best = wh_data.get('best_carrier', '')
        if not ct:
            continue

        # 费用汇总 sheet
        ws_name = f'{wh_name}-汇总'
        ws = wb.create_sheet(ws_name)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
        ws.cell(row=1, column=1, value=f'德国尾程测算 - {wh_name}').font = Font(name='微软雅黑', bold=True, size=14)

        ws.cell(row=3, column=1, value='费用项目').font = hf; ws.cell(row=3, column=1).fill = hfill; ws.cell(row=3, column=1).border = border
        for ci, c in enumerate(carriers_list):
            col = ci + 2
            label = c + (' ★' if c == best else '')
            cell = ws.cell(row=3, column=col, value=label)
            cell.font = hf; cell.fill = hfill; cell.border = border; cell.alignment = Alignment(horizontal='center')

        def _dv(c, k):
            d = ct.get(c) or {}
            return d.get(k, 0)

        fee_rows = [('基础运费 (€)', 'base'), ('超尺寸费 (€)', 'oversize'), ('燃油附加费 (€)', 'fuel')]
        r = 4
        for name, key in fee_rows:
            ws.cell(row=r, column=1, value=name).font = nf; ws.cell(row=r, column=1).border = border
            for ci, c in enumerate(carriers_list):
                cell = ws.cell(row=r, column=ci + 2, value=round(_dv(c, key), 2))
                cell.font = nf; cell.border = border; cell.number_format = '#,##0.00'
            r += 1
        # 运输小计
        ws.cell(row=r, column=1, value='运输小计 (€)').font = Font(name='微软雅黑', bold=True, size=10); ws.cell(row=r, column=1).border = border
        for ci, c in enumerate(carriers_list):
            v = round(_dv(c, 'base') + _dv(c, 'oversize') + _dv(c, 'fuel'), 2)
            cell = ws.cell(row=r, column=ci + 2, value=v)
            cell.font = Font(name='微软雅黑', bold=True, size=10); cell.border = border; cell.number_format = '#,##0.00'
        r += 1
        # 仓库费用
        wh_fees = [
            ('卸货费 (€)', wh_data.get('unloading_fee', 0)),
            ('入库上架费 (€)', wh_data.get('receiving_fee', 0)),
            ('出库操作费 (€)', wh_data.get('outbound_fee', 0)),
            ('仓储费 (€)', wh_data.get('storage_fee', 0)),
            ('旺季附加费 (€)', wh_data.get('peak_surcharge', 0)),
        ]
        for name, val in wh_fees:
            ws.cell(row=r, column=1, value=name).font = nf; ws.cell(row=r, column=1).border = border
            for ci in range(len(carriers_list)):
                cell = ws.cell(row=r, column=ci + 2, value=round(val, 2))
                cell.font = nf; cell.border = border; cell.number_format = '#,##0.00'
            r += 1
        # 总费用
        ws.cell(row=r, column=1, value='总费用 (€)').font = tf; ws.cell(row=r, column=1).fill = tfill; ws.cell(row=r, column=1).border = border
        for ci, c in enumerate(carriers_list):
            t = round(_dv(c, 'base') + _dv(c, 'oversize') + _dv(c, 'fuel') +
                      wh_data.get('unloading_fee', 0) + wh_data.get('receiving_fee', 0) +
                      wh_data.get('outbound_fee', 0) + wh_data.get('storage_fee', 0) +
                      wh_data.get('peak_surcharge', 0), 2)
            cell = ws.cell(row=r, column=ci + 2, value=t)
            cell.font = tf; cell.fill = tfill; cell.border = border; cell.number_format = '#,##0.00'

        ws.column_dimensions['A'].width = 22
        for ci in range(3):
            ws.column_dimensions[chr(66 + ci)].width = 15

        # 产品明细 sheet
        ws2 = wb.create_sheet(f'{wh_name}-明细')
        headers = ['SKU', '数量', '尺寸(cm)', '重量(kg)', '体积(m³)', '最优渠道',
                   '基础运费', '超规类型', '超规详情', '超规费', '燃油费', '渠道小计']
        for c in carriers_list:
            headers += [c + '基础', c + '超规', c + '燃油']
        for c in carriers_list:
            headers += [c + '最终费用']
        headers += ['卸货分摊', '入库分摊', '出库分摊', '仓储分摊', '旺季分摊']
        for col, h in enumerate(headers, 1):
            cell = ws2.cell(row=1, column=col, value=h)
            cell.font = hf; cell.fill = hfill; cell.alignment = Alignment(horizontal='center'); cell.border = border

        items = wh_data.get('items', [])
        valid_n = sum(1 for it in items if not it.get('exceeds_limit')) or 1
        for i, it in enumerate(items):
            row_idx = i + 2
            vals = [it.get('sku', '-'), it.get('qty', 1),
                    f'{it.get("length_cm",0)}×{it.get("width_cm",0)}×{it.get("height_cm",0)}',
                    it.get('weight_kg', 0), round(it.get('volume_cbm', 0), 4),
                    it.get('carrier', '') if not it.get('exceeds_limit') else '超限']
            # 最优渠道费用
            if it.get('exceeds_limit'):
                vals += ['', '', '', '', '', '']
            else:
                vals += [round(it.get('base_freight', 0), 2),
                         it.get('oversize_type') or '-',
                         '; '.join(it.get('oversize_triggers') or []) if it.get('oversize_triggers') else '-',
                         round(it.get('oversize_fee_amount', 0), 2),
                         round(it.get('fuel_fee', 0), 2),
                         round(it.get('base_freight', 0) + it.get('oversize_fee_amount', 0) + it.get('fuel_fee', 0), 2)]
            ac = it.get('all_carriers', {})
            for c in carriers_list:
                d = ac.get(c) or {}
                vals.append(round(d.get('base', 0), 2) if not it.get('exceeds_limit') else '')
                vals.append(round(d.get('oversize', 0), 2) if not it.get('exceeds_limit') else '')
                vals.append(round(d.get('fuel', 0), 2) if not it.get('exceeds_limit') else '')
            u = round(wh_data.get('unloading_fee', 0) / valid_n, 2)
            rc = round(wh_data.get('receiving_fee', 0) / valid_n, 2)
            ob = round(wh_data.get('outbound_fee', 0) / valid_n, 2)
            st = round(wh_data.get('storage_fee', 0) / valid_n, 2)
            pk = round(wh_data.get('peak_surcharge', 0) / valid_n, 2)
            shared = u + rc + ob + st + pk
            # 各渠道最终费用
            for c in carriers_list:
                d = ac.get(c) or {}
                transport = round((d.get('base', 0) if not it.get('exceeds_limit') else 0) +
                                  (d.get('oversize', 0) if not it.get('exceeds_limit') else 0) +
                                  (d.get('fuel', 0) if not it.get('exceeds_limit') else 0), 2)
                vals.append(round(transport + shared, 2) if not it.get('exceeds_limit') else '')
            # 仓库分摊费用
            if it.get('exceeds_limit'):
                vals += ['', '', '', '', '']
            else:
                vals += [u, rc, ob, st, pk]
            for ci, v in enumerate(vals, 1):
                cell = ws2.cell(row=row_idx, column=ci, value=v)
                cell.font = nf; cell.border = border
        ws2.column_dimensions['A'].width = 18

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name='德国最优尾程测算.xlsx')


@app.route('/api/export/uk-last-mile-all', methods=['POST'])
def export_uk_last_mile_all():
    """导出英国最优模式（2个仓库各自一个Sheet）"""
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    data = request.get_json(force=True)
    warehouses = data.get('warehouses', {})
    if not warehouses:
        return jsonify({'error': '无仓库数据'}), 400

    hf = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
    hfill = PatternFill('solid', fgColor='4472C4')
    nf = Font(name='微软雅黑', size=10)
    tf = Font(name='微软雅黑', bold=True, size=11, color='006100')
    tfill = PatternFill('solid', fgColor='D9F2D9')
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))

    wb = Workbook()
    wb.remove(wb.active)

    for wh_name in UK_WAREHOUSES:
        wh_data = warehouses.get(wh_name)
        if not wh_data:
            continue
        ct = wh_data.get('carrier_totals', {})
        best = wh_data.get('best_carrier', '')
        if not ct:
            continue
        carriers = list(ct.keys())

        # 费用汇总
        ws = wb.create_sheet(f'{wh_name}-汇总')
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=1 + len(carriers))
        ws.cell(row=1, column=1, value=f'英国尾程测算 - {wh_name}').font = Font(name='微软雅黑', bold=True, size=14)

        ws.cell(row=3, column=1, value='费用项目').font = hf; ws.cell(row=3, column=1).fill = hfill; ws.cell(row=3, column=1).border = border
        for ci, c in enumerate(carriers):
            col = ci + 2
            label = c + (' ★' if c == best else '')
            cell = ws.cell(row=3, column=col, value=label)
            cell.font = hf; cell.fill = hfill; cell.border = border; cell.alignment = Alignment(horizontal='center')

        def _uv(c, k):
            d = ct.get(c) or {}
            return d.get(k, 0)

        fee_rows = [('基础运费 (£)', 'base'), ('超规附加费 (£)', 'oversize'), ('燃油附加费 (£)', 'fuel')]
        r = 4
        for name, key in fee_rows:
            ws.cell(row=r, column=1, value=name).font = nf; ws.cell(row=r, column=1).border = border
            for ci, c in enumerate(carriers):
                cell = ws.cell(row=r, column=ci + 2, value=round(_uv(c, key), 2))
                cell.font = nf; cell.border = border; cell.number_format = '#,##0.00'
            r += 1
        # 运输小计
        ws.cell(row=r, column=1, value='运输小计 (£)').font = Font(name='微软雅黑', bold=True, size=10); ws.cell(row=r, column=1).border = border
        for ci, c in enumerate(carriers):
            v = round(_uv(c, 'base') + _uv(c, 'oversize') + _uv(c, 'fuel'), 2)
            cell = ws.cell(row=r, column=ci + 2, value=v)
            cell.font = Font(name='微软雅黑', bold=True, size=10); cell.border = border; cell.number_format = '#,##0.00'
        r += 1
        # 仓库费用
        wh_fees = [
            ('卸货费 (£)', wh_data.get('unloading_fee', 0)),
            ('查验费 (£)', wh_data.get('inspection_fee', 0)),
            ('入库费 (£)', wh_data.get('receiving_fee', 0)),
            ('出库费 (£)', wh_data.get('outbound_fee', 0)),
            ('仓储费 (£)', wh_data.get('storage_fee', 0)),
            ('旺季附加费 (£)', wh_data.get('peak_surcharge', 0)),
        ]
        for name, val in wh_fees:
            ws.cell(row=r, column=1, value=name).font = nf; ws.cell(row=r, column=1).border = border
            for ci in range(len(carriers)):
                cell = ws.cell(row=r, column=ci + 2, value=round(val, 2))
                cell.font = nf; cell.border = border; cell.number_format = '#,##0.00'
            r += 1
        # 总费用
        ws.cell(row=r, column=1, value='总费用 (£)').font = tf; ws.cell(row=r, column=1).fill = tfill; ws.cell(row=r, column=1).border = border
        for ci, c in enumerate(carriers):
            t = round(_uv(c, 'base') + _uv(c, 'oversize') + _uv(c, 'fuel') +
                      wh_data.get('unloading_fee', 0) + wh_data.get('inspection_fee', 0) +
                      wh_data.get('receiving_fee', 0) + wh_data.get('outbound_fee', 0) +
                      wh_data.get('storage_fee', 0) + wh_data.get('peak_surcharge', 0), 2)
            cell = ws.cell(row=r, column=ci + 2, value=t)
            cell.font = tf; cell.fill = tfill; cell.border = border; cell.number_format = '#,##0.00'

        ws.column_dimensions['A'].width = 22
        for ci in range(len(carriers)):
            ws.column_dimensions[chr(66 + ci)].width = 14

        # 产品明细
        ws2 = wb.create_sheet(f'{wh_name}-明细')
        headers = ['SKU', '数量', '体积(m³)', '实重(kg)', '最优渠道',
                   '基础运费', '超规类型', '超规详情', '超规费', '燃油费', '小计']
        for c in carriers:
            headers += [c + '基础', c + '超规', c + '燃油']
        for c in carriers:
            headers += [c + '最终费用']
        headers += ['卸货分摊', '查验分摊', '入库费', '出库费', '仓储分摊', '旺季分摊']
        for col, h in enumerate(headers, 1):
            cell = ws2.cell(row=1, column=col, value=h)
            cell.font = hf; cell.fill = hfill; cell.alignment = Alignment(horizontal='center'); cell.border = border

        items = wh_data.get('items', [])
        valid_n = sum(1 for it in items if not it.get('exceeds_limit')) or 1
        for i, it in enumerate(items):
            row_idx = i + 2
            vals = [it.get('sku', '-'), it.get('qty', 1), round(it.get('volume_cbm', 0), 4),
                    it.get('weight_kg', 0), it.get('carrier') if not it.get('exceeds_limit') else '超限']
            # 最优渠道费用
            if it.get('exceeds_limit'):
                vals += ['', '', '', '', '']
            else:
                vals += [round(it.get('base_freight', 0), 2),
                         it.get('oversize_type') or '-',
                         '; '.join(it.get('oversize_triggers') or []) if it.get('oversize_triggers') else '-',
                         round(it.get('oversize_fee_amount', 0), 2),
                         round(it.get('fuel_fee', 0), 2),
                         round(it.get('base_freight', 0) + it.get('oversize_fee_amount', 0) + it.get('fuel_fee', 0), 2)]
            ac = it.get('all_carriers', {})
            for c in carriers:
                d = ac.get(c) or {}
                vals.append(round(d.get('base', 0), 2) if not it.get('exceeds_limit') else '')
                vals.append(round(d.get('oversize', 0), 2) if not it.get('exceeds_limit') else '')
                vals.append(round(d.get('fuel', 0), 2) if not it.get('exceeds_limit') else '')
            u = round(wh_data.get('unloading_fee', 0) / valid_n, 2)
            insp = round(wh_data.get('inspection_fee', 0) / valid_n, 2)
            rc = round(wh_data.get('receiving_fee', 0) / valid_n, 2)
            ob = round(wh_data.get('outbound_fee', 0) / valid_n, 2)
            st = round(wh_data.get('storage_fee', 0) / valid_n, 2)
            pk = round(wh_data.get('peak_surcharge', 0) / valid_n, 2)
            shared = u + insp + rc + ob + st + pk
            # 各渠道最终费用
            for c in carriers:
                d = ac.get(c) or {}
                transport = round((d.get('base', 0) if not it.get('exceeds_limit') else 0) +
                                  (d.get('oversize', 0) if not it.get('exceeds_limit') else 0) +
                                  (d.get('fuel', 0) if not it.get('exceeds_limit') else 0), 2)
                vals.append(round(transport + shared, 2) if not it.get('exceeds_limit') else '')
            # 仓库分摊费用
            if it.get('exceeds_limit'):
                vals += ['', '', '', '', '', '', '']
            else:
                vals += [u, insp, rc, ob, st, pk]
            for ci, v in enumerate(vals, 1):
                cell = ws2.cell(row=row_idx, column=ci, value=v)
                cell.font = nf; cell.border = border
        ws2.column_dimensions['A'].width = 18

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name='英国最优尾程测算.xlsx')


@app.route('/api/export/last-mile-pdf', methods=['POST'])
def export_last_mile_pdf():
    """导出尾程测算结果为 PDF"""
    from io import BytesIO
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib.colors import HexColor
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

    _ensure_cn_font()
    FONT_C = _CN_FONT_NAME
    FONT_B = _CN_FONT_NAME

    data = request.get_json(force=True)
    if not data:
        return jsonify({'error': '无导出数据'}), 400

    warehouses = data.get('warehouses', {})
    if not warehouses:
        return jsonify({'error': '无仓库数据'}), 400

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=15*mm, bottomMargin=15*mm,
                            leftMargin=12*mm, rightMargin=12*mm)
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle('LM_Title', parent=styles['Title'],
                                 fontName=FONT_B, fontSize=16, spaceAfter=6)
    h2_style = ParagraphStyle('LM_H2', parent=styles['Heading2'],
                              fontName=FONT_B, fontSize=13, spaceAfter=4,
                              textColor=HexColor('#1e40af'))
    small_style = ParagraphStyle('LM_Small', parent=styles['Normal'],
                                 fontName=FONT_C, fontSize=7, leading=9,
                                 textColor=HexColor('#6b7280'))
    subh_style = ParagraphStyle('LM_SubH', parent=styles['Heading3'],
                                fontName=FONT_B, fontSize=10, spaceAfter=3)

    elements = []
    elements.append(Paragraph('尾程测算报告', title_style))

    is_multi = len(warehouses) > 1
    if is_multi:
        elements.append(Paragraph('多仓库对比模式', small_style))
    elements.append(Spacer(1, 4*mm))

    for wh_name in list(warehouses.keys()):
        wh = warehouses[wh_name]
        zone = wh.get('zone', '')
        total_qty = wh.get('total_qty', 0)
        total_cbm = wh.get('total_cbm', 0)

        elements.append(Paragraph(f'{wh_name}  Zone {zone}', h2_style))
        elements.append(Paragraph(
            f'总件数: {total_qty} | 总体积: {total_cbm:.4f} m3 | 仓储天数: {wh.get("storage_days", 0)}',
            small_style
        ))

        # 费用汇总表
        fees = [
            ['费用项目', '金额 (USD)', '说明'],
            ['基础运费', f'{wh.get("base_freight",0):,.2f}', f'Zone {zone}，共 {total_qty} 件'],
            ['超尺寸费', f'{wh.get("oversize_fee",0):,.2f}', ''],
            ['住宅附加费', f'{wh.get("residential_surcharge",0):,.2f}', f'{total_qty} 件'],
            ['卸货费', f'{wh.get("unloading_fee",0):,.2f}', wh.get('unloading_detail', {}).get('type', '')],
            ['入库上架费', f'{wh.get("receiving_fee",0):,.2f}', '按各产品实重分别计费'],
            ['出库费', f'{wh.get("outbound_fee",0):,.2f}', f'{total_qty} 件'],
            ['仓储费', f'{wh.get("storage_fee",0):,.2f}', f'{wh.get("storage_days",0)} 天'],
            ['偏远附加费', f'{wh.get("remote_surcharge",0):,.2f}', ''],
            ['燃油附加费', f'{wh.get("fuel_surcharge",0):,.2f}', f'{wh.get("fuel_rate_pct",0)}%'],
            ['旺季附加费', f'{wh.get("peak_surcharge",0):,.2f}', ''],
            ['总费用', f'{wh.get("grand_total",0):,.2f}', 'USD'],
        ]

        t = Table(fees, colWidths=[90, 65, 200])
        t.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), FONT_C),
            ('FONTNAME', (0, 0), (-1, 0), FONT_B),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('BACKGROUND', (0, 0), (-1, 0), HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), HexColor('#FFFFFF')),
            ('BACKGROUND', (0, -1), (-1, -1), HexColor('#D9F2D9')),
            ('FONTNAME', (0, -1), (-1, -1), FONT_B),
            ('GRID', (0, 0), (-1, -1), 0.5, HexColor('#CBD5E1')),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ]))
        elements.append(t)
        elements.append(Spacer(1, 4*mm))

        # 产品明细
        elements.append(Paragraph('产品明细', subh_style))
        items = wh.get('items', [])
        detail_data = [['SKU', '数量', '尺寸(cm)', '实重(kg)', '材积重(lb)', '计费重(lb)', '基础运费', '超尺寸费']]
        for item in items:
            detail_data.append([
                item.get('sku', '-'),
                str(item.get('qty', 1)),
                f'{item.get("length_cm",0):.1f}x{item.get("width_cm",0):.1f}x{item.get("height_cm",0):.1f}',
                f'{item.get("weight_kg",0):.2f}',
                f'{item.get("dim_weight_lb",0):.1f}',
                f'{item.get("billing_weight_final",0):.1f}',
                f'{item.get("base_freight",0):,.2f}',
                f'{item.get("oversize_fee_amount",0):,.2f}',
            ])

        col_w = [50, 30, 55, 35, 35, 35, 40, 40]
        dt = Table(detail_data, colWidths=col_w)
        dt.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), FONT_C),
            ('FONTNAME', (0, 0), (-1, 0), FONT_B),
            ('FONTSIZE', (0, 0), (-1, -1), 7),
            ('BACKGROUND', (0, 0), (-1, 0), HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), HexColor('#FFFFFF')),
            ('GRID', (0, 0), (-1, -1), 0.5, HexColor('#CBD5E1')),
            ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 2),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
        ]))
        elements.append(dt)
        elements.append(Spacer(1, 6*mm))

    doc.build(elements)
    buf.seek(0)

    return send_file(
        buf,
        mimetype='application/pdf',
        as_attachment=True,
        download_name=f'尾程测算报告_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
    )


# ============================================================
#  利润测算 API
# ============================================================

def fetch_exchange_rate():
    """获取美元兑人民币汇率（优先从shishihuilv.com爬取，失败后依次使用API备选）"""
    import urllib.request
    import urllib.error
    import re

    def try_parse_rate(text):
        """从文本中尝试提取汇率数字"""
        if not text:
            return None
        m = re.search(r'(\d+\.?\d*)', str(text))
        if m:
            val = float(m.group(1))
            if 5 < val < 9:
                return round(val, 4)
        return None

    # 1) 优先从 shishihuilv.com 爬取
    try:
        url = 'https://www.shishihuilv.com/'
        req = urllib.request.Request(url, headers={
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
            'Accept-Language': 'zh-CN,zh;q=0.9,en;q=0.8',
        })
        with urllib.request.urlopen(req, timeout=15) as resp:
            html = resp.read().decode('utf-8', errors='ignore')
            # 匹配 <p class="to">6.8005 人民币（CNY）</p>
            m = re.search(r'<p[^>]*class=["\']to["\'][^>]*>\s*([\d.]+)', html)
            if m:
                rate = float(m.group(1))
                if 5 < rate < 9:
                    return round(rate, 4)
    except Exception:
        pass

    # 2) 备选: exchangerate-api 免费接口
    try:
        url = 'https://open.er-api.com/v6/latest/USD'
        req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
        with urllib.request.urlopen(req, timeout=10) as resp:
            data = json.loads(resp.read().decode('utf-8'))
            rate = data.get('rates', {}).get('CNY')
            if rate and rate > 0:
                return round(rate, 4)
    except Exception:
        pass

    # 3) 备选: exchangerate-api v4
    try:
        url = 'https://api.exchangerate-api.com/v4/latest/USD'
        req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
        with urllib.request.urlopen(req, timeout=10) as resp:
            data = json.loads(resp.read().decode('utf-8'))
            rate = data.get('rates', {}).get('CNY')
            if rate and rate > 0:
                return round(rate, 4)
    except Exception:
        pass

    # 默认值
    return 7.25


@app.route('/api/calc/exchange-rate', methods=['GET'])
def api_get_exchange_rate():
    """获取美元兑人民币实时汇率"""
    try:
        rate = fetch_exchange_rate()
        return jsonify({'rate': rate, 'currency_pair': 'USD/CNY', 'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


def calculate_profit(data):
    """利润测算核心计算
    输入参数:
    - items: [{sku, qty, volume_cbm, last_mile_cost_excl_storage, storage_cost, purchase_price_manual}]
    - exchange_rate: 美元汇率
    - tax_per_cbm: 每立方米税务成本（美元），分摊公式：a × 立方米体积
    - management_rate: 管理费率（小数），默认0.04
    - complaint_rate: 客诉费率（小数），默认0.05
    - platform_rate: 平台费率（小数），默认0.15
    - ad_rate: 广告费率（小数），默认0.15
    - profit_rate: 目标利润率（小数），默认0.30
    """
    items = data.get('items', [])
    exchange_rate = float(data.get('exchange_rate', 7.25))
    tax_per_cbm = float(data.get('tax_per_cbm', data.get('tax_total', 0)))
    management_rate = float(data.get('management_rate', 0.04))
    complaint_rate = float(data.get('complaint_rate', 0.05))
    platform_rate = float(data.get('platform_rate', 0.15))
    ad_rate = float(data.get('ad_rate', 0.15))
    profit_rate = float(data.get('profit_rate', 0.30))

    if not items:
        return {'error': '请至少添加一个产品'}

    if exchange_rate <= 0:
        return {'error': '汇率必须大于0'}

    if platform_rate + ad_rate + profit_rate >= 1:
        return {'error': f'平台费率({platform_rate*100}%) + 广告费率({ad_rate*100}%) + 利润率({profit_rate*100}%) = {(platform_rate+ad_rate+profit_rate)*100}% 超过了100%，无法计算'}

    # 获取配置参数：头程基数
    conn = db_connect()
    c = conn.cursor()
    c.execute("SELECT value FROM 配置参数 WHERE key = 'first_leg_base'")
    row = c.fetchone()
    first_leg_base = int(float(row[0])) if row else 600
    conn.close()

    # 先收集所有缺少采购成本的产品
    missing_cost_skus = []
    # 预解析所有产品的采购价，存入 _purchase_price
    for item in items:
        sku = item.get('sku', '-')
        purchase_price = item.get('purchase_price_manual')
        if purchase_price is None and sku and sku != '-':
            try:
                conn2 = db_connect()
                c2 = conn2.cursor()
                c2.execute('SELECT 采购价 FROM 产品信息 WHERE SKU = ? LIMIT 1', (sku,))
                row2 = c2.fetchone()
                conn2.close()
                if row2 and row2[0]:
                    purchase_price = float(row2[0])
            except Exception:
                pass
        # 采购成本为0或未找到时记录
        if not purchase_price or purchase_price <= 0:
            missing_cost_skus.append(sku)
        item['_purchase_price'] = float(purchase_price) if purchase_price else 0.0

    if missing_cost_skus:
        return {'error': f'以下产品缺少采购成本，请在"产品明细"中手动填写采购成本($)后重新测算：{", ".join(missing_cost_skus)}'}

    # 计算总体积（用于显示）
    total_cbm = sum(item.get('volume_cbm', 0) * item.get('qty', 1) for item in items)
    if total_cbm <= 0:
        total_cbm = 1  # 避免除零

    item_results = []

    for item in items:
        sku = item.get('sku', '-')
        qty = int(item.get('qty', 1))
        volume_cbm = float(item.get('volume_cbm', 0))

        # 1. 采购成本（已在预解析中处理好）
        purchase_price = item.get('_purchase_price', 0)
        # 采购价数据库存的是人民币，转为美元统一口径
        purchase_cost = round(float(purchase_price) * qty / exchange_rate, 2)

        # 2. 头程成本 = 头程基数 * 立方米体积 / 美元汇率
        head_leg_cost = round(first_leg_base * volume_cbm * qty / exchange_rate, 2)

        # 3. 税务成本 = 每立方米税务成本 × 体积 × 数量
        if tax_per_cbm > 0:
            tax_cost = round(tax_per_cbm * volume_cbm * qty, 2)
        else:
            tax_cost = 0

        # 4. 尾程成本（不含仓储费），除以数量折算为单件成本
        last_mile_cost = round(float(item.get('last_mile_cost_excl_storage', 0)) / qty, 2)

        # 5. 仓储成本，除以数量折算为单件成本
        storage_cost = round(float(item.get('storage_cost', 0)) / qty, 2)

        # 前5项合计
        subtotal_5 = purchase_cost + head_leg_cost + tax_cost + last_mile_cost + storage_cost

        # 6. 管理成本 = management_rate * subtotal_5
        management_cost = round(management_rate * subtotal_5, 2)

        # 7. 客诉成本 = complaint_rate * subtotal_5
        complaint_cost = round(complaint_rate * subtotal_5, 2)

        # 固定成本 = 7项相加
        fixed_cost = purchase_cost + head_leg_cost + tax_cost + last_mile_cost + storage_cost + management_cost + complaint_cost

        # 价格计算：price = fixed_cost / (1 - platform_rate - ad_rate - profit_rate)
        denominator = 1 - platform_rate - ad_rate - profit_rate
        price = round(fixed_cost / denominator, 2) if denominator > 0 else 0

        # 利润 = price * profit_rate
        profit = round(price * profit_rate, 2)

        # 利润率 = 利润 / 价格
        calc_profit_rate = round(profit / price, 4) if price > 0 else 0

        item_results.append({
            'sku': sku,
            'qty': qty,
            'volume_cbm': round(volume_cbm, 4),
            'purchase_cost': purchase_cost,
            'purchase_price': purchase_price,
            'head_leg_cost': head_leg_cost,
            'tax_cost': tax_cost,
            'last_mile_cost': last_mile_cost,
            'storage_cost': storage_cost,
            'management_cost': management_cost,
            'complaint_cost': complaint_cost,
            'subtotal_5': subtotal_5,
            'fixed_cost': fixed_cost,
            'price': price,
            'profit': profit,
            'profit_rate': calc_profit_rate,
        })

    # 汇总
    total_purchase = sum(r['purchase_cost'] for r in item_results)
    total_head_leg = sum(r['head_leg_cost'] for r in item_results)
    total_tax = sum(r['tax_cost'] for r in item_results)
    total_last_mile = sum(r['last_mile_cost'] for r in item_results)
    total_storage = sum(r['storage_cost'] for r in item_results)
    total_management = sum(r['management_cost'] for r in item_results)
    total_complaint = sum(r['complaint_cost'] for r in item_results)
    total_fixed = sum(r['fixed_cost'] for r in item_results)
    total_price = sum(r['price'] for r in item_results)
    total_profit = sum(r['profit'] for r in item_results)

    return {
        'items': item_results,
        'total_qty': sum(r['qty'] for r in item_results),
        'total_cbm': round(total_cbm, 4),
        'exchange_rate': exchange_rate,
        'first_leg_base': first_leg_base,
        'tax_per_cbm': tax_per_cbm,
        'management_rate': round(management_rate * 100, 1),
        'complaint_rate': round(complaint_rate * 100, 1),
        'platform_rate': round(platform_rate * 100, 1),
        'ad_rate': round(ad_rate * 100, 1),
        'target_profit_rate': round(profit_rate * 100, 1),
        'summary': {
            'purchase_cost': total_purchase,
            'head_leg_cost': total_head_leg,
            'tax_cost': total_tax,
            'last_mile_cost': total_last_mile,
            'storage_cost': total_storage,
            'management_cost': total_management,
            'complaint_cost': total_complaint,
            'fixed_cost': total_fixed,
            'price': total_price,
            'profit': total_profit,
            'profit_rate': round(total_profit / total_price, 4) if total_price > 0 else 0,
        }
    }


@app.route('/api/calc/profit', methods=['POST'])
def api_calc_profit():
    """利润测算API"""
    try:
        data = request.get_json(force=True)
        if not data:
            return jsonify({'error': '请提供测算参数'}), 400

        result = calculate_profit(data)
        if 'error' in result:
            return jsonify(result), 400

        return jsonify(result)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ============================================================
#  多仓库利润测算 API（最优模式）
# ============================================================

US_WAREHOUSES = ['海智链', '大健云仓', '发现']

@app.route('/api/calc/profit-all', methods=['POST'])
def api_calc_profit_all():
    """一键测算美国3个仓库的利润，返回最优仓库标记"""
    try:
        data = request.get_json(force=True)
        if not data:
            return jsonify({'error': '请提供测算参数'}), 400

        items_input = data.get('items', [])
        if not items_input:
            return jsonify({'error': '请至少添加一个产品'}), 400

        zone = int(data.get('zone', 8))
        storage_days = float(data.get('storage_days', 0))
        unloading_type = data.get('unloading_type', 'none')
        remote_mode = data.get('remote_surcharge_mode', 'none')
        peak_mode = data.get('peak_surcharge_mode', 'none')
        fuel_rate_pct = data.get('fuel_rate_pct')

        exchange_rate = float(data.get('exchange_rate', 7.25))
        tax_per_cbm = float(data.get('tax_per_cbm', data.get('tax_total', 0)))
        management_rate = float(data.get('management_rate', 0.04))
        complaint_rate = float(data.get('complaint_rate', 0.05))
        platform_rate = float(data.get('platform_rate', 0.15))
        ad_rate = float(data.get('ad_rate', 0.15))
        profit_rate = float(data.get('profit_rate', 0.30))

        warehouse_results = {}
        optimal_wh = None
        optimal_total_price = float('inf')

        for wh_name in US_WAREHOUSES:
            # 1) 对每个仓库执行尾程测算
            lm_input = {
                'items': items_input,
                'zone': zone,
                'warehouse': wh_name,
                'unloading_type': unloading_type,
                'storage_days': storage_days,
                'remote_surcharge_mode': remote_mode,
                'peak_surcharge_mode': peak_mode,
                'fuel_rate_pct': fuel_rate_pct
            }
            lm_result = calculate_last_mile(lm_input)
            if 'error' in lm_result:
                return jsonify({'error': f'{wh_name} 尾程测算失败: {lm_result["error"]}'}), 400

            # 2) 从尾程结果提取每件产品的尾程成本（不含仓储）和仓储成本
            total_cbm = lm_result.get('total_cbm', 1) or 1
            total_qty = lm_result.get('total_qty', 1) or 1
            total_storage = lm_result.get('storage_fee', 0)

            profit_items = []
            for lm_item in lm_result.get('items', []):
                orig_qty = lm_item.get('qty', 1)
                sku = lm_item.get('sku', '-')
                volume_cbm = float(lm_item.get('volume_cbm', 0))

                per_item_cost = ((lm_item.get('base_freight', 0) or 0) +
                                 (lm_item.get('oversize_fee_amount', 0) or 0) +
                                 (lm_item.get('receiving_fee', 0) or 0)) / orig_qty

                volume_ratio = (volume_cbm * 1) / total_cbm if total_cbm > 0 else (1.0 / len(lm_result['items']))
                unloading_share = (lm_result.get('unloading_fee', 0) or 0) * volume_ratio
                residential_share = (lm_result.get('residential_surcharge', 0) or 0) / total_qty
                outbound_share = (lm_result.get('outbound_fee', 0) or 0) / total_qty
                remote_share = (lm_result.get('remote_surcharge', 0) or 0) * volume_ratio
                fuel_share = (lm_result.get('fuel_surcharge', 0) or 0) * volume_ratio
                peak_share = (lm_result.get('peak_surcharge', 0) or 0) * volume_ratio

                last_mile_excl = per_item_cost + unloading_share + residential_share + outbound_share + remote_share + fuel_share + peak_share
                storage_cost = total_storage * volume_ratio

                profit_items.append({
                    'sku': sku,
                    'qty': 1,
                    'volume_cbm': round(volume_cbm, 6),
                    'last_mile_cost_excl_storage': round(last_mile_excl, 2),
                    'storage_cost': round(storage_cost, 2),
                })

            # 3) 执行利润测算
            pc_input = {
                'items': profit_items,
                'exchange_rate': exchange_rate,
                'tax_per_cbm': tax_per_cbm,
                'management_rate': management_rate,
                'complaint_rate': complaint_rate,
                'platform_rate': platform_rate,
                'ad_rate': ad_rate,
                'profit_rate': profit_rate
            }
            pc_result = calculate_profit(pc_input)
            if 'error' in pc_result:
                return jsonify({'error': f'{wh_name} 利润测算失败: {pc_result["error"]}'}), 400

            warehouse_results[wh_name] = {
                'last_mile': lm_result,
                'profit': pc_result,
            }

            total_price = pc_result['summary']['price']
            if total_price < optimal_total_price:
                optimal_total_price = total_price
                optimal_wh = wh_name

        return jsonify({
            'warehouses': warehouse_results,
            'optimal': optimal_wh,
            'params': {
                'zone': zone,
                'storage_days': storage_days,
                'unloading_type': unloading_type,
                'exchange_rate': exchange_rate,
                'tax_per_cbm': tax_per_cbm,
                'management_rate': round(management_rate * 100, 1),
                'complaint_rate': round(complaint_rate * 100, 1),
                'platform_rate': round(platform_rate * 100, 1),
                'ad_rate': round(ad_rate * 100, 1),
                'profit_rate': round(profit_rate * 100, 1),
            }
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ============================================================
#  德国多仓库尾程测算 API（最优模式）
# ============================================================

DE_WAREHOUSES = ['金仓', '欧品居', '易达云']

@app.route('/api/calc/de/last-mile-all', methods=['POST'])
def api_calc_de_last_mile_all():
    """一键测算德国3个仓库的尾程费用，返回费用对比"""
    try:
        data = request.get_json(force=True)
        if not data:
            return jsonify({'error': '请提供测算参数'}), 400

        items_input = data.get('items', [])
        if not items_input:
            return jsonify({'error': '请至少添加一个产品'}), 400

        unloading_type = data.get('unloading_type', 'none')
        storage_days = float(data.get('storage_days', 0))
        peak_mode = data.get('peak_mode', 'none')
        # 读取所有9个仓库×渠道的燃油附加费率
        fuel_rate_jincang_dpd = float(data.get('fuel_rate_jincang_dpd', 0))
        fuel_rate_jincang_dhl = float(data.get('fuel_rate_jincang_dhl', 0))
        fuel_rate_jincang_gls = float(data.get('fuel_rate_jincang_gls', 0))
        fuel_rate_oupinju_dpd = float(data.get('fuel_rate_oupinju_dpd', 0))
        fuel_rate_oupinju_dhl = float(data.get('fuel_rate_oupinju_dhl', 0))
        fuel_rate_oupinju_gls = float(data.get('fuel_rate_oupinju_gls', 0))
        fuel_rate_yidayun_dpd = float(data.get('fuel_rate_yidayun_dpd', 0))
        fuel_rate_yidayun_dhl = float(data.get('fuel_rate_yidayun_dhl', 0))
        fuel_rate_yidayun_gls = float(data.get('fuel_rate_yidayun_gls', 0))

        # 仓库→燃油费率映射
        WH_FUEL_RATES = {
            '金仓': {'fuel_rate_jincang_dpd': fuel_rate_jincang_dpd,
                      'fuel_rate_jincang_dhl': fuel_rate_jincang_dhl,
                      'fuel_rate_jincang_gls': fuel_rate_jincang_gls},
            '欧品居': {'fuel_rate_oupinju_dpd': fuel_rate_oupinju_dpd,
                       'fuel_rate_oupinju_dhl': fuel_rate_oupinju_dhl,
                       'fuel_rate_oupinju_gls': fuel_rate_oupinju_gls},
            '易达云': {'fuel_rate_yidayun_dpd': fuel_rate_yidayun_dpd,
                       'fuel_rate_yidayun_dhl': fuel_rate_yidayun_dhl,
                       'fuel_rate_yidayun_gls': fuel_rate_yidayun_gls},
        }

        warehouse_results = {}
        optimal_wh = None
        optimal_cost = float('inf')

        for wh_name in DE_WAREHOUSES:
            wh_fuels = WH_FUEL_RATES[wh_name]
            lm_input = {
                'items': items_input,
                'warehouse': wh_name,
                'unloading_type': unloading_type,
                'storage_days': storage_days,
                'peak_mode': peak_mode,
                **wh_fuels,
            }
            lm_result = calc_last_mile_de(lm_input)
            if 'error' in lm_result:
                return jsonify({'error': f'{wh_name} 尾程测算失败: {lm_result["error"]}'}), 400

            warehouse_results[wh_name] = lm_result
            if lm_result.get('grand_total', float('inf')) < optimal_cost:
                optimal_cost = lm_result['grand_total']
                optimal_wh = wh_name

        return jsonify({
            'warehouses': warehouse_results,
            'optimal': optimal_wh,
            'params': {
                'unloading_type': unloading_type,
                'storage_days': storage_days,
                'peak_mode': peak_mode,
                'fuel_rate_jincang_dpd': fuel_rate_jincang_dpd,
                'fuel_rate_jincang_dhl': fuel_rate_jincang_dhl,
                'fuel_rate_jincang_gls': fuel_rate_jincang_gls,
                'fuel_rate_oupinju_dpd': fuel_rate_oupinju_dpd,
                'fuel_rate_oupinju_dhl': fuel_rate_oupinju_dhl,
                'fuel_rate_oupinju_gls': fuel_rate_oupinju_gls,
                'fuel_rate_yidayun_dpd': fuel_rate_yidayun_dpd,
                'fuel_rate_yidayun_dhl': fuel_rate_yidayun_dhl,
                'fuel_rate_yidayun_gls': fuel_rate_yidayun_gls,
            }
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ============================================================
#  利润测算 Excel 导出（多仓库）
# ============================================================

@app.route('/api/export/profit-xlsx', methods=['POST'])
def export_profit_xlsx():
    """导出最优测算结果为 Excel（每个仓库：总览 + 明细）"""
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    data = request.get_json(force=True)
    if not data:
        return jsonify({'error': '无导出数据'}), 400

    warehouses = data.get('warehouses', {})
    if not warehouses:
        return jsonify({'error': '无仓库数据'}), 400

    params = data.get('params', {})

    wb = Workbook()
    wb.remove(wb.active)

    header_font = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='4472C4')
    total_font = Font(name='微软雅黑', bold=True, size=12, color='006100')
    total_fill = PatternFill('solid', fgColor='D9F2D9')
    normal_font = Font(name='微软雅黑', size=10)
    sub_font = Font(name='微软雅黑', size=9, color='808080')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')

    for wh_name in US_WAREHOUSES:
        wh_data = warehouses.get(wh_name)
        if not wh_data:
            continue

        profit = wh_data.get('profit', {})
        items = profit.get('items', [])
        summary = profit.get('summary', {})

        # --- Sheet 1: 总览 ---
        ws1 = wb.create_sheet(f'{wh_name}-总览')
        ws1.merge_cells('A1:D1')
        title_cell = ws1.cell(row=1, column=1, value=f'{wh_name} 利润测算总览')
        title_cell.font = Font(name='微软雅黑', bold=True, size=14)

        ws1.merge_cells('A3:D3')
        ws1.cell(row=3, column=1, value=f'Zone: {params.get("zone","")} | 仓储天数: {params.get("storage_days",0)} | 汇率: {params.get("exchange_rate","")}').font = sub_font

        overview_headers = ['费用项目', '金额 (USD)', '说明', '']
        r = 5
        for col, h in enumerate(overview_headers, 1):
            cell = ws1.cell(row=r, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = center_align

        overview_rows = [
            ('1. 采购成本', summary.get('purchase_cost', 0), 'CNY采购价 / 汇率'),
            ('2. 头程成本', summary.get('head_leg_cost', 0), '头程基数 / 汇率 x 体积'),
            ('3. 税务成本', summary.get('tax_cost', 0), '每立方米税费 × 体积'),
            ('4. 尾程成本(不含仓储)', summary.get('last_mile_cost', 0), '各SKU尾程费用合计'),
            ('5. 仓储成本', summary.get('storage_cost', 0), f'{params.get("storage_days",0)} 天'),
            ('6. 管理成本', summary.get('management_cost', 0), f'{params.get("management_rate",0)}% x 前5项'),
            ('7. 客诉成本', summary.get('complaint_cost', 0), f'{params.get("complaint_rate",0)}% x 前5项'),
        ]
        r = 6
        for name, amount, note in overview_rows:
            ws1.cell(row=r, column=1, value=name).font = normal_font
            c2 = ws1.cell(row=r, column=2, value=round(amount, 2))
            c2.font = normal_font
            c2.number_format = '#,##0.00'
            ws1.cell(row=r, column=3, value=note).font = sub_font
            for col in range(1, 5):
                ws1.cell(row=r, column=col).border = thin_border
            r += 1

        summary_rows = [
            ('固定成本合计', summary.get('fixed_cost', 0), '7项之和'),
            ('平台费率', None, f'{params.get("platform_rate",0)}%'),
            ('广告费率', None, f'{params.get("ad_rate",0)}%'),
            ('目标利润率', None, f'{params.get("profit_rate",0)}%'),
            ('总价格', summary.get('price', 0), '= 固定成本 / (1-平台-广告-利润)'),
            ('总利润', summary.get('profit', 0), f'利润率 {round(summary.get("profit_rate", 0) * 100, 2)}%'),
        ]
        for name, amount, note in summary_rows:
            ws1.cell(row=r, column=1, value=name).font = Font(name='微软雅黑', bold=True, size=11)
            if amount is not None:
                c2 = ws1.cell(row=r, column=2, value=round(amount, 2))
                c2.font = Font(name='微软雅黑', bold=True, size=11)
                c2.number_format = '#,##0.00'
            ws1.cell(row=r, column=3, value=note).font = sub_font
            if name in ('总价格', '总利润'):
                ws1.cell(row=r, column=1).fill = total_fill
                if amount is not None:
                    ws1.cell(row=r, column=2).fill = total_fill
            for col in range(1, 5):
                ws1.cell(row=r, column=col).border = thin_border
            r += 1

        ws1.column_dimensions['A'].width = 22
        ws1.column_dimensions['B'].width = 16
        ws1.column_dimensions['C'].width = 34
        ws1.column_dimensions['D'].width = 5

        # --- Sheet 2: 明细 ---
        ws2 = wb.create_sheet(f'{wh_name}-明细')
        ws2.merge_cells('A1:N1')
        ws2.cell(row=1, column=1, value=f'{wh_name} 各SKU费用明细').font = Font(name='微软雅黑', bold=True, size=14)

        detail_headers = [
            'SKU', '数量', '体积(m3)', '采购成本', '头程成本', '税务成本',
            '尾程成本', '仓储成本', '管理成本', '客诉成本', '固定成本',
            '价格', '利润', '利润率'
        ]
        r = 3
        for col, h in enumerate(detail_headers, 1):
            cell = ws2.cell(row=r, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = center_align

        r = 4
        for item in items:
            vals = [
                item.get('sku', '-'),
                item.get('qty', 1),
                round(item.get('volume_cbm', 0), 6),
                round(item.get('purchase_cost', 0), 2),
                round(item.get('head_leg_cost', 0), 2),
                round(item.get('tax_cost', 0), 2),
                round(item.get('last_mile_cost', 0), 2),
                round(item.get('storage_cost', 0), 2),
                round(item.get('management_cost', 0), 2),
                round(item.get('complaint_cost', 0), 2),
                round(item.get('fixed_cost', 0), 2),
                round(item.get('price', 0), 2),
                round(item.get('profit', 0), 2),
                f'{round(item.get("profit_rate", 0) * 100, 2)}%',
            ]
            for col, v in enumerate(vals, 1):
                cell = ws2.cell(row=r, column=col, value=v)
                cell.font = normal_font
                cell.border = thin_border
                if col >= 2:
                    cell.alignment = Alignment(horizontal='right')
            r += 1

        # 合计行
        total_vals = [
            '合计', profit.get('total_qty', 0),
            round(profit.get('total_cbm', 0), 6),
            round(summary.get('purchase_cost', 0), 2),
            round(summary.get('head_leg_cost', 0), 2),
            round(summary.get('tax_cost', 0), 2),
            round(summary.get('last_mile_cost', 0), 2),
            round(summary.get('storage_cost', 0), 2),
            round(summary.get('management_cost', 0), 2),
            round(summary.get('complaint_cost', 0), 2),
            round(summary.get('fixed_cost', 0), 2),
            round(summary.get('price', 0), 2),
            round(summary.get('profit', 0), 2),
            f'{round(summary.get("profit_rate", 0) * 100, 2)}%',
        ]
        for col, v in enumerate(total_vals, 1):
            cell = ws2.cell(row=r, column=col, value=v)
            cell.font = total_font
            cell.fill = total_fill
            cell.border = thin_border
            if col >= 2:
                cell.alignment = Alignment(horizontal='right')

        ws2.column_dimensions['A'].width = 16
        for c_letter in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']:
            ws2.column_dimensions[c_letter].width = 12

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'利润测算最优结果_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    )


# ============================================================
#  利润测算 PDF 导出
# ============================================================

@app.route('/api/export/profit-pdf', methods=['POST'])
def export_profit_pdf():
    """导出利润测算结果为 PDF"""
    from io import BytesIO
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib.colors import HexColor
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak

    _ensure_cn_font()
    FONT_C = 'MicrosoftYaHei'
    FONT_B = 'MicrosoftYaHei'

    data = request.get_json(force=True)
    if not data:
        return jsonify({'error': '无导出数据'}), 400

    warehouses = data.get('warehouses', {})
    if not warehouses:
        return jsonify({'error': '无仓库数据'}), 400

    params = data.get('params', {})
    optimal = data.get('optimal', '')

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=15*mm, bottomMargin=15*mm,
                            leftMargin=12*mm, rightMargin=12*mm)
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle('Title_CN', parent=styles['Title'],
                                 fontName=FONT_B, fontSize=16, spaceAfter=6)
    h2_style = ParagraphStyle('H2_CN', parent=styles['Heading2'],
                              fontName=FONT_B, fontSize=13, spaceAfter=4,
                              textColor=HexColor('#1e40af'))
    small_style = ParagraphStyle('Small_CN', parent=styles['Normal'],
                                 fontName=FONT_C, fontSize=7, leading=9,
                                 textColor=HexColor('#6b7280'))
    subh_style = ParagraphStyle('SubH', parent=styles['Heading3'],
                                fontName=FONT_B, fontSize=10, spaceAfter=3)

    elements = []
    elements.append(Paragraph('利润测算报告', title_style))
    elements.append(Paragraph(
        f'Zone: {params.get("zone","")} | 仓储天数: {params.get("storage_days",0)} | '
        f'汇率: {params.get("exchange_rate","")} | '
        f'平台费率: {params.get("platform_rate",0)}% | '
        f'广告费率: {params.get("ad_rate",0)}% | '
        f'目标利润率: {params.get("profit_rate",0)}%',
        small_style
    ))
    elements.append(Spacer(1, 6*mm))

    for wh_name in US_WAREHOUSES:
        wh_data = warehouses.get(wh_name)
        if not wh_data:
            continue

        profit = wh_data.get('profit', {})
        summary = profit.get('summary', {})
        items = profit.get('items', [])

        is_optimal = wh_name == optimal
        badge = ' ★ 最优' if is_optimal else ''
        elements.append(Paragraph(f'{wh_name}{badge}', h2_style))

        # 总览表
        overview_data = [
            ['费用项目', '金额 (USD)', '说明'],
            ['1. 采购成本', f'{summary.get("purchase_cost",0):,.2f}', 'CNY采购价 / 汇率'],
            ['2. 头程成本', f'{summary.get("head_leg_cost",0):,.2f}', '头程基数 / 汇率 x 体积'],
            ['3. 税务成本', f'{summary.get("tax_cost",0):,.2f}', '每立方米税费 × 体积'],
            ['4. 尾程成本', f'{summary.get("last_mile_cost",0):,.2f}', '各SKU尾程合计'],
            ['5. 仓储成本', f'{summary.get("storage_cost",0):,.2f}', f'{params.get("storage_days",0)}天'],
            ['6. 管理成本', f'{summary.get("management_cost",0):,.2f}', f'{params.get("management_rate",0)}% x 前5项'],
            ['7. 客诉成本', f'{summary.get("complaint_cost",0):,.2f}', f'{params.get("complaint_rate",0)}% x 前5项'],
            ['固定成本合计', f'{summary.get("fixed_cost",0):,.2f}', '7项之和'],
            ['总价格', f'{summary.get("price",0):,.2f}', f'利润率{round(summary.get("profit_rate",0)*100,2)}%'],
            ['总利润', f'{summary.get("profit",0):,.2f}', ''],
        ]

        header_color = HexColor('#D9F2D9') if is_optimal else HexColor('#E2E8F0')
        t = Table(overview_data, colWidths=[160, 90, 270])
        t.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), FONT_C),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('BACKGROUND', (0, 0), (-1, 0), HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), HexColor('#FFFFFF')),
            ('FONTNAME', (0, 0), (-1, 0), FONT_B),
            ('BACKGROUND', (0, -2), (-1, -1), header_color),
            ('FONTNAME', (0, -2), (-1, -1), FONT_B),
            ('GRID', (0, 0), (-1, -1), 0.5, HexColor('#CBD5E1')),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ]))
        elements.append(t)
        elements.append(Spacer(1, 4*mm))

        # 明细表
        elements.append(Paragraph('各SKU费用明细', subh_style))
        detail_data = [['SKU', '数量', '采购', '头程', '税务', '尾程', '仓储', '管理', '客诉', '固定', '价格', '利润', '利润率']]
        for item in items:
            detail_data.append([
                item.get('sku', '-'),
                str(item.get('qty', 1)),
                f'{item.get("purchase_cost",0):,.2f}',
                f'{item.get("head_leg_cost",0):,.2f}',
                f'{item.get("tax_cost",0):,.2f}',
                f'{item.get("last_mile_cost",0):,.2f}',
                f'{item.get("storage_cost",0):,.2f}',
                f'{item.get("management_cost",0):,.2f}',
                f'{item.get("complaint_cost",0):,.2f}',
                f'{item.get("fixed_cost",0):,.2f}',
                f'{item.get("price",0):,.2f}',
                f'{item.get("profit",0):,.2f}',
                f'{round(item.get("profit_rate",0)*100,2)}%',
            ])
        detail_data.append([
            '合计', str(profit.get('total_qty', 0)),
            f'{summary.get("purchase_cost",0):,.2f}',
            f'{summary.get("head_leg_cost",0):,.2f}',
            f'{summary.get("tax_cost",0):,.2f}',
            f'{summary.get("last_mile_cost",0):,.2f}',
            f'{summary.get("storage_cost",0):,.2f}',
            f'{summary.get("management_cost",0):,.2f}',
            f'{summary.get("complaint_cost",0):,.2f}',
            f'{summary.get("fixed_cost",0):,.2f}',
            f'{summary.get("price",0):,.2f}',
            f'{summary.get("profit",0):,.2f}',
            f'{round(summary.get("profit_rate",0)*100,2)}%',
        ])

        col_widths_detail = [55, 25, 36, 34, 34, 36, 34, 34, 34, 36, 36, 36, 36]
        dt = Table(detail_data, colWidths=col_widths_detail)
        dt.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), FONT_C),
            ('FONTNAME', (0, 0), (-1, 0), FONT_B),
            ('FONTSIZE', (0, 0), (-1, -1), 7),
            ('BACKGROUND', (0, 0), (-1, 0), HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), HexColor('#FFFFFF')),
            ('BACKGROUND', (0, -1), (-1, -1), header_color),
            ('FONTNAME', (0, -1), (-1, -1), FONT_B),
            ('GRID', (0, 0), (-1, -1), 0.5, HexColor('#CBD5E1')),
            ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 2),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
        ]))
        elements.append(dt)

        # 仓库间加分页
        current_idx = US_WAREHOUSES.index(wh_name)
        if current_idx < len(US_WAREHOUSES) - 1 and US_WAREHOUSES[current_idx + 1] in warehouses:
            elements.append(PageBreak())

    doc.build(elements)
    buf.seek(0)

    return send_file(
        buf,
        mimetype='application/pdf',
        as_attachment=True,
        download_name=f'利润测算报告_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
    )


# ============================================================
#  SKU 搜索与详情 API
# ============================================================

@app.route('/api/sku/search', methods=['GET'])
def api_sku_search():
    """搜索SKU（前缀匹配），用于自动补全"""
    q = request.args.get('q', '').strip()
    if not q:
        return jsonify({'skus': []})
    try:
        conn = db_connect()
        c = conn.cursor()
        c.execute("SELECT DISTINCT SKU FROM 产品信息 WHERE SKU LIKE ? ORDER BY SKU LIMIT 20", (q + '%',))
        rows = c.fetchall()
        conn.close()
        return jsonify({'skus': [r[0] for r in rows]})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/sku/info', methods=['GET'])
def api_sku_info():
    """获取SKU的尺寸和重量信息"""
    sku = request.args.get('sku', '').strip()
    if not sku:
        return jsonify({'error': '请提供SKU'}), 400
    try:
        conn = db_connect()
        c = conn.cursor()
        c.execute("SELECT 长度, 宽度, 高度, 重量, 采购价 FROM 产品信息 WHERE SKU = ? LIMIT 1", (sku,))
        row = c.fetchone()
        conn.close()
        if not row:
            return jsonify({'error': f'SKU "{sku}" 未找到'}), 404
        return jsonify({
            'sku': sku,
            'length_cm': row[0],
            'width_cm': row[1],
            'height_cm': row[2],
            'weight_kg': row[3],
            'purchase_price': row[4],
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ============================================================
#  产品详情看板 API
# ============================================================

def match_sku_to_board(source_sku: str, sku_mapping: dict = None, region: str = None) -> str:
    """
    将来源SKU（销量仓库SKU / 在途货物SKU）匹配到看板基础SKU
    规则：
    1. 若含 "-"，截取 "-" 前的部分（去掉仓库/平台后缀如 -DY, -FBA, -FG 等）
    2. 在 SKU映射 中按 (地区, 原始SKU) 查找匹配后SKU
    3. 若未找到映射，返回截取后的 SKU 本身
    """
    if not source_sku:
        return ""
    sku = source_sku.strip()
    if '-' in sku:
        sku = sku.split('-')[0].strip()
    if sku_mapping:
        # 优先按 (地区, SKU) 精确匹配
        if region and (region, sku) in sku_mapping:
            return sku_mapping[(region, sku)]
        # 全局映射（地区='-'）
        if ('-', sku) in sku_mapping:
            return sku_mapping[('-', sku)]
    return sku


# 国家代码→映射地区 字典（从数据库"国家映射"表加载）
_country_code_map = {}
_default_country_region = 'DE'  # 兜底默认地区

# SKU映射缓存（5分钟TTL，避免每次请求全量加载SKU映射表）
_sku_mapping_cache = None    # (sku_mapping, reverse_sku_mapping)
_sku_mapping_cache_time = 0


def load_country_mapping(conn):
    """从数据库加载国家代码→映射地区的映射关系"""
    global _country_code_map, _default_country_region
    _country_code_map.clear()
    c = conn.cursor()
    c.execute("SELECT 映射地区, 国家代码 FROM 国家映射 ORDER BY id")
    regions = set()
    for row in c.fetchall():
        region, code = row[0], row[1]
        regions.add(region)
        if code == '*':
            _default_country_region = region  # 兜底地区
        else:
            _country_code_map[code] = region
    # 补充身份映射：映射地区本身也要能映射到自己（UK→UK, US→US, DE→DE）
    for r in regions:
        if r not in _country_code_map:
            _country_code_map[r] = r


def map_country_code(code: str) -> str:
    """将销量表中的国家或地区代码映射为SKU映射表的地区标识
    规则（从数据库国家映射表加载）：
      US/CA → US
      GB/IE → UK
      其余 → DE（默认）
    """
    if not code or code == 'nan':
        return _default_country_region
    return _country_code_map.get(code, _default_country_region)


def map_currency_to_region(currency: str) -> str:
    """将销量表中的币种映射为地区标识
    规则：
      USD/CAD → US（美国地区）
      GBP → UK（英国地区）
      EUR → DE（欧洲地区）
    """
    if not currency or currency == 'nan':
        return 'DE'  # 默认归欧洲
    cur = currency.strip().upper()
    if cur in ('USD', 'CAD'):
        return 'US'
    elif cur == 'GBP':
        return 'UK'
    elif cur == 'EUR':
        return 'DE'
    return 'DE'


def get_sku_mapping_cached(conn):
    """获取SKU映射（带5分钟缓存，避免每次请求全量加载）"""
    global _sku_mapping_cache, _sku_mapping_cache_time
    now = time.time()
    if _sku_mapping_cache and (now - _sku_mapping_cache_time) < 300:
        return _sku_mapping_cache
    c = conn.cursor()
    sku_mapping = {}  # {(地区, 原始SKU): 匹配后SKU}
    reverse_sku_mapping = {}  # {(地区, 匹配后SKU): [原始SKU1, ...]}
    c.execute("SELECT 地区, 原始SKU, 匹配后SKU FROM SKU映射 WHERE id != 67")
    for row in c.fetchall():
        region, orig, matched = row[0] or '', row[1], row[2]
        if orig and matched and region:
            sku_mapping[(region, orig)] = matched
            if (region, matched) not in reverse_sku_mapping:
                reverse_sku_mapping[(region, matched)] = []
            reverse_sku_mapping[(region, matched)].append(orig)
    _sku_mapping_cache = (sku_mapping, reverse_sku_mapping)
    _sku_mapping_cache_time = now
    return _sku_mapping_cache


@app.route('/api/sku-board/list', methods=['GET'])
def api_sku_board_list():
    """获取SKU列表聚合数据"""
    try:
        conn = db_connect()
        c = conn.cursor()

        # 加载国家映射关系
        load_country_mapping(conn)

        # 1. 获取所有唯一SKU（全链接：批次库存表 + 工厂库存表 + 在途货物表 + 产品信息表）
        c.execute("""
            SELECT DISTINCT SKU FROM (
                SELECT SKU FROM 批次库存表 WHERE SKU IS NOT NULL AND SKU != ''
                UNION
                SELECT SKU FROM 工厂库存 WHERE SKU IS NOT NULL AND SKU != ''
                UNION
                SELECT SKU FROM 在途货物 WHERE SKU IS NOT NULL AND SKU != ''
                UNION
                SELECT SKU FROM 产品信息 WHERE SKU IS NOT NULL AND SKU != ''
            ) ORDER BY SKU
        """)
        sku_rows = c.fetchall()
        raw_skus = [r[0] for r in sku_rows]

        # 预加载 SKU映射（地区感知）
        sku_mapping = {}  # {(地区, 原始SKU): 匹配后SKU}
        reverse_sku_mapping = {}  # {(地区, 匹配后SKU): [原始SKU1, ...]}
        c.execute("SELECT 地区, 原始SKU, 匹配后SKU FROM SKU映射 WHERE id != 67")
        for row in c.fetchall():
            region, orig, matched = row[0] or '', row[1], row[2]
            if orig and matched and region:
                sku_mapping[(region, orig)] = matched
                if (region, matched) not in reverse_sku_mapping:
                    reverse_sku_mapping[(region, matched)] = []
                reverse_sku_mapping[(region, matched)].append(orig)

        # SKU归一化（仅用于SKU列表构建，不做地区特定映射）
        # 规则：去"-"获取base，仅应用通用映射（地区='-'），不应用US/UK/DE特定映射
        def normalize_sku(raw):
            base = raw.split('-')[0].strip() if '-' in raw else raw.strip()
            # 仅检查通用映射
            if ('-', base) in sku_mapping:
                return sku_mapping[('-', base)]
            if ('-', raw) in sku_mapping:
                return sku_mapping[('-', raw)]
            return base

        # 建立 原始SKU→归一化SKU 映射，以及 归一化SKU→原始SKU集合 反向映射
        raw_to_norm = {}
        norm_to_raws = {}
        for raw in raw_skus:
            norm = normalize_sku(raw)
            raw_to_norm[raw] = norm
            if norm not in norm_to_raws:
                norm_to_raws[norm] = set()
            norm_to_raws[norm].add(raw)

        all_skus = sorted(norm_to_raws.keys())
        board_sku_set = set(all_skus)

        # 构建 region-aware 反向映射 norm_reverse，用于销量/在途匹配
        # 规则：地区特定映射优先（从sku_mapping表），否则fallback到归一化结果
        norm_reverse = {}  # {(地区, 来源SKU): 看板归一化SKU}
        # 1. 归一化SKU的身份映射 (地区, norm) → norm
        for norm in all_skus:
            for region in ('US', 'UK', 'DE', '-'):
                norm_reverse[(region, norm)] = norm
        # 2. 原始SKU→归一化SKU的映射 (地区, raw) → norm
        for norm, raws in norm_to_raws.items():
            for raw in raws:
                for region in ('US', 'UK', 'DE', '-'):
                    norm_reverse[(region, raw)] = norm
        # 3. SKU映射表中的地区特定映射（仅在对应地区生效）
        for (region, orig), matched in sku_mapping.items():
            if region in ('US', 'UK', 'DE'):
                norm_reverse[(region, orig)] = matched
        # 4. reverse_sku_mapping 中的别名（按地区）
        for (region, matched), origs in reverse_sku_mapping.items():
            if region in ('US', 'UK', 'DE'):
                for orig in origs:
                    norm_reverse[(region, orig)] = matched

        # 2. 获取日期和近8天日期（t-1往回推8天，即t-1到t-8）
        fixed_sales_start = datetime(2026, 6, 3)
        eight_dates = [
            (fixed_sales_start + timedelta(days=i)).strftime('%Y-%m-%d')
            for i in range(8)
        ]

        # 预构建 来源SKU→看板SKU 映射（用于在途货物/销量匹配，地区感知）
        # 优先用归一化映射 norm_reverse，fallback到match_sku_to_board
        source_to_board = norm_reverse  # 直接复用

        # 加载负责人表：{(SKU, 店铺, 国家): 负责人}
        c.execute("SELECT SKU, 店铺, 国家, 负责人 FROM 负责人 WHERE 负责人 IS NOT NULL")
        owner_data = {}
        all_owners_set = set()
        for row in c.fetchall():
            o_sku, o_plat, o_country, o_name = row[0] or '', row[1] or '', row[2] or '', row[3]
            if o_sku and o_name:
                all_owners_set.add(o_name)
                owner_data[(o_sku, o_plat, o_country)] = o_name
        all_owners = sorted(all_owners_set)

        # 批量查询在途货物，按看板SKU聚合（Layer1:总量, Layer2:按市场代码）
        c.execute("SELECT SKU, 市场代码, SUM(数量) FROM 在途货物 GROUP BY SKU, 市场代码")
        transit_rows = c.fetchall()
        transit_by_sku = {}
        for row in transit_rows:
            zt_sku, market_code, qty = row[0] or '', row[1] or '', row[2] or 0
            matched = source_to_board.get((market_code, zt_sku))
            if matched is None:
                matched = source_to_board.get(('-', zt_sku))
            if matched is None:
                matched = normalize_sku(zt_sku)
            if matched is None or matched not in board_sku_set:
                continue
            if matched not in transit_by_sku:
                transit_by_sku[matched] = {'total': 0, 'by_country': {}}
            transit_by_sku[matched]['total'] += qty
            transit_by_sku[matched]['by_country'][market_code] = transit_by_sku[matched]['by_country'].get(market_code, 0) + qty

        # 批量查询90天销量，按看板SKU聚合
        ninety_ago_slash = (datetime.now() - timedelta(days=90)).strftime('%Y/%m/%d')
        yesterday_end = (datetime.now() - timedelta(days=1)).strftime('%Y/%m/%d') + ' 23:59:59'
        c.execute("""SELECT 付款时间, 仓库SKU, 数量, 平台, 账号, 币种
                     FROM 销量 WHERE 付款时间 >= ? AND 付款时间 <= ?""",
                  (ninety_ago_slash, yesterday_end))
        sales_rows = c.fetchall()

        sales_by_sku = {}
        for row in sales_rows:
            pay_time = row[0] or ''
            wh_sku = row[1] or ''
            qty = row[2] or 0
            platform = row[3] or ''
            account = row[4] or ''
            raw_currency = row[5] or ''
            country_code = map_currency_to_region(raw_currency)
            matched = source_to_board.get((country_code, wh_sku))
            if matched is None:
                matched = source_to_board.get(('-', wh_sku))
            if matched is None:
                matched = match_sku_to_board(wh_sku, sku_mapping, country_code)
            if not matched or matched not in board_sku_set:
                if matched and matched.strip():
                    board_sku_set.add(matched)
                    all_skus.append(matched)
                    continue
                continue
            date_key = pay_time[:10].replace('/', '-') if pay_time else ''
            if matched not in sales_by_sku:
                sales_by_sku[matched] = {}
            if date_key not in sales_by_sku[matched]:
                sales_by_sku[matched][date_key] = {'total': 0, 'by_platform': {}, 'by_account': {}, 'by_country': {}, 'by_plat_country': {}, 'by_acct_country': {}}
            sales_by_sku[matched][date_key]['total'] += qty
            sales_by_sku[matched][date_key]['by_platform'][platform] = sales_by_sku[matched][date_key]['by_platform'].get(platform, 0) + qty
            sales_by_sku[matched][date_key]['by_account'][account] = sales_by_sku[matched][date_key]['by_account'].get(account, 0) + qty
            sales_by_sku[matched][date_key]['by_country'][country_code] = sales_by_sku[matched][date_key]['by_country'].get(country_code, 0) + qty
            # 平台+地区 组合key（同一平台在不同国家有不同负责人）
            pc_key = f"{platform}|{country_code}"
            sales_by_sku[matched][date_key]['by_plat_country'][pc_key] = sales_by_sku[matched][date_key]['by_plat_country'].get(pc_key, 0) + qty
            # 账号+地区 组合key
            ac_key = f"{account}|{country_code}"
            sales_by_sku[matched][date_key]['by_acct_country'][ac_key] = sales_by_sku[matched][date_key]['by_acct_country'].get(ac_key, 0) + qty

        # 预构建 事业部 查找字典：{规范化SKU: 所属事业部}
        division_map = {}
        # 直接按产品信息表SKU精确匹配
        c.execute("SELECT SKU, 所属事业部 FROM 产品信息 WHERE 所属事业部 IS NOT NULL AND 所属事业部 != ''")
        for row in c.fetchall():
            prod_sku = row[0] or ''
            division = row[1] or ''
            if not prod_sku or not division:
                continue
            # 直接匹配（不做映射，因为产品信息表SKU已经是归一化后的格式）
            division_map[prod_sku] = division
            # 同时去掉 "-" 后缀作为备用key
            if '-' in prod_sku:
                base = prod_sku.split('-')[0].strip()
                if base not in division_map:
                    division_map[base] = division

            # 归一化后的SKU也加入division_map
            norm = normalize_sku(prod_sku)
            if norm not in division_map:
                division_map[norm] = division

        # 批量查询工厂库存 → 按归一化SKU聚合
        c.execute("SELECT SKU, COALESCE(国内在库,0) FROM 工厂库存 WHERE 国内在库 > 0")
        factory_stock_raw = {}
        for row in c.fetchall():
            fs_sku, fs_qty = row[0], row[1] or 0
            factory_stock_raw[fs_sku] = factory_stock_raw.get(fs_sku, 0) + fs_qty

        # 批量查询批次库存（仅用于库龄匹配，库存量改用海外仓库存表）
        c.execute("SELECT SKU, 国家, 仓库, 库存, 库龄 FROM 批次库存表 WHERE 库存 > 0 ORDER BY SKU, 国家, 仓库")
        batch_rows_all = c.fetchall()

        # 批量查询海外仓库存（替代批次库存表作为海外库存量数据源）
        c.execute("SELECT SKU, 仓库名称, SUM(可用数量), 市场 FROM 海外仓库存 WHERE 可用数量 > 0 GROUP BY SKU, 仓库名称")
        overseas_inv_rows = c.fetchall()
        overseas_inv_by_sku = {}       # {归一化SKU: {'total': int, 'by_country': {US: int, ...}}}
        overseas_wh_by_sku_market = {} # {归一化SKU: {US: [{'warehouse': 'xxx', 'qty': 10}, ...]}}
        for row in overseas_inv_rows:
            oiv_sku, oiv_wh, oiv_qty, oiv_market = row[0] or '', row[1] or '', row[2] or 0, row[3] or ''
            if not oiv_market:
                oiv_market = 'US'
            matched = source_to_board.get((oiv_market, oiv_sku))
            if matched is None:
                matched = source_to_board.get(('-', oiv_sku))
            if matched is None:
                matched = normalize_sku(oiv_sku)
            if not matched or matched not in board_sku_set:
                continue
            if matched not in overseas_inv_by_sku:
                overseas_inv_by_sku[matched] = {'total': 0, 'by_country': {}}
                overseas_wh_by_sku_market[matched] = {}
            overseas_inv_by_sku[matched]['total'] += oiv_qty
            overseas_inv_by_sku[matched]['by_country'][oiv_market] = overseas_inv_by_sku[matched]['by_country'].get(oiv_market, 0) + oiv_qty
            # 仓库级明细
            if oiv_market not in overseas_wh_by_sku_market[matched]:
                overseas_wh_by_sku_market[matched][oiv_market] = []
            overseas_wh_by_sku_market[matched][oiv_market].append({'warehouse': oiv_wh, 'qty': oiv_qty})

        result_skus = []

        for sku in all_skus:
            # 该归一化SKU对应的所有原始SKU
            raws = norm_to_raws.get(sku, {sku})
            item = {'sku': sku}

            # 事业部查找：归一化SKU直接匹配 → 原始SKU去"-"匹配
            item['division'] = division_map.get(sku, '')
            if not item['division']:
                for raw in raws:
                    item['division'] = division_map.get(raw, '')
                    if item['division']: break
                    if '-' in raw:
                        item['division'] = division_map.get(raw.split('-')[0].strip(), '')
                        if item['division']: break

            # 销量数据（sales_by_sku 已按归一化SKU聚合）
            sku_sales = sales_by_sku.get(sku, {})

            # 国内仓库存：汇总该归一化SKU下所有原始SKU的工厂库存
            domestic_stock = 0
            for raw in raws:
                domestic_stock += factory_stock_raw.get(raw, 0)
            item['domestic_stock'] = domestic_stock

            # 海外仓库存：从海外仓库存表取量，替代批次库存表
            raws_set = set(raws)
            overseas_inv = overseas_inv_by_sku.get(sku, {'total': 0, 'by_country': {}})
            overseas_total = overseas_inv['total']

            # 库龄：仅从批次库存表匹配，按SKU和市场维度（不做仓库级展开）
            age_weighted_sum = 0
            age_total_qty = 0
            max_age_val = 0
            age_by_country = {}  # {country: {'weighted': int, 'total_qty': int, 'max': int}}
            for row in batch_rows_all:
                b_sku = row[0] or ''
                if b_sku not in raws_set:
                    continue
                country = row[1] or '未知'
                country = map_country_code(country)
                qty = row[3] or 0
                age = row[4] or 0
                age_weighted_sum += age * qty
                age_total_qty += qty
                if age > max_age_val:
                    max_age_val = age
                if country not in age_by_country:
                    age_by_country[country] = {'weighted': 0, 'total_qty': 0, 'max': 0}
                age_by_country[country]['weighted'] += age * qty
                age_by_country[country]['total_qty'] += qty
                if age > age_by_country[country]['max']:
                    age_by_country[country]['max'] = age

            item['overseas_stock'] = overseas_total
            item['max_age'] = max_age_val
            item['avg_age'] = round(age_weighted_sum / age_total_qty, 1) if age_total_qty > 0 else 0

            # 在途货物（从预聚合数据取）
            transit = transit_by_sku.get(sku, {'total': 0, 'by_country': {}})
            item['in_transit'] = transit['total']
            item['in_transit_by_country'] = transit['by_country']

            # 国家列表：海外仓库存的市场 + 在途的市场 + 销量中出现的国家
            countries = set(overseas_inv['by_country'].keys())
            for c in transit['by_country']:
                if c: countries.add(c)
            for d in sku_sales:
                for c in sku_sales[d].get('by_country', {}):
                    if c:
                        countries.add(c)
            item['countries'] = sorted(list(countries))

            # 销量数据 - 从预聚合数据取
            all_dates = sorted(sku_sales.keys())

            # 近8天销量
            sales_8d = [round(sku_sales.get(d, {}).get('total', 0), 2) for d in eight_dates]
            item['sales_8d'] = sales_8d
            item['sales_8d_dates'] = eight_dates

            # 计算各时段平均日销
            def calc_avg(days):
                cutoff = (datetime.now() - timedelta(days=days)).strftime('%Y-%m-%d')
                relevant = [sku_sales[d]['total'] for d in all_dates if d >= cutoff]
                total = sum(relevant)
                return round(total / days, 2) if days > 0 else 0

            item['avg_7d'] = calc_avg(7)
            item['avg_14d'] = calc_avg(14)
            item['avg_30d'] = calc_avg(30)

            # 加权日均销量
            cutoff_30 = (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d')
            cutoff_60 = (datetime.now() - timedelta(days=60)).strftime('%Y-%m-%d')
            relevant_31_60 = [sku_sales[d]['total'] for d in all_dates if cutoff_60 <= d < cutoff_30]
            avg_31_60_val = round(sum(relevant_31_60) / 30, 2) if relevant_31_60 else 0

            cutoff_90 = (datetime.now() - timedelta(days=90)).strftime('%Y-%m-%d')
            relevant_61_90 = [sku_sales[d]['total'] for d in all_dates if cutoff_90 <= d < cutoff_60]
            avg_61_90_val = round(sum(relevant_61_90) / 30, 2) if relevant_61_90 else 0

            weighted_daily = round(0.5 * item['avg_30d'] + 0.3 * avg_31_60_val + 0.2 * avg_61_90_val, 2)
            item['weighted_daily'] = weighted_daily

            # 售罄天数 = 海外仓库存 / (7天+14天+30天日均的平均值)
            avg_daily = (item['avg_7d'] + item['avg_14d'] + item['avg_30d']) / 3
            item['sellout_days'] = round(overseas_total / avg_daily, 1) if avg_daily > 0 else 9999

            # 子行数据：按国家聚合
            country_data = {}
            for country in item['countries']:
                cd = {'country': country}

                # 该国家的海外仓库存（来自海外仓库存表）
                cd['overseas_stock'] = overseas_inv['by_country'].get(country, 0)
                # 该国家的库龄（来自批次库存表，按市场匹配）
                country_age = age_by_country.get(country, {'weighted': 0, 'total_qty': 0, 'max': 0})
                cd['max_age'] = country_age['max']
                cd['avg_age'] = round(country_age['weighted'] / country_age['total_qty'], 1) if country_age['total_qty'] > 0 else 0

                # 该国家的近8天销量（按country代码匹配）
                cd['sales_8d'] = [
                    round(sku_sales.get(d, {}).get('by_country', {}).get(country, 0), 2)
                    for d in eight_dates
                ]

                # 该国各时段
                def calc_country_avg(days):
                    c_cutoff = (datetime.now() - timedelta(days=days)).strftime('%Y-%m-%d')
                    rel = []
                    for d in all_dates:
                        if d >= c_cutoff:
                            rel.append(sku_sales[d].get('by_country', {}).get(country, 0))
                    total = sum(rel)
                    return round(total / days, 2) if days > 0 else 0

                cd['avg_7d'] = calc_country_avg(7)
                cd['avg_14d'] = calc_country_avg(14)
                cd['avg_30d'] = calc_country_avg(30)

                # 该国家的在途货物数量
                cd['in_transit'] = transit['by_country'].get(country, 0)

                # 仓库明细：该SKU在该市场下各仓库的库存（来自海外仓库存表）
                sku_wh = overseas_wh_by_sku_market.get(sku, {})
                cd['warehouses'] = sku_wh.get(country, [])

                country_data[country] = cd

            item['country_data'] = country_data

            # 按账号聚合销量（用于第三层展开：按账号+国家）
            account_sales = {}         # key=账号, value={total, daily: {日期: 数量}}
            acct_country_sales = {}    # key=账号|国家, value={total, daily: {日期: 数量}}
            for d in all_dates:
                for acct, qty in sku_sales[d].get('by_account', {}).items():
                    if acct not in account_sales:
                        account_sales[acct] = {'total': 0, 'daily': {}}
                    account_sales[acct]['total'] += qty
                    account_sales[acct]['daily'][d] = account_sales[acct]['daily'].get(d, 0) + qty
                for ac_key, qty in sku_sales[d].get('by_acct_country', {}).items():
                    if ac_key not in acct_country_sales:
                        acct_country_sales[ac_key] = {'total': 0, 'daily': {}}
                    acct_country_sales[ac_key]['total'] += qty
                    acct_country_sales[ac_key]['daily'][d] = acct_country_sales[ac_key]['daily'].get(d, 0) + qty

            # 构建 account_sales 输出，每个账号下增加 by_country 子对象
            cutoff_7 = (datetime.now() - timedelta(days=7)).strftime('%Y-%m-%d')
            cutoff_14 = (datetime.now() - timedelta(days=14)).strftime('%Y-%m-%d')
            cutoff_30 = (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d')

            def build_as_output(daily_dict):
                """根据 daily 字典构建 {total_90d, sales_8d, avg_7/14/30d}"""
                return {
                    'total_90d': round(daily_dict.get('total', 0), 2),
                    'sales_8d': [round(daily_dict.get(d, 0), 2) for d in eight_dates],
                    'avg_7d': round(sum(daily_dict.get(d, 0) for d in all_dates if d >= cutoff_7) / 7, 2),
                    'avg_14d': round(sum(daily_dict.get(d, 0) for d in all_dates if d >= cutoff_14) / 14, 2),
                    'avg_30d': round(sum(daily_dict.get(d, 0) for d in all_dates if d >= cutoff_30) / 30, 2),
                }

            item['account_sales'] = {}
            for acct, aps in account_sales.items():
                entry = build_as_output({'total': aps['total'], **aps['daily']})
                # 按国家拆分的子数据
                entry['by_country'] = {}
                for country in item['countries']:
                    ac_key = f"{acct}|{country}"
                    if ac_key in acct_country_sales:
                        acs = acct_country_sales[ac_key]
                        entry['by_country'][country] = build_as_output({'total': acs['total'], **acs['daily']})
                    else:
                        entry['by_country'][country] = build_as_output({})
                item['account_sales'][acct] = entry

            # 负责人：按(SKU, 账号, 市场)三元组匹配，无匹配显示"未匹配"
            item['owners'] = {}
            for acct in account_sales:
                item['owners'][acct] = {}
                for country in item['countries']:
                    owner = owner_data.get((sku, acct, country), '')
                    item['owners'][acct][country] = owner if owner else '未匹配'

            result_skus.append(item)

        # 过滤：自动隐藏 30天无销量 且 无库存(国内+海外=0) 且 无在途 的SKU
        result_skus = [s for s in result_skus if not (
            s['avg_30d'] == 0 and
            (s['domestic_stock'] + s['overseas_stock']) == 0 and
            s['in_transit'] == 0
        )]

        conn.close()
        return jsonify({'skus': result_skus, 'all_owners': all_owners, 'date_range': {'from': ninety_ago_slash, 'to': yesterday_end}})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/api/sku-board/export', methods=['POST'])
def api_sku_board_export():
    """按筛选条件导出SKU看板数据为Excel"""
    try:
        data = request.get_json(silent=True) or {}
        skus = data.get('skus', [])
        if not skus:
            return jsonify({'error': '无数据可导出'}), 400

        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from io import BytesIO

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '产品详情看板'

        # 获取日期列
        dates = skus[0].get('sales_8d_dates', []) if skus else []
        date_headers = [d[5:] for d in dates]  # MM-DD

        # 表头
        headers = ['SKU', '事业部', '国内仓库存', '海外仓库存', '在途货物',
                   '最高库龄(天)', '平均库龄(天)']
        headers += date_headers
        headers += ['7天日均', '14天日均', '30天日均', '售罄天数']

        # 样式
        header_font = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
        header_fill = PatternFill(start_color='3B82F6', end_color='3B82F6', fill_type='solid')
        header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin', color='D0D0D0'),
            right=Side(style='thin', color='D0D0D0'),
            top=Side(style='thin', color='D0D0D0'),
            bottom=Side(style='thin', color='D0D0D0'),
        )
        data_font = Font(name='微软雅黑', size=10)
        data_align = Alignment(horizontal='center', vertical='center')

        # 写入表头
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        # 写入数据
        for row_idx, sku in enumerate(skus, 2):
            values = [
                sku.get('sku', ''),
                sku.get('division', ''),
                sku.get('domestic_stock', 0),
                sku.get('overseas_stock', 0),
                sku.get('in_transit', 0),
                sku.get('max_age', 0),
                sku.get('avg_age', 0),
            ]
            values += sku.get('sales_8d', [0]*8)
            values += [
                sku.get('avg_7d', 0),
                sku.get('avg_14d', 0),
                sku.get('avg_30d', 0),
                sku.get('sellout_days', 0),
            ]
            for col_idx, val in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.font = data_font
                cell.alignment = data_align
                cell.border = thin_border

        # 自动列宽
        for col_idx in range(1, len(headers) + 1):
            max_len = len(str(headers[col_idx - 1]))
            for row_idx in range(2, len(skus) + 2):
                cell_val = str(ws.cell(row=row_idx, column=col_idx).value or '')
                # 中文字符约占2个字符宽度
                cn_count = sum(1 for c in cell_val if '\u4e00' <= c <= '\u9fff')
                cell_len = len(cell_val) + cn_count
                if cell_len > max_len:
                    max_len = cell_len
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_len + 4, 25)

        # 冻结首行
        ws.freeze_panes = 'A2'

        # 写入BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'产品详情看板_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
        )
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/api/sku-board/detail/<sku>', methods=['GET'])
def api_sku_board_detail(sku):
    """获取单个SKU的详细数据（用于BI看板）
    支持 ?owner=负责人名 参数，按负责人（账号→平台映射）筛选销量数据
    """
    try:
        conn = db_connect()
        c = conn.cursor()

        # 加载国家映射关系
        load_country_mapping(conn)

        # 加载 SKU映射（带缓存，避免每次请求全量加载）
        sku_mapping, reverse_sku_mapping = get_sku_mapping_cached(conn)

        # ----- 负责人筛选参数 -----
        owner_filter = request.args.get('owner', '').strip()
        # 该负责人在本SKU下负责的 (账号, 国家) 组合列表
        # 关键：同一账号在不同国家有不同负责人，必须按 (账号+币种→国家) 组合筛选
        owner_acct_country_set = set()  # {(账号, 国家), ...}
        if owner_filter:
            c.execute("SELECT DISTINCT 店铺, 国家 FROM 负责人 WHERE SKU = ? AND 负责人 = ? AND 店铺 IS NOT NULL AND 店铺 != ''",
                      (sku, owner_filter))
            for r in c.fetchall():
                owner_acct_country_set.add((r[0] or '', r[1] or ''))

        def _match_owner(account, currency):
            """检查 (账号, 币种) 是否匹配当前负责人筛选"""
            if not owner_acct_country_set:
                return True  # 无筛选 = 全部通过
            country = map_currency_to_region(currency)
            return (account, country) in owner_acct_country_set

        # 产品信息
        c.execute("SELECT 采购价 FROM 产品信息 WHERE SKU = ? LIMIT 1", (sku,))
        row = c.fetchone()
        purchase_price = row[0] if row else 0

        # 国内仓库存（工厂库存）
        c.execute("""SELECT 国内在库, 库龄天数, 工厂简写, 合同交期
                     FROM 工厂库存 WHERE SKU = ? AND 国内在库 > 0
                     ORDER BY 库龄天数""", (sku,))
        domestic_rows = c.fetchall()
        domestic_batches = [{
            'qty': r[0], 'age': r[1] or 0, 'original': r[0],
            'warehouse': r[2] or '工厂',
            'delivery_date': r[3] or ''
        } for r in domestic_rows]

        # 海外仓库存：数量取海外仓库存表，库龄取批次库存表
        c.execute("""SELECT 仓库名称, SUM(可用数量), 市场
                     FROM 海外仓库存 WHERE SKU = ? AND 可用数量 > 0
                     GROUP BY 仓库名称""", (sku,))
        overseas_inv_rows = c.fetchall()
        overseas_qty_map = {}  # {仓库名称: 数量}
        overseas_market_map = {}  # {仓库名称: 市场}
        for r in overseas_inv_rows:
            overseas_qty_map[r[0] or ''] = r[1] or 0
            overseas_market_map[r[0] or ''] = r[2] or 'US'

        c.execute("""SELECT 仓库, 库存, 库龄, 国家
                     FROM 批次库存表 WHERE SKU = ? AND 库存 > 0
                     ORDER BY 库龄 DESC""", (sku,))
        batch_rows = c.fetchall()
        overseas_batches = []
        for r in batch_rows:
            wh = r[0] or '海外仓'
            qty = overseas_qty_map.get(wh, r[1] or 0)  # 优先海外仓库存表的数量
            age = r[2] or 0
            country = r[3] or ''
            market = overseas_market_map.get(wh, map_country_code(country))
            overseas_batches.append({
                'warehouse': f'{wh}({market})' if market else wh,
                'qty': qty, 'age': age, 'original': qty
            })

        # 在途货物（直接匹配 + 反向SKU映射匹配，地区感知）
        match_skus = [sku]
        for region in ('US', 'UK', 'DE', '-'):
            if (region, sku) in reverse_sku_mapping:
                match_skus.extend(reverse_sku_mapping[(region, sku)])
        placeholders = ','.join(['?' for _ in match_skus])
        c.execute(f"SELECT COALESCE(SUM(数量),0) FROM 在途货物 WHERE SKU IN ({placeholders})", match_skus)
        in_transit = c.fetchone()[0]

        # --- 90天销量数据（SQL下推过滤，避免全表扫描后在Python中逐行过滤）---
        # 预计算该SKU所有可能的仓库SKU匹配值
        sales_match_skus = [sku]           # 精确匹配
        sales_like_patterns = [sku + '-%']  # 带后缀（如 A001-FBA）
        # 从SKU映射中收集所有能匹配到该SKU的原始SKU
        for (region, orig), matched in sku_mapping.items():
            if matched == sku and orig != sku and region in ('US', 'UK', 'DE', '-'):
                sales_match_skus.append(orig)
                sales_like_patterns.append(orig + '-%')
        for region in ('US', 'UK', 'DE', '-'):
            if (region, sku) in reverse_sku_mapping:
                for orig in reverse_sku_mapping[(region, sku)]:
                    if orig not in sales_match_skus:
                        sales_match_skus.append(orig)
                        sales_like_patterns.append(orig + '-%')

        ninety_ago_slash = (datetime.now() - timedelta(days=90)).strftime('%Y/%m/%d')
        yesterday_end = (datetime.now() - timedelta(days=1)).strftime('%Y/%m/%d') + ' 23:59:59'

        # 构建 SQL：时间范围 + (IN精确匹配 OR LIKE前缀匹配)
        # 始终带上 账号 + 币种 字段，用于 Python 层的 (账号, 币种→国家) 负责人精确筛选
        in_ph = ','.join(['?' for _ in sales_match_skus])
        like_clauses = ' OR '.join(['仓库SKU LIKE ?' for _ in sales_like_patterns])
        sales_sql = f"""SELECT 付款时间, 仓库SKU, 数量, 平台, 账号, 币种
                        FROM 销量
                        WHERE 付款时间 >= ? AND 付款时间 <= ?
                        AND (仓库SKU IN ({in_ph}) OR {like_clauses})
                        ORDER BY 付款时间"""
        params = [ninety_ago_slash, yesterday_end] + sales_match_skus + sales_like_patterns
        c.execute(sales_sql, params)
        all_sales_rows = c.fetchall()

        # Python 层按 (账号, 币种→国家) 精确筛选负责人
        if owner_acct_country_set:
            sales_rows = [r for r in all_sales_rows if _match_owner(r[4] or '', r[5] or '')]
        else:
            sales_rows = all_sales_rows

        # 按天汇总（SQL已过滤，循环直接累加，无需逐行match_sku_to_board）
        daily_sales = {}
        all_dates = []
        d = datetime.now() - timedelta(days=90)
        end_d = datetime.now() - timedelta(days=1)
        while d <= end_d:
            all_dates.append(d.strftime('%Y-%m-%d'))
            d += timedelta(days=1)

        for row in sales_rows:
            pay_time = row[0] or ''
            qty = row[2] or 0
            if not pay_time:
                continue
            date_key = pay_time[:10].replace('/', '-')
            daily_sales[date_key] = daily_sales.get(date_key, 0) + qty

        # 填充所有日期
        daily_sales_list = [round(daily_sales.get(d, 0), 2) for d in all_dates]
        date_labels = all_dates

        # 仓储费计算（从批次库存表汇总当日仓储费）
        c.execute("SELECT COALESCE(SUM(当日仓储费),0), COALESCE(SUM(CASE WHEN 库龄>=180 THEN 当日仓储费 ELSE 0 END),0), COALESCE(SUM(CASE WHEN 库龄<180 THEN 当日仓储费 ELSE 0 END),0) FROM 批次库存表 WHERE SKU = ?", (sku,))
        row = c.fetchone()
        storage_total = row[0] or 0
        storage_gt180 = row[1] or 0
        storage_lt180 = row[2] or 0

        c.execute("SELECT COALESCE(SUM(CASE WHEN 库龄>=180 THEN 库存 ELSE 0 END),0), COALESCE(SUM(CASE WHEN 库龄<180 THEN 库存 ELSE 0 END),0) FROM 批次库存表 WHERE SKU = ?", (sku,))
        row = c.fetchone()
        gt180_qty = row[0] or 0
        lt180_qty = row[1] or 0

        total_overseas = sum(b['qty'] for b in overseas_batches)
        total_domestic = sum(b['qty'] for b in domestic_batches)
        total_inv = total_overseas + total_domestic

        # 销量统计
        def calc_period(days):
            p = daily_sales_list[-days:]
            t = round(sum(p), 2)
            return {'total': t, 'avg': round(t/days, 2)}

        s7 = calc_period(7)
        s14 = calc_period(14)
        s30 = calc_period(30)
        s90 = calc_period(90)

        p31_60 = daily_sales_list[30:60] if len(daily_sales_list) >= 60 else []
        t31_60 = sum(p31_60)
        avg31_60 = round(t31_60/30, 2) if p31_60 else 0

        p61_90 = daily_sales_list[:30]
        t61_90 = sum(p61_90)
        avg61_90 = round(t61_90/30, 2) if p61_90 else 0

        weighted_daily = round(0.5*s30['avg'] + 0.3*avg31_60 + 0.2*avg61_90, 2)
        # 用户要求：日均销量取7/14/30天平均值
        avg_daily_3period = round((s7['avg'] + s14['avg'] + s30['avg']) / 3, 2)

        sellout_days = round(total_overseas / avg_daily_3period, 1) if avg_daily_3period > 0 else 9999
        stockout_date = datetime.now() + timedelta(days=min(int(sellout_days), 9999))

        # 月度对比
        yesterday = (datetime.now() - timedelta(days=1)).strftime('%Y-%m-%d')
        may_cutoff = f'{datetime.now().year}-05-01'
        apr_cutoff = f'{datetime.now().year}-04-01'
        may_sales = [daily_sales.get(d, 0) for d in all_dates if d >= may_cutoff and d <= yesterday]
        apr_sales = [daily_sales.get(d, 0) for d in all_dates if d >= apr_cutoff and d < may_cutoff]
        may_total = round(sum(may_sales), 2)
        apr_total = round(sum(apr_sales), 2)
        may_daily = round(may_total / len(may_sales), 2) if may_sales else 0
        apr_daily = round(apr_total / 30, 2)
        may_projected = round(may_daily * 31, 1)
        mom_change = round((may_projected - apr_total) / apr_total * 100, 1) if apr_total > 0 else 0

        # 近30天各账号销量（带SKU映射 + 可选负责人精确筛选(账号,币种)）
        thirty_ago_slash = (datetime.now() - timedelta(days=30)).strftime('%Y/%m/%d')
        account_sales_30d = defaultdict(int)
        c.execute("""SELECT 账号, 仓库SKU, 数量, 币种
                     FROM 销量 WHERE 付款时间 >= ? AND 付款时间 <= ?""",
                  (thirty_ago_slash, yesterday_end))
        for row in c.fetchall():
            acct, wh_sku, qty, raw_currency = row[0] or '未知', row[1] or '', row[2] or 0, row[3] or ''
            # 负责人精确筛选：(账号, 币种→国家) 组合匹配
            if not _match_owner(acct, raw_currency):
                continue
            country_code = map_currency_to_region(raw_currency)
            matched = match_sku_to_board(wh_sku, sku_mapping, country_code)
            if matched == sku:
                account_sales_30d[acct] += qty

        conn.close()

        return jsonify({
            'sku': sku,
            'purchase_price': purchase_price,
            'domestic_batches': domestic_batches,
            'overseas_batches': overseas_batches,
            'in_transit': in_transit,
            'total_overseas': total_overseas,
            'total_domestic': total_domestic,
            'total_inv': total_inv,
            'overseas_value': total_overseas * purchase_price,
            'domestic_value': total_domestic * purchase_price,
            'total_value': total_inv * purchase_price,
            'daily_sales': daily_sales_list,
            'date_labels': date_labels,
            's7': s7, 's14': s14, 's30': s30, 's90': s90,
            'avg31_60': avg31_60, 'avg61_90': avg61_90,
            'weighted_daily': weighted_daily,
            'avg_daily_3period': avg_daily_3period,
            'sellout_days': sellout_days,
            'stockout_date': stockout_date.strftime('%Y-%m-%d'),
            'account_sales_30d': account_sales_30d,
            'storage_total': round(storage_total, 2),
            'storage_gt180': round(storage_gt180, 2),
            'storage_lt180': round(storage_lt180, 2),
            'gt180_qty': gt180_qty,
            'lt180_qty': lt180_qty,
            'may_total': may_total, 'may_daily': may_daily, 'may_projected': may_projected,
            'apr_total': apr_total, 'apr_daily': apr_daily,
            'mom_change': mom_change
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


# ============================================================
#  问题反馈 API — 共享问题追踪
# ============================================================

@app.route('/api/issues', methods=['GET'])
def api_issues_list():
    """获取所有问题反馈（未完成/已完成分列，各自独立序号）"""
    try:
        conn = db_connect()
        c = conn.cursor()
        c.execute('SELECT id, 模块, 内容, 提交人, 紧急程度, 开发者备注, 状态, 创建时间 FROM 问题反馈 ORDER BY id')
        rows = c.fetchall()
        conn.close()

        undone = []
        done = []
        for row in rows:
            item = {
                'id': row[0],
                'module': row[1] or '',
                'content': row[2] or '',
                'submitter': row[3] or '',
                'urgency': row[4] or '一般',
                'dev_note': row[5] or '',
                'status': row[6] or '未完成',
                'created_at': row[7] or ''
            }
            if item['status'] in ('已完成', 'done', 'Done', 'DONE'):
                done.append(item)
            else:
                undone.append(item)

        # 各自独立编号
        for idx, item in enumerate(undone, 1):
            item['seq'] = idx
        for idx, item in enumerate(done, 1):
            item['seq'] = idx

        return jsonify({
            'undone': undone,
            'done': done,
            'undone_count': len(undone),
            'done_count': len(done)
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/issues', methods=['POST'])
def api_issues_create():
    """新增问题反馈"""
    try:
        data = request.get_json(force=True)
    except Exception:
        return jsonify({'error': '无效的JSON数据'}), 400

    content = (data.get('content') or '').strip()
    if not content:
        return jsonify({'error': '内容不能为空'}), 400

    try:
        conn = db_connect()
        c = conn.cursor()
        c.execute('''INSERT INTO 问题反馈 (模块, 内容, 提交人, 紧急程度, 开发者备注, 状态, 创建时间)
                     VALUES (?, ?, ?, ?, ?, '未完成', datetime('now','localtime'))''',
                  (data.get('module', ''), content, data.get('submitter', ''),
                   data.get('urgency', '一般'), data.get('dev_note', '')))
        conn.commit()
        new_id = c.lastrowid
        conn.close()
        return jsonify({'id': new_id, 'message': '新增成功'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/issues/<int:issue_id>', methods=['PUT'])
def api_issues_update(issue_id):
    """更新问题（切换状态或修改字段）"""
    try:
        data = request.get_json(force=True)
    except Exception:
        return jsonify({'error': '无效的JSON数据'}), 400

    try:
        conn = db_connect()
        c = conn.cursor()

        # 仅切换状态（接受 status 或 状态 字段名）
        status_val = data.get('status') or data.get('状态')
        if status_val and len(data) <= 2:
            c.execute('UPDATE 问题反馈 SET 状态=? WHERE id=?', (status_val, issue_id))
        else:
            # 更新字段
            fields = []
            values = []
            for key in ['模块', '内容', '提交人', '紧急程度', '开发者备注', '状态']:
                if key in data:
                    # 字段名是中文，直接拼接
                    fields.append(f'"{key}"=?')
                    values.append(data[key])
            if not fields:
                return jsonify({'error': '没有要更新的字段'}), 400
            values.append(issue_id)
            sql = f"UPDATE 问题反馈 SET {', '.join(fields)} WHERE id=?"
            c.execute(sql, values)

        if c.rowcount == 0:
            conn.close()
            return jsonify({'error': '问题不存在'}), 404
        conn.commit()
        conn.close()
        return jsonify({'message': '更新成功'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/issues/<int:issue_id>', methods=['DELETE'])
def api_issues_delete(issue_id):
    """删除问题反馈"""
    try:
        conn = db_connect()
        c = conn.cursor()
        c.execute('DELETE FROM 问题反馈 WHERE id=?', (issue_id,))
        if c.rowcount == 0:
            conn.close()
            return jsonify({'error': '问题不存在'}), 404
        conn.commit()
        conn.close()
        return jsonify({'message': '删除成功'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ============================================================
#  开发者日志 API — 自动扫描项目模块结构
# ============================================================

@app.route('/api/devlog', methods=['GET'])
def devlog():
    """返回项目所有模块的文件位置、行号、说明等定位信息"""
    import re
    import os as _os

    # app.py 各模块功能说明
    APP_SECTION_DESC = {
        '地区分类规则（与 skill 一致）': '国家代码→地区映射（US/UK/DE），供销量/库存查询统一使用',
        'SQLite 查询': '数据库连接管理（WAL模式）、SKU解析（去-FBA后缀）、SKU映射解析、事业部分类',
        'Function Calling 工具定义': '7个Function Calling工具：销量/库存/在途/工厂查询 + FedEx/汇率/利润计算',
        '工具执行函数': 'query_sales、query_inventory、query_in_transit、query_factory_stock 四个查询函数的具体SQL实现',
        'AI 对话 API': '/api/chat POST，多轮工具调用循环（最多5轮），支持DeepSeek/OpenAI/小米',
        '同步状态管理': '4类数据同步的后台线程管理 + /api/sync/<type> 触发 + /api/sync/status 状态轮询',
        '参数配置 API': '头程基数/资金利息的增改查，FedEx运价/其他费用的只读查询',
        '尾程测算引擎': 'FedEx运费计算：cm→inch、kg→lb、材积重(DIM 250)、AHS超大件费、燃油附加费',
        '利润测算 API': '7项固定成本 + 平台广告费的综合利润计算（采购/头程/税务/尾程/仓储/管理/客诉）',
        'SKU 搜索与详情 API': 'SKU自动补全搜索 + 尺寸重量查询（供尾程测算/利润测算调用）',
        '开发者日志 API — 自动扫描项目模块结构': '扫描所有项目文件的模块、路由、行号，生成开发者定位索引',
        '主程序': 'Waitress WSGI 服务启动（端口8081），server.log 日志配置',
    }

    result = {
        'project': '博昱科技智能助手后端',
        'base_dir': BASE_DIR,
        'files': []
    }

    # --- 1. app.py 核心模块 ---
    app_py = _os.path.join(BASE_DIR, 'app.py')
    app_lines = 0
    routes_info = []
    sections = []
    current_section = ''
    section_start = 0

    try:
        with open(app_py, 'r', encoding='utf-8') as f:
            lines_data = f.readlines()
        app_lines = len(lines_data)

        for i, line in enumerate(lines_data, 1):
            stripped = line.strip()
            # 检测注释分隔块（# =======...）
            if stripped.startswith('# ==') and stripped.endswith('=='):
                inner = stripped.strip('# =')
                # 内联标题格式: # ========== 同步状态管理 ==========
                if inner and len(inner) > 2:
                    if current_section and section_start > 0:
                        sections.append({
                            'name': current_section.strip(),
                            'start_line': section_start,
                            'end_line': i - 1
                        })
                    current_section = inner.strip()
                    section_start = i + 1
                    continue
                # 纯分隔线: 向前看下一行，如果也是分隔线则跳过（这是结束分隔线）
                next_is_sep = False
                if i < len(lines_data):
                    nxt = lines_data[i].strip()
                    if nxt.startswith('# ==') and nxt.endswith('=='):
                        next_is_sep = True
                if next_is_sep:
                    continue  # 跳过结束分隔线
                # 这是开始分隔线，下一行是标题
                # 但如果紧接着上一节的结束分隔线（当前section才1-2行），跳过
                if current_section and section_start > 0 and i - section_start <= 2:
                    continue
                if i < len(lines_data):
                    next_line = lines_data[i].strip()
                    title = next_line.lstrip('#').strip()
                    if title and not title.startswith('==='):
                        if current_section and section_start > 0:
                            sections.append({
                                'name': current_section.strip(),
                                'start_line': section_start,
                                'end_line': i - 1
                            })
                        current_section = title
                        # 跳过标题行和可能的结束分隔线
                        section_start = i + 2  # 跳过标题行(i+1)
                        # 如果下一行是结束分隔线，再跳一行
                        if i + 1 < len(lines_data):
                            nxt2 = lines_data[i + 1].strip()
                            if nxt2.startswith('# ==') and nxt2.endswith('==') and not nxt2.strip('# ='):
                                section_start = i + 3
            # 检测路由定义
            m = re.match(r"@app\.route\(['\"]([^'\"]+)['\"]", stripped)
            if m:
                routes_info.append({
                    'route': m.group(1),
                    'line': i
                })

        # 最后一个section
        if current_section:
            sections.append({
                'name': current_section.strip(),
                'start_line': section_start,
                'end_line': app_lines
            })

        # --- 扫描所有 def 函数定义及注释 ---
        func_pattern = re.compile(r'^def\s+(\w+)\s*\(')
        all_functions = []  # [{name, line, desc}]
        for i, line in enumerate(lines_data, 1):
            m = func_pattern.match(line.strip())
            if m:
                func_name = m.group(1)
                # 尝试提取函数下的docstring作为描述
                desc = ''
                if i < len(lines_data):
                    next_l = lines_data[i].strip()
                    doc_m = re.match(r'["\']{3}\s*(.+?)\s*["\']{3}', next_l)
                    if doc_m:
                        desc = doc_m.group(1)[:80]
                all_functions.append({'name': func_name, 'line': i, 'desc': desc})

        # 合并路由、区间、函数信息
        modules = []
        for s in sections:
            r_in_section = [r for r in routes_info if s['start_line'] <= r['line'] <= s['end_line']]
            f_in_section = [f for f in all_functions if s['start_line'] <= f['line'] <= s['end_line']]
            if r_in_section or f_in_section or s['name']:
                mod = {
                    'name': s['name'],
                    'start_line': s['start_line'],
                    'end_line': s['end_line'],
                    'lines': s['end_line'] - s['start_line'] + 1,
                    'routes': r_in_section,
                    'functions': f_in_section,  # 新增：该模块内所有函数定位
                    'desc': APP_SECTION_DESC.get(s['name'].strip(), ''),
                }
                modules.append(mod)
    except Exception:
        pass

    result['files'].append({
        'file': 'app.py',
        'path': 'app.py',
        'total_lines': app_lines,
        'type': 'Flask 后端主程序',
        'modules': modules
    })

    # --- 2. static/index.html 前端 ---
    html_path = _os.path.join(BASE_DIR, 'static', 'index.html')
    html_lines = 0
    html_modules = []
    try:
        with open(html_path, 'r', encoding='utf-8') as f:
            html_data = f.readlines()
        html_lines = len(html_data)

        # 通过检测 id="...View" 和对应的类来自动定位各视图模块
        view_pattern = re.compile(r'<div\s+class="([^"]*)"\s+id="(\w+View)"')
        view_starts = {}
        for i, line in enumerate(html_data, 1):
            m = view_pattern.search(line)
            if m:
                view_starts[m.group(2)] = {'class': m.group(1), 'line': i}

        # 检测 CSS 注释标记的功能块
        css_sections = []
        for i, line in enumerate(html_data, 1):
            m = re.search(r'/\*\s*(.+?)\s*\*/', line)
            if m and i < 300:
                css_sections.append({'name': m.group(1).strip(), 'line': i})

        # 检测 JS 注释标记的功能块
        js_sections = []
        in_js = False
        js_start = 0
        for i, line in enumerate(html_data, 1):
            if '<script>' in line:
                in_js = True
                js_start = i
            if in_js and re.match(r'\s*//\s*=+\s*(\S.+)', line):
                m = re.match(r'\s*//\s*=+\s*(\S.+)', line)
                js_sections.append({'name': m.group(1).strip(), 'line': i})
            if '</script>' in line and in_js:
                in_js = False
                js_sections.append({'name': 'JavaScript 结束', 'line': i})

        # 根据当前已知结构手动划分关键模块
        html_modules = [
            {'name': '全局样式 (CSS)', 'start_line': 7, 'end_line': 470, 'lines': 464,
             'desc': 'CSS变量、布局、侧边栏、消息、数据表、配置、利润、开发者日志等全部样式'},
            {'name': '侧边栏导航', 'start_line': 476, 'end_line': 570, 'lines': 95,
             'desc': '6个导航项：供应链查询/数据库/参数配置/利润测算/API配置/开发者日志'},
            {'name': 'AI对话视图', 'start_line': 580, 'end_line': 620, 'lines': 41,
             'desc': '消息列表、快捷操作卡片、输入框'},
            {'name': '数据库浏览视图', 'start_line': 621, 'end_line': 640, 'lines': 20,
             'desc': '表选择标签、搜索框、数据表格、翻页'},
            {'name': '参数配置视图 (FedEx+海智)', 'start_line': 640, 'end_line': 710, 'lines': 71,
             'desc': 'FedEx运价表、其他费用配置'},
            {'name': '利润测算视图', 'start_line': 710, 'end_line': 915, 'lines': 206,
             'desc': '尾程测算 + 利润计算的表单和结果展示'},
            {'name': '开发者日志视图', 'start_line': 917, 'end_line': 923, 'lines': 7,
             'desc': '从 /api/devlog 动态加载项目模块定位信息'},
            {'name': 'API配置弹窗', 'start_line': 925, 'end_line': 975, 'lines': 51,
             'desc': 'DeepSeek/OpenAI/小米 API 配置'},
            {'name': 'JavaScript 核心逻辑', 'start_line': 978, 'end_line': html_lines, 'lines': html_lines - 977,
             'desc': '视图切换、对话、数据库浏览、同步、配置、利润计算、开发者日志' +
                     (' (含 loadDevLog 函数)' if any('loadDevLog' in l for l in html_data) else '')},
        ]
    except Exception:
        pass

    result['files'].append({
        'file': 'index.html',
        'path': 'static/index.html',
        'total_lines': html_lines,
        'type': '前端单页面 (SPA)',
        'modules': html_modules
    })

    # --- 3. daemon.py ---
    daemon_path = _os.path.join(BASE_DIR, 'daemon.py')
    try:
        with open(daemon_path, 'r', encoding='utf-8') as f:
            d_lines_data = f.readlines()
        d_lines = len(d_lines_data)
        d_funcs = []
        df_pat = re.compile(r'^def\s+(\w+)\s*\(')
        for i, line in enumerate(d_lines_data, 1):
            dm = df_pat.match(line.strip())
            if dm:
                desc = ''
                if i < len(d_lines_data):
                    nl = d_lines_data[i].strip()
                    dm2 = re.match(r'["\']{3}\s*(.+?)\s*["\']{3}', nl)
                    if dm2:
                        desc = dm2.group(1)[:80]
                d_funcs.append({'name': dm.group(1), 'line': i, 'desc': desc})
        result['files'].append({
            'file': 'daemon.py',
            'path': 'daemon.py',
            'total_lines': d_lines,
            'type': '进程守护',
            'modules': [{
                'name': '进程守护 + 健康检查 + 自动重启',
                'start_line': 1, 'end_line': d_lines, 'lines': d_lines,
                'functions': d_funcs
            }]
        })
    except Exception:
        pass

    # --- 4. data/ 目录下的脚本 ---
    data_dir = _os.path.join(BASE_DIR, 'data')
    if _os.path.isdir(data_dir):
        data_files_info = [
            ('dingtalk_sync.py', '钉钉数据同步'),
            ('eccang_sales_sync.py', '易仓销量同步'),
            ('eccang_sync.py', '易仓库存同步'),
            ('import_data.py', '数据导入 (product_info/sku_mapping)'),
            ('import_fedex_price.py', 'FedEx运价表导入'),
        ]
        for fname, desc in data_files_info:
            fpath = _os.path.join(data_dir, fname)
            if _os.path.isfile(fpath):
                try:
                    with open(fpath, 'r', encoding='utf-8') as f:
                        fl_data = f.readlines()
                    fl = len(fl_data)
                    fl_funcs = []
                    fl_pat = re.compile(r'^def\s+(\w+)\s*\(')
                    for i, line in enumerate(fl_data, 1):
                        fm = fl_pat.match(line.strip())
                        if fm:
                            fdesc = ''
                            if i < len(fl_data):
                                nl = fl_data[i].strip()
                                fd2 = re.match(r'["\']{3}\s*(.+?)\s*["\']{3}', nl)
                                if fd2:
                                    fdesc = fd2.group(1)[:80]
                            fl_funcs.append({'name': fm.group(1), 'line': i, 'desc': fdesc})
                    result['files'].append({
                        'file': fname,
                        'path': f'data/{fname}',
                        'total_lines': fl,
                        'type': desc,
                        'modules': [{
                            'name': desc,
                            'start_line': 1, 'end_line': fl, 'lines': fl,
                            'functions': fl_funcs
                        }]
                    })
                except Exception:
                    pass

    # --- 5. 数据库文件 ---
    db_path = _os.path.join(BASE_DIR, 'data', 'supply_chain.db')
    if _os.path.isfile(db_path):
        size_kb = round(_os.path.getsize(db_path) / 1024, 1)
        tables_info = []
        try:
            conn = db_connect()
            c = conn.cursor()
            c.execute("SELECT name FROM sqlite_master WHERE type='table' AND name != 'sqlite_sequence' ORDER BY name")
            for row in c.fetchall():
                tname = row[0]
                c.execute(f'SELECT COUNT(*) FROM "{tname}"')
                cnt = c.fetchone()[0]
                tables_info.append({'name': tname, 'rows': cnt})
            conn.close()
        except Exception:
            pass
        result['files'].append({
            'file': 'supply_chain.db',
            'path': 'data/supply_chain.db',
            'type': f'SQLite数据库 ({size_kb}KB)',
            'total_lines': 0,
            'tables': tables_info
        })

    # --- 6. 配置文件 ---
    config_path = _os.path.join(BASE_DIR, 'config', 'mcporter.json')
    if _os.path.isfile(config_path):
        try:
            with open(config_path, 'r', encoding='utf-8') as f:
                cl = len(f.readlines())
            result['files'].append({
                'file': 'mcporter.json',
                'path': 'config/mcporter.json',
                'total_lines': cl,
                'type': 'MCP桥接配置',
                'modules': [{'name': 'mcporter CLI 配置', 'start_line': 1, 'end_line': cl, 'lines': cl}]
            })
        except Exception:
            pass

    # --- 汇总 ---
    total_code_lines = sum(f.get('total_lines', 0) for f in result['files'])
    result['summary'] = {
        'total_files': len(result['files']),
        'total_code_lines': total_code_lines,
        'python_files': sum(1 for f in result['files'] if f.get('file','').endswith('.py')),
        'html_files': sum(1 for f in result['files'] if f.get('file','').endswith('.html')),
        'db_files': sum(1 for f in result['files'] if f.get('file','').endswith('.db')),
    }

    return jsonify(result)


# ============================================================
#  需求预测 API
# ============================================================

import statistics
import math

# 预测参数
FORECAST_Z = 1.65          # 95% 服务水平
FORECAST_ALPHA = 0.25      # 指数平滑系数
FORECAST_WINDOW = 90       # 预测窗口天数
FORECAST_TARGET_DAYS = 100 # 目标库龄
FORECAST_MOQ = 50          # 默认最小起订量
FORECAST_SAFETY_FLOOR = 0.10  # 安全库存下限 L × 日均 × 10%

def _get_daily_sales():
    """从销量表获取每个SKU的每日销量序列"""
    conn = db_connect()
    c = conn.cursor()
    c.execute('''
        SELECT 仓库SKU, DATE(substr(付款时间,1,10)) as dt, SUM(数量) as qty
        FROM 销量
        WHERE 付款时间 IS NOT NULL AND 付款时间 != ''
          AND (仓库SKU NOT LIKE 'W%' AND 仓库SKU NOT LIKE 'w%')
          AND 仓库SKU IS NOT NULL AND 仓库SKU != ''
        GROUP BY 仓库SKU, dt
        ORDER BY 仓库SKU, dt
    ''')
    rows = c.fetchall()
    conn.close()

    sku_sales = {}
    for row in rows:
        sku_raw, dt, qty = row[0], row[1], int(row[2] or 0)
        sku = sku_raw.split('-')[0] if '-' in sku_raw else sku_raw
        if sku not in sku_sales:
            sku_sales[sku] = []
        sku_sales[sku].append({'date': dt, 'qty': qty})

    return sku_sales


def _exp_smooth(daily_qty_list, alpha=FORECAST_ALPHA, window=FORECAST_WINDOW):
    """指数平滑预测日均销量，取最近window天的数据"""
    if not daily_qty_list:
        return 0, 0
    recent = daily_qty_list[-window:] if len(daily_qty_list) > window else daily_qty_list
    vals = [d['qty'] for d in recent]
    if not vals:
        return 0, 0
    forecast = vals[0]
    for v in vals[1:]:
        forecast = alpha * v + (1 - alpha) * forecast
    sigma = statistics.stdev(vals) if len(vals) >= 2 else max(forecast * 0.3, 0.5)
    return round(forecast, 2), round(sigma, 2)


def _get_seasonal_factor(sku_sales, target_month):
    """计算季节性系数：去年同月日均 / 去年全年日均"""
    if len(sku_sales) < 365:
        return 1.0
    year_ago_data = {}
    for d in sku_sales:
        year_ago_data.setdefault(d['date'][:7], []).append(d['qty'])
    months = sorted(year_ago_data.keys())
    if len(months) < 12:
        return 1.0
    last_12 = months[-12:]
    yearly_avg = sum(sum(year_ago_data[m]) / len(year_ago_data[m]) for m in last_12) / 12
    target_key = target_month
    matching = [m for m in last_12 if m[-2:] == target_key[-2:]]
    if not matching or yearly_avg == 0:
        return 1.0
    month_avg = sum(year_ago_data[matching[0]]) / len(year_ago_data[matching[0]])
    return max(month_avg / yearly_avg, 0.3)


def _get_production_cycle(sku):
    """从工厂库存表获取SKU的生产周期中位数（实际交期 - 下单日期）"""
    conn = db_connect()
    c = conn.cursor()
    c.execute('''
        SELECT 下单日期, 实际交期 FROM 工厂库存
        WHERE SKU = ? AND 下单日期 IS NOT NULL AND 下单日期 != ''
          AND 实际交期 IS NOT NULL AND 实际交期 != ''
    ''', (sku,))
    rows = c.fetchall()
    conn.close()
    cycles = []
    for r in rows:
        try:
            d1 = datetime.strptime(r[0][:10], '%Y-%m-%d')
            d2 = datetime.strptime(r[1][:10], '%Y-%m-%d')
            diff = (d2 - d1).days
            if 5 < diff < 365:
                cycles.append(diff)
        except:
            pass
    if cycles:
        return int(statistics.median(cycles))
    return 30  # 默认30天


def _get_shipping_cycles():
    """从海运周期表获取各仓库海运周期"""
    conn = db_connect()
    c = conn.cursor()
    c.execute('SELECT 仓库类型, 海运周期 FROM 海运周期')
    rows = c.fetchall()
    conn.close()
    return {r[0]: int(r[1]) for r in rows}


def _get_inventory():
    """获取完整的库存快照"""
    conn = db_connect()
    c = conn.cursor()

    # 海外仓库存：按仓库名称汇总
    c.execute('SELECT SKU, 仓库名称, SUM(可用数量) FROM 海外仓库存 GROUP BY SKU, 仓库名称')
    overseas = {}
    for r in c.fetchall():
        sku, wh, qty = r[0], r[1], int(r[2] or 0)
        overseas.setdefault(sku, {})[wh] = qty

    # 国内仓库存 + 在产
    c.execute('SELECT SKU, SUM(国内在库), SUM(在产数量) FROM 工厂库存 GROUP BY SKU')
    domestic = {}
    in_production = {}
    for r in c.fetchall():
        sku, dom, prod = r[0], int(r[1] or 0), int(r[2] or 0)
        domestic[sku] = dom
        in_production[sku] = prod

    # 在途货物：按SKU和目的仓库汇总
    c.execute('SELECT SKU, 仓库名称, SUM(数量) FROM 在途货物 GROUP BY SKU, 仓库名称')
    in_transit = {}
    for r in c.fetchall():
        sku, wh, qty = r[0], r[1] or '未知', int(r[2] or 0)
        in_transit.setdefault(sku, {})[wh] = qty

    conn.close()
    return {
        'overseas': overseas,
        'domestic': domestic,
        'in_production': in_production,
        'in_transit': in_transit,
    }


def _map_to_market(warehouse_name):
    """将仓库名称映射到市场"""
    name = warehouse_name.upper() if warehouse_name else ''
    if '美西' in name or '美东' in name or 'US' in name or 'LA' in name or 'NY' in name or 'CG' in name or 'NJ' in name or 'CA' in name:
        return '美国地区'
    if '英国' in name or 'UK' in name or 'GB' in name or '英' in name:
        return '英国地区'
    if '德国' in name or 'DE' in name or 'EU' in name or '法国' in name or 'FR' in name or '意大利' in name or 'IT' in name or '西班牙' in name or 'ES' in name:
        return '欧洲地区'
    return '美国地区'


def _get_market_ship_cycle(market, shipping_cycles):
    """获取市场的海运周期：美国取美西美东平均，其他直接取"""
    if market == '美国地区':
        west = shipping_cycles.get('FBM美西', 35)
        east = shipping_cycles.get('FBM美东', 57)
        return (west + east) // 2
    if market == '英国地区':
        return shipping_cycles.get('FBM英国', 70)
    if market == '欧洲地区':
        return shipping_cycles.get('FBM德国', 65)
    return 50


def _map_overseas_to_shipping(warehouse_name):
    """将海外仓仓库名称映射到海运周期表的仓库类型"""
    name = warehouse_name.upper() if warehouse_name else ''
    if '美西' in name or 'USW' in name or 'LA' in name:
        return 'FBM美西'
    if '美东' in name or 'USE' in name or 'NY' in name:
        return 'FBM美东'
    if '英国' in name or 'UK' in name or 'GB' in name:
        return 'FBM英国'
    if '德国' in name or 'DE' in name or 'EU' in name:
        return 'FBM德国'
    if 'CG' in name or '中大' in name:
        return 'CG美国'
    return 'FBM美西'  # 默认


def _calc_daily_stats(sales_list):
    """从今日往前推90天分三段加权(0.5/0.3/0.2)，每段总数除以该段日历天数"""
    if not sales_list:
        return 0, 0

    today = datetime.now().date()
    today_ord = today.toordinal()
    records = []
    min_ord = None
    for d in sales_list:
        try:
            parts = (d.get('date') or '')[:10].replace('/', '-').strip().split('-')
            day = datetime(int(parts[0]), int(parts[1]), int(parts[2])).date()
        except Exception:
            continue
        day_ord = day.toordinal()
        qty = d.get('qty', 0) or 0
        records.append((day_ord, qty))
        if min_ord is None or day_ord < min_ord:
            min_ord = day_ord

    if not records or min_ord is None:
        return 0, 0

    total_span = today_ord - min_ord
    window_days = min(90, max(total_span, 30))
    start_ord = today_ord - window_days
    seg_days = window_days // 3
    seg3_end = today_ord
    seg3_start = today_ord - seg_days
    seg2_end = seg3_start
    seg2_start = today_ord - seg_days * 2
    seg1_end = seg2_start

    s1 = s2 = s3 = 0
    window_qty = []
    for day_ord, qty in records:
        if seg3_start < day_ord <= seg3_end:
            s3 += qty
        elif seg2_start < day_ord <= seg2_end:
            s2 += qty
        elif start_ord < day_ord <= seg1_end:
            s1 += qty
        if start_ord < day_ord <= today_ord:
            window_qty.append(qty)

    d3 = max(1, min(seg3_end - seg3_start, 30))
    d2 = max(1, min(seg2_end - seg2_start, 30))
    d1 = max(1, min(seg1_end - start_ord, 30))
    daily_avg = round((s3 / d3) * 0.5 + (s2 / d2) * 0.3 + (s1 / d1) * 0.2, 2)
    sigma = round(statistics.stdev(window_qty), 2) if len(window_qty) >= 2 else round(max(daily_avg * 0.3, 0.5), 2)
    return daily_avg, sigma

    today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)

    # 找到最早销售日，判断总跨度
    dates = [datetime.strptime(d['date'], '%Y-%m-%d') for d in sales_list]
    max_date = max(dates)
    min_date = min(dates)
    total_span = (today - min_date).days  # 从最早销售日到今天的日历天数

    # 实际窗口天数：不超过90天，且不超过数据跨度
    window_days = min(90, max(total_span, 30))
    start_date = today - timedelta(days=window_days)

    # 窗口内分三等份
    seg_days = window_days // 3
    seg3_end = today
    seg3_start = today - timedelta(days=seg_days)
    seg2_end = seg3_start
    seg2_start = today - timedelta(days=seg_days * 2)
    seg1_end = seg2_start
    seg1_start = start_date

    def seg_qty(start_d, end_d):
        total = 0
        for d in sales_list:
            dt = datetime.strptime(d['date'], '%Y-%m-%d')
            if start_d < dt <= end_d:
                total += d['qty']
        return total

    s3 = seg_qty(seg3_start, seg3_end)   # 最近 1/3，权重 0.5
    s2 = seg_qty(seg2_start, seg2_end)   # 中间 1/3，权重 0.3
    s1 = seg_qty(seg1_start, seg1_end)   # 最远 1/3，权重 0.2

    d3 = min((seg3_end - seg3_start).days, 30)
    d2 = min((seg2_end - seg2_start).days, 30)
    d1 = min((seg1_end - seg1_start).days, 30)

    w_avg = (s3 / d3) * 0.5 + (s2 / d2) * 0.3 + (s1 / d1) * 0.2
    daily_avg = round(w_avg, 2)

    # sigma：窗口内的日销量波动
    window_qty = []
    for d in sales_list:
        dt = datetime.strptime(d['date'], '%Y-%m-%d')
        if start_date < dt <= today:
            window_qty.append(d['qty'])
    if len(window_qty) >= 2:
        sigma = round(statistics.stdev(window_qty), 2)
    else:
        sigma = round(max(daily_avg * 0.3, 0.5), 2)

    return daily_avg, sigma


def _get_season_cached(sales_list, target_month):
    """计算季节性系数"""
    season = 1.0
    month_data = {}
    for d in sales_list:
        month_data.setdefault(d['month'], []).append(d['qty'])
    months = sorted(month_data.keys())
    if len(months) >= 12:
        last_12 = months[-12:]
        yearly_avg = sum(sum(month_data[m]) / len(month_data[m]) for m in last_12) / 12
        if yearly_avg > 0:
            matching = [m for m in last_12 if m[-2:] == target_month[-2:]]
            if matching:
                month_avg = sum(month_data[matching[0]]) / len(month_data[matching[0]])
                season = max(month_avg / yearly_avg, 0.3)
    return season


def _urgency_stars(days_remaining, L_shipping):
    """计算紧急程度星级"""
    if days_remaining <= 0:
        return 5
    ratio = days_remaining / L_shipping
    if ratio < 0.3:
        return 5
    if ratio < 0.6:
        return 4
    if ratio < 0.9:
        return 3
    if ratio < 1.2:
        return 2
    if ratio < 1.5:
        return 1
    return 0  # 不需要补货


@app.route('/api/forecast/calculate', methods=['GET'])
def forecast_calculate():
    """需求预测计算：返回工厂下单建议和海运建议"""
    try:
        conn = db_connect()
        c = conn.cursor()

        # 1. 批量获取销量（按SKU+市场分组）
        c.execute('''
            SELECT 仓库SKU,
                CASE WHEN UPPER(COALESCE(币种,"")) IN ("USD","CAD") THEN "US"
                     WHEN UPPER(COALESCE(币种,"")) = "GBP" THEN "UK"
                     ELSE "DE" END as market,
                REPLACE(substr(付款时间,1,10), '/', '-') as dt, SUM(数量) as qty
            FROM 销量
            WHERE 付款时间 IS NOT NULL AND 付款时间 != ""
              AND (仓库SKU NOT LIKE "W%" AND 仓库SKU NOT LIKE "w%")
              AND 仓库SKU IS NOT NULL AND 仓库SKU != ""
            GROUP BY 仓库SKU, market, dt
            ORDER BY 仓库SKU, market, dt
        ''')
        sales_rows = c.fetchall()
        sku_sales = {}       # 全球
        sku_market_sales = {}  # {sku: {market: [...]}}
        for row in sales_rows:
            sku_raw, mkt_code, dt, qty = row[0], row[1], row[2], int(row[3] or 0)
            sku = sku_raw.split('-')[0] if '-' in sku_raw else sku_raw
            mkt_name = '美国地区' if mkt_code == 'US' else ('英国地区' if mkt_code == 'UK' else '欧洲地区')
            rec = {'date': dt, 'qty': qty, 'month': dt[:7]}
            sku_sales.setdefault(sku, []).append(rec)
            sku_market_sales.setdefault(sku, {}).setdefault(mkt_name, []).append(rec)

        # 2. 批量获取生产周期（一次查询所有SKU）
        c.execute('''
            SELECT SKU, 下单日期, 实际交期 FROM 工厂库存
            WHERE 下单日期 IS NOT NULL AND 下单日期 != ''
              AND 实际交期 IS NOT NULL AND 实际交期 != ''
        ''')
        prod_rows = c.fetchall()
        prod_cycles_raw = {}
        for r in prod_rows:
            sku, d1, d2 = r[0], r[1][:10] if r[1] else '', r[2][:10] if r[2] else ''
            try:
                odate = datetime.strptime(d1, '%Y-%m-%d')
                ddate = datetime.strptime(d2, '%Y-%m-%d')
                diff = (ddate - odate).days
                if 5 < diff < 365:
                    prod_cycles_raw.setdefault(sku, []).append(diff)
            except:
                pass
        prod_cycles = {}
        for sku, cycles in prod_cycles_raw.items():
            prod_cycles[sku] = int(statistics.median(cycles))

        # 3. 海运周期
        c.execute('SELECT 仓库类型, 海运周期 FROM 海运周期')
        shipping_cycles = {r[0]: int(r[1]) for r in c.fetchall()}

        # 4. 海外仓库存
        c.execute('SELECT SKU, 仓库名称, SUM(可用数量) FROM 海外仓库存 GROUP BY SKU, 仓库名称')
        overseas = {}
        for r in c.fetchall():
            overseas.setdefault(r[0], {})[r[1]] = int(r[2] or 0)

        # 5. 国内仓 + 在产
        c.execute('SELECT SKU, SUM(国内在库), SUM(在产数量) FROM 工厂库存 GROUP BY SKU')
        domestic = {}; in_production = {}
        for r in c.fetchall():
            domestic[r[0]] = int(r[1] or 0); in_production[r[0]] = int(r[2] or 0)

        # 6. 在途
        c.execute('SELECT SKU, 仓库名称, SUM(数量) FROM 在途货物 GROUP BY SKU, 仓库名称')
        in_transit = {}
        for r in c.fetchall():
            in_transit.setdefault(r[0], {})[r[1] or '未知'] = int(r[2] or 0)

        conn.close()

        # 7. 计算
        today = datetime.now()
        target_month = today.strftime('%Y-%m')
        all_skus = set(sku_sales) | set(overseas) | set(domestic) | set(in_transit)
        max_ship = max(shipping_cycles.values()) if shipping_cycles else 50

        factory_rows = []
        shipping_rows = []

        for sku in all_skus:
            sales_data = sku_sales.get(sku, [])
            if not sales_data:
                continue
            daily_avg, sigma = _calc_daily_stats(sales_data)
            if daily_avg <= 0:
                continue
            season = _get_season_cached(sales_data, target_month)
            daily_adj = daily_avg * season
            prod_cycle = prod_cycles.get(sku, 30)
            domestic_stock = domestic.get(sku, 0)
            total_overseas = sum(overseas.get(sku, {}).values())
            total_transit = sum(in_transit.get(sku, {}).values())

            # ===== 工厂 =====
            L_global = prod_cycle + max_ship
            global_net = domestic_stock + total_overseas + total_transit
            ss_global = round(max(FORECAST_Z * sigma * math.sqrt(L_global), L_global * daily_adj * FORECAST_SAFETY_FLOOR), 0)
            reorder_global = round(L_global * daily_adj + ss_global, 0)
            days_left_global = round(global_net / daily_adj, 0) if daily_adj > 0 else 999
            stars_f = _urgency_stars(days_left_global, L_global)

            if stars_f > 0:
                target_stock = FORECAST_TARGET_DAYS * daily_adj
                need_qty = round(max(target_stock - global_net, FORECAST_MOQ), 0)
                factory_rows.append({'sku': sku, 'daily_avg': round(daily_adj, 1), 'sigma': round(sigma, 1),
                    'prod_cycle': prod_cycle, 'global_net': round(global_net, 0), 'days_left': days_left_global,
                    'reorder_point': reorder_global, 'need_qty': need_qty, 'stars': stars_f, 'safety_stock': ss_global})

            # ===== 海运建议：按市场聚合 + 按市场日销 =====
            market_stock = {}
            for wh_name, wh_qty in overseas.get(sku, {}).items():
                mkt = _map_to_market(wh_name)
                market_stock[mkt] = market_stock.get(mkt, 0) + wh_qty

            market_transit = {}
            for wh_name, wh_qty in in_transit.get(sku, {}).items():
                mkt = _map_to_market(wh_name)
                market_transit[mkt] = market_transit.get(mkt, 0) + wh_qty

            for mkt_name in set(list(market_stock.keys()) + list(market_transit.keys())):
                # 该SKU在该市场的日销量
                mkt_sales = sku_market_sales.get(sku, {}).get(mkt_name, [])
                mkt_daily, mkt_sigma = _calc_daily_stats(mkt_sales)
                if mkt_daily <= 0:
                    continue
                mkt_season = _get_season_cached(mkt_sales, target_month)
                mkt_daily_adj = mkt_daily * mkt_season

                mkt_stock = market_stock.get(mkt_name, 0)
                mkt_transit = market_transit.get(mkt_name, 0)
                L_ship = _get_market_ship_cycle(mkt_name, shipping_cycles)
                mkt_net = mkt_stock + mkt_transit

                ss_ship = round(max(FORECAST_Z * mkt_sigma * math.sqrt(L_ship), L_ship * mkt_daily_adj * FORECAST_SAFETY_FLOOR), 0)
                reorder_ship = round(L_ship * mkt_daily_adj + ss_ship, 0)
                days_left_ship = round(mkt_net / mkt_daily_adj, 0) if mkt_daily_adj > 0 else 999
                stars_s = _urgency_stars(days_left_ship, L_ship)
                if stars_s > 0:
                    target_mkt = FORECAST_TARGET_DAYS * mkt_daily_adj
                    need_ship = round(max(target_mkt - mkt_net, 0), 0)
                    shipping_rows.append({'sku': sku, 'market': mkt_name,
                        'L_ship': L_ship, 'daily_avg': round(mkt_daily_adj, 1), 'sigma': round(mkt_sigma, 1),
                        'mkt_stock': mkt_stock, 'mkt_transit': mkt_transit, 'mkt_net': round(mkt_net, 0),
                        'days_left': days_left_ship, 'reorder_point': reorder_ship,
                        'need_qty': need_ship, 'stars': stars_s, 'safety_stock': ss_ship})

        factory_rows.sort(key=lambda x: (-x['stars'], x['days_left']))
        shipping_rows.sort(key=lambda x: (-x['stars'], x['days_left']))

        return jsonify({
            'factory': factory_rows, 'shipping': shipping_rows,
            'factory_count': len(factory_rows), 'shipping_count': len(shipping_rows),
            'params': {'Z': FORECAST_Z, 'alpha': FORECAST_ALPHA, 'window': FORECAST_WINDOW,
                'target_days': FORECAST_TARGET_DAYS, 'MOQ': FORECAST_MOQ, 'safety_floor': FORECAST_SAFETY_FLOOR}
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/forecast/shipping/recommend', methods=['POST'])
def forecast_shipping_recommend():
    """根据输入的发运体积，推荐应优先发运的SKU及数量"""
    try:
        data = request.get_json(force=True)
        max_volume = float(data.get('max_volume', 0))
        if max_volume <= 0:
            return jsonify({'error': '请输入有效的发运体积'}), 400

        conn = db_connect()
        c = conn.cursor()

        # 销量（按市场）
        c.execute('''
            SELECT 仓库SKU,
                CASE WHEN UPPER(COALESCE(币种,"")) IN ("USD","CAD") THEN "US"
                     WHEN UPPER(COALESCE(币种,"")) = "GBP" THEN "UK"
                     ELSE "DE" END as market,
                substr(付款时间,1,10) as dt, SUM(数量) as qty
            FROM 销量 WHERE 付款时间 IS NOT NULL AND 付款时间 != ""
              AND (仓库SKU NOT LIKE "W%" AND 仓库SKU NOT LIKE "w%")
              AND 仓库SKU IS NOT NULL AND 仓库SKU != ""
            GROUP BY 仓库SKU, market, dt ORDER BY 仓库SKU, market, dt
        ''')
        sku_sales = {}
        for row in c.fetchall():
            sku_raw, mkt_code, dt, qty = row[0], row[1], row[2], int(row[3] or 0)
            sku = sku_raw.split('-')[0] if '-' in sku_raw else sku_raw
            mkt_name = '美国地区' if mkt_code == 'US' else ('英国地区' if mkt_code == 'UK' else '欧洲地区')
            sku_sales.setdefault(sku, {}).setdefault(mkt_name, []).append({'date': dt, 'qty': qty, 'month': dt[:7]})

        c.execute('SELECT 仓库类型, 海运周期 FROM 海运周期')
        shipping_cycles = {r[0]: int(r[1]) for r in c.fetchall()}

        c.execute('SELECT SKU, 仓库名称, SUM(可用数量) FROM 海外仓库存 GROUP BY SKU, 仓库名称')
        overseas = {}
        for r in c.fetchall():
            overseas.setdefault(r[0], {})[r[1]] = int(r[2] or 0)

        c.execute('SELECT SKU, 仓库名称, SUM(数量) FROM 在途货物 GROUP BY SKU, 仓库名称')
        in_transit = {}
        for r in c.fetchall():
            in_transit.setdefault(r[0], {})[r[1] or '未知'] = int(r[2] or 0)

        conn.close()

        today = datetime.now()
        target_month = today.strftime('%Y-%m')
        all_skus = set(sku_sales) | set(overseas) | set(in_transit)
        candidates = []

        for sku in all_skus:
            for mkt_name in sku_sales.get(sku, {}):
                mkt_sales = sku_sales[sku][mkt_name]
                mkt_daily, mkt_sigma = _calc_daily_stats(mkt_sales)
                if mkt_daily <= 0:
                    continue
                mkt_season = _get_season_cached(mkt_sales, target_month)
                mkt_daily_adj = mkt_daily * mkt_season

                # 按市场聚合库存
                mkt_stock = 0
                for wh_name, wh_qty in overseas.get(sku, {}).items():
                    if _map_to_market(wh_name) == mkt_name:
                        mkt_stock += wh_qty
                mkt_transit = 0
                for wh_name, wh_qty in in_transit.get(sku, {}).items():
                    if _map_to_market(wh_name) == mkt_name:
                        mkt_transit += wh_qty

                L_ship = _get_market_ship_cycle(mkt_name, shipping_cycles)
                mkt_net = mkt_stock + mkt_transit
                ss_ship = max(FORECAST_Z * mkt_sigma * math.sqrt(L_ship), L_ship * mkt_daily_adj * FORECAST_SAFETY_FLOOR)
                reorder_ship = L_ship * mkt_daily_adj + ss_ship
                days_left = mkt_net / mkt_daily_adj if mkt_daily_adj > 0 else 999
                if mkt_net <= reorder_ship:
                    target_mkt = FORECAST_TARGET_DAYS * mkt_daily_adj
                    need_qty = max(target_mkt - mkt_net, 0)
                    stars = _urgency_stars(days_left, L_ship)
                    urgency_score = stars * (L_ship / max(days_left, 1))
                    candidates.append({'sku': sku, 'market': mkt_name, 'need_qty': round(need_qty, 0),
                        'days_left': round(days_left, 0), 'stars': stars, 'urgency_score': round(urgency_score, 2)})

        candidates.sort(key=lambda x: -x['urgency_score'])
        recommended = []
        used_volume = 0
        for c in candidates:
            if used_volume >= max_volume:
                break
            can_take = min(c['need_qty'], int(max_volume - used_volume))
            if can_take > 0:
                c['recommended_qty'] = can_take
                recommended.append(c)
                used_volume += can_take

        return jsonify({'max_volume': max_volume, 'used_volume': used_volume,
            'recommended': recommended, 'count': len(recommended),
            'remaining_candidates': len(candidates) - len(recommended)})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ============================================================
#  主程序
# ============================================================

if __name__ == '__main__':
    port = int(sys.argv[1]) if len(sys.argv) > 1 else 8081

    # 日志配置（pythonw.exe无控制台时写入文件）
    import logging
    APP_DIR = os.path.dirname(os.path.abspath(__file__))
    LOG_PATH = os.path.join(APP_DIR, 'server.log')
    _handler = logging.FileHandler(LOG_PATH, encoding='utf-8')
    _handler.setFormatter(logging.Formatter('%(asctime)s [%(levelname)s] %(message)s'))
    logging.basicConfig(level=logging.INFO, handlers=[_handler])
    log = logging.getLogger('server')

    log.info("=" * 50)
    log.info("博昱科技智能助手 服务器启动")
    log.info(f"端口: {port}")
    log.info(f"数据库: {DB_PATH}")

    db_exists = os.path.exists(DB_PATH)
    if db_exists:
        try:
            conn = db_connect()
            c = conn.cursor()
            c.execute("SELECT name FROM sqlite_master WHERE type='table'")
            tables = [row[0] for row in c.fetchall()]
            for t in tables:
                c.execute(f'SELECT COUNT(*) FROM "{t}"')
                count = c.fetchone()[0]
                log.info(f"  {t}: {count} 行")
            # 确保"问题反馈"表存在
            c.execute('''CREATE TABLE IF NOT EXISTS 问题反馈 (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                模块 TEXT DEFAULT '',
                内容 TEXT NOT NULL,
                提交人 TEXT DEFAULT '',
                紧急程度 TEXT DEFAULT '一般',
                开发者备注 TEXT DEFAULT '',
                状态 TEXT DEFAULT '未完成',
                创建时间 TEXT DEFAULT (datetime('now','localtime'))
            )''')
            # 确保"海运周期"表存在（静态参考数据）
            c.execute('''CREATE TABLE IF NOT EXISTS 海运周期 (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                仓库类型 TEXT NOT NULL UNIQUE,
                海运周期 INTEGER NOT NULL
            )''')
            c.execute("SELECT COUNT(*) FROM 海运周期")
            if c.fetchone()[0] == 0:
                c.executemany(
                    "INSERT INTO 海运周期 (仓库类型, 海运周期) VALUES (?, ?)",
                    [("FBM德国", 65), ("FBM英国", 70), ("FBM美东", 57),
                     ("FBM美西", 35), ("CG美国", 50)]
                )
            # 确保德国价卡费用表存在
            c.execute('''CREATE TABLE IF NOT EXISTS 德国基础运费 (
                weight_kg INTEGER PRIMARY KEY,
                金仓DPD REAL,
                金仓DHL REAL,
                金仓GLS REAL,
                欧品居DPD REAL,
                欧品居DHL REAL,
                欧品居GLS REAL,
                易达云DPD REAL,
                易达云DHL REAL,
                易达云GLS REAL
            )''')
            c.execute('''CREATE TABLE IF NOT EXISTS 德国超规费用 (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                oversize_type TEXT,
                carrier TEXT,
                description TEXT,
                金仓 REAL,
                欧品居 REAL,
                易达云 REAL
            )''')
            c.execute('''CREATE TABLE IF NOT EXISTS 德国其他费用 (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                category TEXT,
                fee_name TEXT,
                condition_desc TEXT,
                condition_min REAL,
                condition_max REAL,
                condition_unit TEXT,
                rate REAL,
                unit TEXT,
                notes TEXT,
                sort_order INTEGER,
                仓库名称 TEXT DEFAULT '金仓'
            )''')
            log.info("德国价卡费用表已就绪")
            # 确保英国价卡费用表存在
            c.execute('''CREATE TABLE IF NOT EXISTS 英国基础运费 (
                weight_kg INTEGER PRIMARY KEY,
                易达云Royalmail REAL,
                易达云Yodel REAL,
                易达云Hermes REAL,
                易达云Parcelforce REAL,
                易达云DPD REAL,
                大健云仓Hermes REAL,
                大健云仓Whistl REAL,
                大健云仓TNT REAL
            )''')
            c.execute('''CREATE TABLE IF NOT EXISTS 英国超规费用 (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                超规类型 TEXT,
                仓库 TEXT,
                渠道 TEXT,
                描述 TEXT,
                费用 REAL
            )''')
            c.execute('''CREATE TABLE IF NOT EXISTS 英国其他费用 (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                category TEXT,
                fee_name TEXT,
                condition_desc TEXT,
                condition_min REAL,
                condition_max REAL,
                condition_unit TEXT,
                rate REAL,
                unit TEXT,
                notes TEXT,
                sort_order INTEGER,
                仓库名称 TEXT
            )''')
            log.info("英国价卡费用表已就绪")
            # 清理此前误添加的价卡表市场字段（SQLite 3.35+ 支持 DROP COLUMN）
            for ct in ['美国基础运费', '美国超规费用', '美国其他费用']:
                try:
                    c.execute(f"ALTER TABLE [{ct}] DROP COLUMN 市场")
                    log.info(f"  [{ct}] 已移除「市场」字段")
                except Exception:
                    pass
            # 自动创建性能索引（IF NOT EXISTS 保证幂等，不影响已有功能）
            log.info("初始化数据库索引...")
            indexes = [
                ("idx_销量_付款时间", "销量", "付款时间"),
                ("idx_销量_仓库SKU_付款时间", "销量", "仓库SKU, 付款时间"),
                ("idx_批次库存表_SKU", "批次库存表", "SKU"),
                ("idx_产品信息_SKU", "产品信息", "SKU"),
                ("idx_海外仓库存_SKU", "海外仓库存", "SKU"),
                ("idx_在途货物_SKU", "在途货物", "SKU"),
                ("idx_工厂库存_SKU", "工厂库存", "SKU"),
            ]
            created = 0
            for idx_name, table, col in indexes:
                try:
                    c.execute(f"CREATE INDEX IF NOT EXISTS {idx_name} ON [{table}]({col})")
                    created += 1
                except Exception:
                    pass  # 表不存在时跳过
            log.info(f"索引初始化完成: {created} 个就绪")
            conn.commit()
            conn.close()
        except Exception as e:
            log.error(f"读取数据库出错: {e}")
    else:
        log.warning(f"数据库不存在: {DB_PATH}")

    # 单次启动（不再自重启，由 daemon.py 统一管理重启逻辑，避免双重守护冲突）
    from waitress import serve
    log.info(f"waitress 启动 (端口 {port}, 线程 16)")
    serve(app, host='0.0.0.0', port=port, threads=16, channel_timeout=120)
