"""
导入德国基础运费到数据库
数据源: 德国费用.xlsx
"""
import sqlite3
import openpyxl
import os

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DB_PATH = os.path.join(BASE_DIR, 'data', 'supply_chain.db')
EXCEL_PATH = os.path.join(os.path.expanduser('~'), 'Desktop', '德国费用.xlsx')

COLS = ['金仓DPD', '金仓DHL', '金仓GLS',
        '欧品居DPD', '欧品居DHL', '欧品居GLS',
        '易达云DPD', '易达云DHL', '易达云GLS']

conn = sqlite3.connect(DB_PATH)
c = conn.cursor()

c.execute('''CREATE TABLE IF NOT EXISTS 德国基础运费 (
    weight_kg INTEGER PRIMARY KEY,
    金仓DPD REAL, 金仓DHL REAL, 金仓GLS REAL,
    欧品居DPD REAL, 欧品居DHL REAL, 欧品居GLS REAL,
    易达云DPD REAL, 易达云DHL REAL, 易达云GLS REAL
)''')

c.execute('DELETE FROM 德国基础运费')

wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
ws = wb['基础费用']

col_placeholders = ', '.join(['?'] * len(COLS))
sql = f'INSERT OR REPLACE INTO 德国基础运费 (weight_kg, {", ".join(COLS)}) VALUES (?, {col_placeholders})'

count = 0
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    weight = row[0]
    if weight is None:
        continue
    try:
        weight = int(float(weight))
    except (ValueError, TypeError):
        continue

    vals = []
    for i in range(len(COLS)):
        v = row[i + 1]
        vals.append(float(v) if v is not None else None)

    c.execute(sql, [weight] + vals)
    count += 1

conn.commit()
conn.close()
wb.close()
print(f'[OK] 德国基础运费已导入 {count} 条（1-{count}kg）')
