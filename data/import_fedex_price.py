"""
导入2026 FedEx报价到数据库
- 基础运费：重量1-150磅 × Zone-2~8
- 超尺寸费：AHS-Dimensions / AHS-Weight / Oversize-Commercial / Oversize-Residential
- 四个超尺寸费互斥，择大收取
"""
import sqlite3
import os
import openpyxl

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DB_PATH = os.path.join(BASE_DIR, 'data', 'supply_chain.db')
EXCEL_PATH = r'R:\3. 仓库管理\1.价格试算\001 原始价卡\2026海智链价卡.xlsx'

def create_tables(conn):
    cur = conn.cursor()
    
    # 参数配置表
    cur.execute('''
        CREATE TABLE IF NOT EXISTS 配置参数 (
            key TEXT PRIMARY KEY,
            value TEXT NOT NULL,
            updated_at TEXT
        )
    ''')
    
    # Fedex基础运费表：重量 × Zone + 仓库名称
    cur.execute('''
        CREATE TABLE IF NOT EXISTS 美国基础运费 (
            weight_lb INTEGER PRIMARY KEY,
            zone2 REAL,
            zone3 REAL,
            zone4 REAL,
            zone5 REAL,
            zone6 REAL,
            zone7 REAL,
            zone8 REAL,
            仓库名称 TEXT DEFAULT '海智链'
        )
    ''')
    
    # Fedex超大件费率表
    cur.execute('''
        CREATE TABLE IF NOT EXISTS 美国超规费用 (
            charge_type TEXT NOT NULL,
            zone2 REAL,
            zone3 REAL,
            zone4 REAL,
            zone5 REAL,
            zone6 REAL,
            zone7 REAL,
            zone8 REAL,
            仓库名称 TEXT DEFAULT '海智链',
            PRIMARY KEY (charge_type)
        )
    ''')
    
    conn.commit()
    print("[OK] 数据库表已创建")

def import_fedex_base_rate(conn):
    """从FedEx报价Sheet解析基础运费"""
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb['FedEx报价']
    
    cur = conn.cursor()
    cur.execute('DELETE FROM 美国基础运费')
    
    count = 0
    for row_idx in range(8, ws.max_row + 1):
        # 重量在B列(col=2)，空行跳过
        weight_cell = ws.cell(row=row_idx, column=2).value
        if weight_cell is None:
            continue
        try:
            weight = int(float(weight_cell))
        except (ValueError, TypeError):
            continue
        
        zones = {}
        # Zone-2~8 在 C~I 列 (col 3~9)
        for z_idx, z_name in enumerate(range(2, 9)):
            col_idx = 3 + z_idx  # Zone-2=col3, Zone-8=col9
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None:
                try:
                    zones[f'zone{z_name}'] = float(val)
                except (ValueError, TypeError):
                    zones[f'zone{z_name}'] = None
            else:
                zones[f'zone{z_name}'] = None
        
        cur.execute('''
            INSERT OR REPLACE INTO 美国基础运费 (weight_lb, zone2, zone3, zone4, zone5, zone6, zone7, zone8)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        ''', (weight, zones['zone2'], zones['zone3'], zones['zone4'],
              zones['zone5'], zones['zone6'], zones['zone7'], zones['zone8']))
        count += 1
    
    conn.commit()
    wb.close()
    print(f"[OK] 基础运费已导入 {count} 条（1-{count}磅）")

def import_fedex_oversize(conn):
    """从FedEx报价Sheet解析超尺寸费（四个互斥，择大收取）"""
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb['FedEx报价']
    
    cur = conn.cursor()
    cur.execute('DELETE FROM 美国超规费用')
    
    # 定义四种超尺寸费在Excel中的位置
    # 格式: (charge_type, start_row, zone_col)
    # zone价格在右侧列（比当前列+1）
    oversize_defs = [
        # AHS - Dimensions: R21-R27, col M(13)是zone名称, col N(14)是价格
        {
            'type': 'AHS_Dimensions',
            'label': 'AHS - Dimensions（额外处理费-超尺寸附加费）',
            'rows': [(21, 2), (22, 3), (23, 4), (24, 5), (25, 6), (26, 7), (27, 8)]
        },
        # AHS - Weight: R28-R34, col M(13)是zone名称, col N(14)是价格
        {
            'type': 'AHS_Weight',
            'label': 'AHS - Weight（额外处理费-超重附加费）',
            'rows': [(28, 2), (29, 3), (30, 4), (31, 5), (32, 6), (33, 7), (34, 8)]
        },
        # Oversize Charge 商业: R42-R48, col M(13)和N(14)
        {
            'type': 'Oversize_Commercial',
            'label': 'Oversize Charge（超尺寸附加费-商业地址）',
            'rows': [(42, 2), (43, 3), (44, 4), (45, 5), (46, 6), (47, 7), (48, 8)]
        },
        # Oversize Charge 住宅: R49-R55, col M(13)和N(14)
        {
            'type': 'Oversize_Residential',
            'label': 'Oversize Charge（超尺寸附加费-住宅地址）',
            'rows': [(49, 2), (50, 3), (51, 4), (52, 5), (53, 6), (54, 7), (55, 8)]
        },
    ]
    
    for od in oversize_defs:
        zones = {}
        for row_idx, zone_num in od['rows']:
            val = ws.cell(row=row_idx, column=14).value  # col N = 14
            if val is not None:
                try:
                    zones[f'zone{zone_num}'] = float(val)
                except (ValueError, TypeError):
                    zones[f'zone{zone_num}'] = None
            else:
                zones[f'zone{zone_num}'] = None
        
        cur.execute('''
            INSERT OR REPLACE INTO 美国超规费用 (charge_type, zone2, zone3, zone4, zone5, zone6, zone7, zone8)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            od['type'],
            zones.get('zone2'), zones.get('zone3'), zones.get('zone4'),
            zones.get('zone5'), zones.get('zone6'), zones.get('zone7'), zones.get('zone8')
        ))
        print(f"  [OK] {od['label']}: {zones}")
    
    conn.commit()
    wb.close()
    print(f"[OK] 超尺寸费已导入 4 类")

def init_config_params(conn):
    """初始化默认参数"""
    cur = conn.cursor()
    now = __import__('datetime').datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    
    defaults = {
        'first_leg_base': ('0', '头程基数（整数，单位：RMB）'),
        'capital_interest': ('0.00', '资金利息（百分比，如5.00表示5%）'),
    }
    
    for key, (val, desc) in defaults.items():
        cur.execute('''
            INSERT OR IGNORE INTO 配置参数 (key, value, updated_at)
            VALUES (?, ?, ?)
        ''', (key, val, now))
    
    conn.commit()
    print(f"[OK] 默认参数已初始化")

def main():
    conn = sqlite3.connect(DB_PATH)
    conn.execute("PRAGMA journal_mode=WAL")
    
    print("=" * 50)
    print("导入2026 FedEx报价")
    print("=" * 50)
    
    create_tables(conn)
    import_fedex_base_rate(conn)
    import_fedex_oversize(conn)
    init_config_params(conn)
    
    # 验证
    cur = conn.cursor()
    cur.execute('SELECT COUNT(*) FROM 美国基础运费')
    print(f"\n[验证] 基础运费: {cur.fetchone()[0]} 条")
    cur.execute('SELECT charge_type, zone2, zone3, zone4, zone5, zone6, zone7, zone8 FROM 美国超规费用')
    for row in cur.fetchall():
        print(f"  {row[0]}: Z2={row[1]} Z3={row[2]} Z4={row[3]} Z5={row[4]} Z6={row[5]} Z7={row[6]} Z8={row[7]}")
    cur.execute('SELECT * FROM 配置参数')
    for row in cur.fetchall():
        print(f"  {row[0]} = {row[1]} ({row[2]})")
    
    conn.close()
    print("\n[DONE] 导入完成!")

if __name__ == '__main__':
    main()
